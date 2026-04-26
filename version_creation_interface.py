import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox
import datetime
import webbrowser
import pandas as pd
from pathlib import Path
import re
import threading
from openpyxl import load_workbook

import version_creation_worker


class VersionCreationTab(ttk.Frame):
    def __init__(self, parent, controller):
        super().__init__(parent)
        self.parent = parent
        self.controller = controller

        if self.controller.DATABASE_PATH:
            self.DB_PATH = self.controller.DATABASE_PATH / 'main_database.xlsx'
        else:
            self.DB_PATH = None
        if self.controller.PROMPTS_PATH:
            self.PROMPT_PATH = self.controller.PROMPTS_PATH / 'version_rewriter_prompt.txt'
        else:
            self.PROMPT_PATH = None
        self.SECRETS_PATH = self.controller.SECRETS_PATH

        self.version_var = tk.StringVar(value='v2')
        self.sorting_var = tk.StringVar(value='original_asc')
        self.intro_limit_var = tk.StringVar(value='15')
        self.main_limit_var = tk.StringVar(value='25')
        self.loaded_data = None
        self.actual_sheet_name = None  # Stores actual sheet name as in file
        self.worker_thread = None
        self.stop_event = threading.Event()

        # Load databases for name and song lookup
        self.master_db = None
        self.songs_db = None
        self._load_databases()

        self.columnconfigure(0, weight=1)
        self.rowconfigure(2, weight=1)
        self.rowconfigure(4, weight=0, minsize=150)

        self._configure_styles()
        self._create_widgets()
        self._initial_log_message()

    def _configure_styles(self):
        style = ttk.Style(self)
        style.configure('Accent.TButton', font=('Helvetica', 10, 'bold'))

    def _load_databases(self):
        """Loads artist and song databases into cache"""
        try:
            if self.DB_PATH:
                self.master_db = pd.read_excel(self.DB_PATH, sheet_name='MasterDatabase_v1')
                self.songs_db = pd.read_excel(self.DB_PATH, sheet_name='Songs_Database')
                self.log_message("Databases successfully loaded into cache")
        except Exception as e:
            self.log_message(f"Warning: could not load databases: {e}")
            self.master_db = pd.DataFrame()
            self.songs_db = pd.DataFrame()

    def get_artist_name(self, eid: str) -> str:
        """Returns artist name by EID"""
        if not eid or self.master_db is None or self.master_db.empty:
            return ""

        result = self.master_db[self.master_db['EID'] == eid]
        if not result.empty:
            return result.iloc[0]['Name']
        return ""

    def get_song_name(self, ssid: str) -> str:
        """Returns song name by SSID"""
        if not ssid or self.songs_db is None or self.songs_db.empty:
            return ""

        result = self.songs_db[self.songs_db['SSID'] == ssid]
        if not result.empty:
            return result.iloc[0]['Song_Name']
        return ""

    def _create_widgets(self):
        top_controls_frame = ttk.Frame(self)
        top_controls_frame.grid(row=0, column=0, padx=10, pady=(10, 5), sticky="ew")
        top_controls_frame.columnconfigure(3, weight=1)
        ttk.Label(top_controls_frame, text="Video ID:").grid(row=0, column=0, padx=(0, 5), sticky='w')
        self.video_id_entry = ttk.Entry(top_controls_frame, width=15, font=('Helvetica', 11))
        self.video_id_entry.grid(row=0, column=1, sticky='w')
        self.video_id_entry.insert(0, "8")  # Changed to simple format for demonstration
        self.load_button = ttk.Button(top_controls_frame, text="Load", command=self._on_load_click)
        self.load_button.grid(row=0, column=2, padx=(10, 20), sticky='w')
        version_frame = ttk.Frame(top_controls_frame)
        version_frame.grid(row=0, column=3, sticky='w')
        ttk.Label(version_frame, text="Version:").pack(side="left")
        versions = ['v1', 'v2', 'v3', 'v4', 'v5', 'v6']
        for v_text in versions:
            rb = ttk.Radiobutton(version_frame, text=f"<{v_text}>", variable=self.version_var, value=v_text,
                                 takefocus=False)
            rb.pack(side="left", padx=3)
        url_frame = ttk.Frame(top_controls_frame)
        url_frame.grid(row=0, column=4, sticky='e')
        ttk.Label(url_frame, text="Donor URL:").pack(side='left', padx=(20, 5))
        self.url_label = tk.Label(url_frame, text="<load ID>", fg="grey", cursor="hand2")
        self.url_label.pack(side='left')
        self.url_label.bind("<Button-1>", self._open_url)
        sorting_frame = ttk.Labelframe(self, text="Sorting and Block Numbering", padding=(10, 5))
        sorting_frame.grid(row=1, column=0, padx=10, pady=10, sticky="ew")
        sorting_options = {"Original Ascending": "original_asc", "Original Descending": "original_desc",
                           "Random Ascending": "random_asc", "Random Descending": "random_desc"}
        for i, (text, value) in enumerate(sorting_options.items()):
            rb = ttk.Radiobutton(sorting_frame, text=text, variable=self.sorting_var, value=value, takefocus=False)
            rb.grid(row=0, column=i, padx=5, pady=5)

        ttk.Label(sorting_frame, text="Intro Limit:").grid(row=0, column=4, padx=(20, 2), pady=5)
        ttk.Entry(sorting_frame, textvariable=self.intro_limit_var, width=4).grid(row=0, column=5, padx=2, pady=5)
        ttk.Label(sorting_frame, text="Block Limit:").grid(row=0, column=6, padx=(10, 2), pady=5)
        ttk.Entry(sorting_frame, textvariable=self.main_limit_var, width=4).grid(row=0, column=7, padx=2, pady=5)
        tree_frame = ttk.Labelframe(self, text="Loaded Plan")
        tree_frame.grid(row=2, column=0, padx=10, pady=5, sticky="nsew")
        tree_frame.columnconfigure(0, weight=1)
        tree_frame.rowconfigure(0, weight=1)
        self.tree = ttk.Treeview(tree_frame, columns=('Block', 'Artist', 'EID', 'SongName', 'SSID', 'Caption'), show='headings', height=10)
        self.tree.heading('Block', text='Block');
        self.tree.heading('Artist', text='Artist');
        self.tree.heading('EID', text='EID')
        self.tree.heading('SongName', text='Song')
        self.tree.heading('SSID', text='SSID')
        self.tree.heading('Caption', text='Caption')
        # Set initial minimum widths and allow resizing
        self.tree.column('Block', width=60, minwidth=50, anchor='center', stretch=False)
        self.tree.column('Artist', width=180, minwidth=120, stretch=True)
        self.tree.column('EID', width=80, minwidth=70, anchor='center', stretch=False)
        self.tree.column('SongName', width=180, minwidth=120, stretch=True)
        self.tree.column('SSID', width=80, minwidth=70, anchor='center', stretch=False)
        self.tree.column('Caption', width=200, minwidth=150, stretch=True)
        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        self.tree.grid(row=0, column=0, sticky="nsew");
        scrollbar.grid(row=0, column=1, sticky="ns")
        action_frame = ttk.Frame(self)
        action_frame.grid(row=3, column=0, padx=10, pady=(10, 5), sticky="e")
        self.process_button = ttk.Button(action_frame, text="Process", command=self._on_process_click,
                                         style="Accent.TButton", state="disabled")
        self.process_button.pack(side="left", padx=5)
        self.stop_button = ttk.Button(action_frame, text="Stop", command=self._on_stop_click, state="disabled")
        self.stop_button.pack(side="left", padx=5)
        log_frame = ttk.Labelframe(self, text="Execution Log")
        log_frame.grid(row=4, column=0, padx=10, pady=(5, 10), sticky="nsew")
        log_frame.rowconfigure(0, weight=1);
        log_frame.columnconfigure(0, weight=1)
        self.log_widget = scrolledtext.ScrolledText(log_frame, wrap=tk.WORD, state="disabled", font=('Menlo', 10))
        self.log_widget.grid(row=0, column=0, sticky="nsew")

    def _initial_log_message(self):
        self.log_message("Version creation interface ready.")

    def log_message(self, message: str):
        if not self.winfo_exists(): return
        timestamp = datetime.datetime.now().strftime('%H:%M:%S')
        full_message = f"[{timestamp}] {message}\n"

        def _insert():
            if self.winfo_exists():
                self.log_widget.configure(state="normal")
                self.log_widget.insert(tk.END, full_message)
                self.log_widget.configure(state="disabled")
                self.log_widget.see(tk.END)

        self.after(0, _insert)

    def _open_url(self, event=None):
        url = self.url_label.cget("text")
        if url.startswith("http"):
            webbrowser.open(url)
        else:
            self.log_message("URL is not a valid link.")

    def _clear_plan(self):
        for item in self.tree.get_children(): self.tree.delete(item)
        self.loaded_data = None
        self.actual_sheet_name = None
        self.process_button.config(state="disabled")
        self.url_label.config(text="<load ID>", fg="grey")

    def _normalize_video_id(self, input_text: str) -> str:
        """
        Normalizes the entered text into Video ID format.
        Rules:
        - 4 -> VID0004s
        - 221 -> VID0221s
        - 0006S -> VID0006s
        - VID0008s -> VID0008s (no changes)
        """
        text = input_text.strip()

        # If already starts with VID (case-insensitive), just ensure it ends with 's'
        if text.upper().startswith('VID'):
            if not text.lower().endswith('s'):
                text = text + 's'
            return text.upper()[:3] + text.lower()[3:]  # VID uppercase, s lowercase

        # Remove leading 's' or 'S' if present (e.g., '0006S' -> '0006')
        if text.lower().endswith('s'):
            text = text[:-1]

        # Extract only digits
        import re
        numbers = re.findall(r'\d+', text)
        if not numbers:
            # If no digits, return as is or could raise error
            return text

        # Take only the sequence of digits
        number_str = numbers[0]

        # Pad with zeros to 4 digits
        padded_number = number_str.zfill(4)

        return f"VID{padded_number}s"

    def _find_sheet_case_insensitive(self, video_id: str):
        """
        Searches for a sheet in Excel file case-insensitively.
        Returns sheet name as it is in the file or None if not found.
        """
        try:
            import openpyxl
            workbook = openpyxl.load_workbook(self.DB_PATH, read_only=True)

            # Search for sheet ignoring case
            for sheet_name in workbook.sheetnames:
                if sheet_name.lower() == video_id.lower():
                    workbook.close()
                    return sheet_name

            workbook.close()
            return None
        except Exception as e:
            self.log_message(f"Error finding sheet: {e}")
            return None

    def _on_load_click(self):
        input_text = self.video_id_entry.get().strip()
        if not input_text:
            messagebox.showwarning("Error", "Please enter Video ID.", parent=self)
            return

        # Normalize entered ID
        video_id = self._normalize_video_id(input_text)

        # Update input field with normalized value
        self.video_id_entry.delete(0, tk.END)
        self.video_id_entry.insert(0, video_id)

        # Reload databases to pull in new artists/songs
        self._load_databases()

        self._clear_plan()
        self.log_message(f"Searching data for ID: {video_id} (entered: '{input_text}')...")

        # Search for sheet case-insensitively
        actual_sheet_name = self._find_sheet_case_insensitive(video_id)
        if not actual_sheet_name:
            self.log_message(f"ERROR: Sheet named '{video_id}' not found in database.")
            return

        # Save actual sheet name
        self.actual_sheet_name = actual_sheet_name

        try:
            df = pd.read_excel(self.DB_PATH, sheet_name=actual_sheet_name, header=None)
            self.log_message(f"File successfully read. Sheet found: '{actual_sheet_name}'. Forming plan...")

            # URL is now in column G (index 6)
            url = df.iloc[1, 6] if len(df.columns) > 6 and len(df) > 1 and pd.notna(df.iloc[1, 6]) else ""
            if url and str(url).startswith('http'):
                self.url_label.config(text=url, fg="blue")
            else:
                self.url_label.config(text="not found", fg="red")

            # Gather data from correct columns
            # A (0) - Block, B (1) - Text, C (2) - SSID, D (3) - EID, E (4) - Caption
            blocks_data = []
            for index, row in df.iloc[2:].iterrows():
                block_id = str(row.iloc[0]) if pd.notna(row.iloc[0]) and str(row.iloc[0]).strip() else ''
                if not block_id:  # Skip empty lines
                    continue

                text = str(row.iloc[1]) if pd.notna(row.iloc[1]) else ''
                ssid = str(row.iloc[2]) if len(df.columns) > 2 and pd.notna(row.iloc[2]) else ''
                eid = str(row.iloc[3]) if len(df.columns) > 3 and pd.notna(row.iloc[3]) else ''
                caption = str(row.iloc[4]) if len(df.columns) > 4 and pd.notna(row.iloc[4]) else ''

                # Special processing for Intro block
                if block_id.upper() == 'B01' or caption.upper() == 'INTRO':
                    artist_name = 'Intro'
                    song_name = ''
                else:
                    # Get artist name from database by EID
                    artist_name = self.get_artist_name(eid)

                    # Get song name from database by SSID
                    song_name = self.get_song_name(ssid)

                blocks_data.append({
                    'Block': block_id,
                    'Text': text,
                    'EID': eid,
                    'Artist': artist_name,
                    'SongName': song_name,
                    'SSID': ssid,
                    'Caption': caption
                })

            self.loaded_data = blocks_data
            if not self.loaded_data:
                self.log_message("WARNING: Plan is empty. No data for processing.");
                return

            # Output data to treeview with full content
            for row in self.loaded_data:
                # Format text for display (full version in tooltip)
                display_text = row['Text']
                if len(display_text) > 60:
                    display_text = display_text[:57] + '...'

                display_caption = row['Caption']
                if len(display_caption) > 50:
                    display_caption = display_caption[:47] + '...'

                # For B01 don't show "Unknown" in Song column
                song_display = row['SongName'] or 'Unknown'
                if row['Block'].upper() == 'B01':
                    song_display = ''

                self.tree.insert("", "end", values=(
                    row['Block'],
                    row['Artist'] or 'Unknown',
                    row['EID'],
                    song_display,
                    row['SSID'],
                    display_caption
                ))

            self.log_message(f"Plan successfully loaded. Blocks found: {len(self.loaded_data)}")
            self.process_button.config(state="normal")
        except FileNotFoundError:
            self.log_message(f"ERROR: Database file not found: {self.DB_PATH}")
        except ValueError:
            self.log_message(f"ERROR: Sheet named '{actual_sheet_name}' not found in database.")
        except Exception as e:
            self.log_message(f"ERROR during loading: {e}")

    def _get_base_video_id(self, video_id_str):
        """
        Extracts base ID from full format (e.g., VID0004s -> VID0004)
        """
        # Normalize into correct format if needed
        text = video_id_str.strip()
        if text.upper().startswith('VID'):
            text = text.upper()[:3] + text.lower()[3:]

        # Remove trailing 's' if it exists
        if text.lower().endswith('s'):
            return text[:-1]
        return text

    def _on_process_click(self):
        if not self.loaded_data:
            self.log_message("Error: no data to process. Load the plan first.");
            return

        # Check that sheet was loaded
        if not self.actual_sheet_name:
            self.log_message("Error: source sheet not found. Reload data.");
            return

        base_id = self._get_base_video_id(self.video_id_entry.get().strip())
        target_version = self.version_var.get()
        new_sheet_name = f"{base_id}{target_version.lower()}"
        try:
            workbook = load_workbook(self.DB_PATH)
            if new_sheet_name in workbook.sheetnames:
                if not messagebox.askyesno("Confirmation",
                                           f"Sheet '{new_sheet_name}' already exists.\nOverwrite it?", parent=self):
                    self.log_message("Operation cancelled by user.");
                    return
        except FileNotFoundError:
            self.log_message(f"ERROR: Cannot check sheets as file is not found: {self.DB_PATH}");
            return
        settings = {
            'video_id': self.actual_sheet_name, 'db_path': self.DB_PATH,
            'prompt_path': self.PROMPT_PATH, 'ai_manager': self.controller.ai_manager,
            'new_sheet_name': new_sheet_name, 'sorting_mode': self.sorting_var.get(),
            'sound_alerts_path': self.controller.SOUND_ALERTS_PATH,
            'target_version': target_version,  # Pass version (v1, v2, etc.)
            'intro_limit': self.intro_limit_var.get(),
            'main_limit': self.main_limit_var.get(),
        }
        self.stop_event.clear()
        self.worker_thread = threading.Thread(
            target=version_creation_worker.run_version_creation, args=(settings, self.log_message, self.stop_event, settings['sound_alerts_path']),
            daemon=True
        )
        self.process_button.config(state="disabled");
        self.load_button.config(state="disabled");
        self.stop_button.config(state="normal")
        self.worker_thread.start()
        self._check_worker_status()

    def _check_worker_status(self):
        if self.worker_thread and self.worker_thread.is_alive():
            self.after(200, self._check_worker_status)
        else:
            self._on_worker_finish()

    def _on_worker_finish(self):
        if self.winfo_exists():
            self.log_message("Process complete.")
            self.process_button.config(state="normal" if self.loaded_data else "disabled")
            self.load_button.config(state="normal")
            self.stop_button.config(state="disabled")
            self.worker_thread = None

    def _on_stop_click(self):
        if self.worker_thread and self.worker_thread.is_alive():
            self.log_message("Stop signal sent...")
            self.stop_event.set()
            self.stop_button.config(state="disabled")