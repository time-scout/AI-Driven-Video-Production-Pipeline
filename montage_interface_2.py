# --- START OF FILE montage_interface_2.py ---
#
import tkinter as tk
from tkinter import ttk, filedialog, colorchooser, scrolledtext, font, messagebox
from pathlib import Path
import pandas as pd
import json
import threading
import traceback
import montage_processor_2
from datetime import datetime
import webbrowser
import os
import re
from montage_processor_2 import stop_flag

# --- Audio imports (pydub) ---
from pydub import AudioSegment
from pydub.playback import play


def _parse_audio_filename(filename: str):
    """Extracts block number from audio filename (e.g., 1.mp3 -> 1)."""
    match = re.search(r'(\d+)', Path(filename).stem)
    return int(match.group(1)) if match else None


def format_path_for_display(full_path_str: str, widget_width_chars: int) -> str:
    """
    Formats long path for display in a widget of limited width.
    Shows the end of the path and as many previous components as will fit.
    """
    if not full_path_str:
        return ""

    # Approximate estimation of character width. May need adjustment.
    char_width_approx = 8
    widget_width_px = widget_width_chars * char_width_approx

    # Use system separator
    separator = os.sep
    parts = full_path_str.split(separator)

    # Collect path from the end while it fits
    display_path = ""
    for i in range(len(parts) - 1, -1, -1):
        current_part = parts[i]

        # If this is not the first part added, add separator
        if display_path:
            temp_path = current_part + separator + display_path
        else:
            temp_path = current_part

        # Use font.measure for accurate width measurement
        tk_font = font.nametofont("TkDefaultFont")
        if tk_font.measure(temp_path) <= widget_width_px:
            display_path = temp_path
        else:
            # If adding a new part exceeds the limit, stop
            # and add ellipsis
            display_path = "..." + separator + display_path
            break

    return display_path


class MontageTab2(ttk.Frame):
    AUDIO_EXTENSIONS = ('.mp3', '.wav', '.aac', '.m4a')

    def __init__(self, parent, controller):
        super().__init__(parent)
        self.controller = controller
        self.root = controller.root

        # --- State variables ---
        self.video_id = tk.StringVar()
        self.loaded_project_id = None  # CRITICAL: Track ID of actually loaded project
        self.project_name_var = tk.StringVar(value="Project not loaded")
        self.project_url_var = tk.StringVar(value="")
        self.montage_plan_df = None
        self.entity_id_map = {}
        self.ssid_map = {} # Map for finding song paths by SSID

        # --- Queue and threads ---
        self.montage_queue = []
        self.queue_thread = None
        self.plan_loading_thread = None
        self.is_queue_running = threading.Event()

        # --- Paths ---
        self.projects_root_path = self.controller.WORK_ROOT_PATH
        if self.controller.DATABASE_PATH:
            self.db_path = self.controller.DATABASE_PATH / 'main_database.xlsx'
        else:
            self.db_path = None
        self.archive_path = Path(self.controller.get_setting('media_archive_path', "")) if self.controller.get_setting('media_archive_path') else None

        # --- Manual intro ---
        self.manual_intro_enabled = tk.BooleanVar(value=False)
        self.manual_intro_default_folder = tk.StringVar(value="")
        self.manual_intro_list1 = []
        self.manual_intro_list2 = []
        self.manual_intro_list3 = []
        self.manual_intro_bg_songs = {1: None, 2: None, 3: None}
        self.manual_intro_b01_01_row_idx = None
        self.manual_intro_b01_02_row_idx = None
        self.manual_intro_text1_widget = None
        self.manual_intro_text2_widget = None
        self.manual_intro_text3_widget = None # Added for the third block
        self.manual_intro_radio_var = tk.StringVar(value="auto")

        # Stores SSID of the selected treeview row (for file selection buttons)
        self.active_ssid = None

        # --- Assets and settings ---
        self._load_config()

        self._build_ui()
        self._update_caption_widgets_state()
        self._update_queue_display()

    def _play_sound(self, sound_type: str):
        sound_map = {
            'success': self.controller.SOUND_ALERTS_PATH / "video_montage_complete.mp3",
            'error': self.controller.SOUND_ALERTS_PATH / "video_montage_error.mp3",
        }
        path = sound_map.get(sound_type)
        if not path or not path.exists():
            self.log(f"Montage2 WARN: Audio file for '{sound_type}' not found at path: {path}")
            return

        def task():
            try:
                audio = AudioSegment.from_file(path)
                play(audio)
            except Exception as e:
                print(f"Audio playback error: {e}")

        threading.Thread(target=task, daemon=True).start()

    def _load_config(self):
        if not self.controller.COMMON_ASSETS_PATH:
            self.log("Montage2 WARN: Resource path not set, configuration loaded by default.")
            # Set minimum necessary empty values to avoid crash
            self.assets = {}
            self.music_playlist = []
            self.formulas = {}
            self.captions = {}
            self.resources = {}
            self.sound_settings = {}
            self.general_prefixes = []
            self.live_clip_duration = tk.DoubleVar(value=3.9)
            return
        default_assets_path = self.controller.COMMON_ASSETS_PATH
        default_music = str(default_assets_path / 'music_background' / 'Jasmine Whipers - Patrick Patrikios.mp3')

        central_config_path = self.controller.CONFIG_FILE_PATH
        if central_config_path.exists():
            try:
                with open(central_config_path, 'r', encoding='utf-8') as f:
                    config = json.load(f)
            except (json.JSONDecodeError, TypeError):
                config = {}
        else:
            config = {}

        montage2_config = config.get('montage2', {})

        # 3. Extraction of nested dictionaries with backward compatibility
        assets_cfg = montage2_config.get('assets', {})
        formulas_cfg = montage2_config.get('formulas', {})
        captions_cfg = montage2_config.get('captions', {})
        sound_cfg = montage2_config.get('sound', {})
        res_cfg = montage2_config.get('resources', {})
        gen_cfg = montage2_config.get('general', {})

        self.assets = {
            'prefix': tk.StringVar(value=assets_cfg.get('prefix', montage2_config.get('prefix', ''))),
            'overlay': tk.StringVar(value=assets_cfg.get('overlay', montage2_config.get('overlay',
                                                              str(default_assets_path / 'video_overlays' / 'Dirt_vintage_film_dust_scratches.mp4')))),
            'font': tk.StringVar(
                value=assets_cfg.get('font', montage2_config.get('font', str(default_assets_path / 'fonts' / 'Egidia_Captions.ttf')))),
            'audio_folder': tk.StringVar(value=assets_cfg.get('audio_folder', montage2_config.get('audio_folder', ''))),
            'output_folder': tk.StringVar(
                value=assets_cfg.get('output_folder', montage2_config.get('output_folder', str(self.projects_root_path / "PUBLICATION_OUTPUT")))),
            'temp_folder': tk.StringVar(value=assets_cfg.get('temp_folder', montage2_config.get('temp_folder',
                                                                  str(self.projects_root_path / "PUBLICATION_OUTPUT" / "montage2_temp")))),
            # New variables for glitch transition and advertisement block
            'glitch_effect': tk.StringVar(value=assets_cfg.get('glitch_effect', montage2_config.get('glitch_effect', ''))),
            'ad_file': tk.StringVar(value=assets_cfg.get('ad_file', montage2_config.get('ad_file', ''))),
            'ad_placement': tk.StringVar(value=assets_cfg.get('ad_placement', montage2_config.get('ad_placement', '0'))),
            # Variables for manual intro (not saved to config, current session only)
            'manual_intro_live1': tk.StringVar(value=''),
            'manual_intro_live2': tk.StringVar(value=''),
            'montage_mode': tk.StringVar(value='zigzag')
        }
        self.music_playlist = gen_cfg.get('music_playlist', montage2_config.get('music_playlist', [default_music]))

        self.formulas = {
            "Intro": tk.StringVar(value=formulas_cfg.get("Intro", montage2_config.get("Intro", "L+C+L"))),
            "Main Part": tk.StringVar(value=formulas_cfg.get("Main Part", montage2_config.get("Main Part", "C+L+C")))
        }
        self.live_clip_duration = tk.DoubleVar(value=gen_cfg.get('live_clip_duration', montage2_config.get('live_clip_duration', 3.9)))

        self.captions = {
            'font_size': tk.IntVar(value=captions_cfg.get('font_size', montage2_config.get('font_size', 150))),
            'duration': tk.DoubleVar(value=captions_cfg.get('duration', montage2_config.get('duration', 5.8))),
            'mode': tk.StringVar(value=captions_cfg.get('mode', captions_cfg.get('caption_mode', montage2_config.get('caption_mode', 'fixed')))),
            'offset_end': tk.DoubleVar(value=captions_cfg.get('offset_end', montage2_config.get('offset_end', 5.0))),
            'font_color': tk.StringVar(value=captions_cfg.get('font_color', montage2_config.get('font_color', "#560078"))),
            'outline_width': tk.IntVar(value=captions_cfg.get('outline_width', montage2_config.get('outline_width', 9))),
            'position': tk.StringVar(value=captions_cfg.get('position', montage2_config.get('position', "Middle"))),
            'outline_color': tk.StringVar(value=captions_cfg.get('outline_color', montage2_config.get('outline_color', "#f98500")))
        }
        self.resources = {
            'normalization_threads': tk.IntVar(value=res_cfg.get('normalization_threads', montage2_config.get('normalization_threads', 8)))
        }

        self.sound_settings = {
            'target_source_lufs': tk.StringVar(value=sound_cfg.get('target_source_lufs', montage2_config.get('target_source_lufs', '-16'))),
            'bg_music_db': tk.StringVar(value=sound_cfg.get('bg_music_db', montage2_config.get('bg_music_db', '-25'))),
            'combo_vs_live_db': tk.StringVar(value=sound_cfg.get('combo_vs_live_db', montage2_config.get('combo_vs_live_db', '0'))),
            'final_limiter_tp': tk.StringVar(value=sound_cfg.get('final_limiter_tp', montage2_config.get('final_limiter_tp', '-1.5')))
        }

        self.use_individual_var = tk.BooleanVar(value=gen_cfg.get('use_individual_prefixes', montage2_config.get('use_individual_prefixes', False)))
        self.general_prefixes = gen_cfg.get('general_prefixes', montage2_config.get('general_prefixes', []))

        # Variables for fade effects
        self.use_fade_transition = tk.BooleanVar(value=gen_cfg.get('use_fade_transition', montage2_config.get('use_fade_transition', False)))
        self.fade_duration = tk.DoubleVar(value=gen_cfg.get('fade_duration', montage2_config.get('fade_duration', 0.2)))

        # Loading manual intro settings
        manual_intro_cfg = montage2_config.get('manual_intro', {})
        self.manual_intro_enabled.set(manual_intro_cfg.get('enabled', False))
        self.manual_intro_default_folder.set(manual_intro_cfg.get('default_browse_folder', ''))

        # 4. Ensure "instant" saving (Trace)
        # Assets, Formulas, Titles, Resources, Audio
        for d in [self.assets, self.formulas, self.captions, self.resources, self.sound_settings]:
            for var in d.values():
                var.trace_add('write', self._save_config)

        # Single variables
        for var in [self.live_clip_duration, self.use_individual_var, self.use_fade_transition, self.fade_duration,
                    self.manual_intro_enabled, self.manual_intro_default_folder]:
            var.trace_add('write', self._save_config)

        self.log("Montage2 configuration successfully loaded from central config under key 'montage2'.")
        # Save existing listeners for font (add after main trace_add)
        self.assets['font'].trace_add('write', self._update_caption_widgets_state)

    def _save_config(self, *args):
        """Saving Montage2 configuration. *args for compatibility with trace_add."""
        try:
            central_config_path = self.controller.CONFIG_FILE_PATH
            if central_config_path.exists():
                with open(central_config_path, 'r', encoding='utf-8') as f:
                    config = json.load(f)
            else:
                config = {}

            # 2. Unification of structure in _save_config - strict nesting
            montage2_data = {
                'assets': {key: var.get() for key, var in self.assets.items()},
                'formulas': {key: var.get() for key, var in self.formulas.items()},
                'captions': {key: var.get() for key, var in self.captions.items()},
                'sound': {key: var.get() for key, var in self.sound_settings.items()},
                'resources': {key: var.get() for key, var in self.resources.items()},
                'general': {
                    'music_playlist': self.music_playlist,
                    'general_prefixes': self.general_prefixes,
                    'use_individual_prefixes': self.use_individual_var.get(),
                    'live_clip_duration': self.live_clip_duration.get(),
                    'use_fade_transition': self.use_fade_transition.get(),
                    'fade_duration': self.fade_duration.get()
                },
                'manual_intro': {
                    'enabled': self.manual_intro_enabled.get(),
                    'default_browse_folder': self.manual_intro_default_folder.get()
                }
            }

            config['montage2'] = montage2_data

            central_config_path.parent.mkdir(parents=True, exist_ok=True)
            with open(central_config_path, 'w', encoding='utf-8') as f:
                json.dump(config, f, indent=4, ensure_ascii=False)

            self.log("Montage2 configuration saved in central config under key 'montage2'.")
        except Exception as e:
            self.log(f"Error saving Montage2 configuration: {e}")

    def _build_ui(self):
        main_paned_window = ttk.PanedWindow(self, orient='vertical')
        main_paned_window.pack(fill='both', expand=True, padx=10, pady=10)
        top_frame = ttk.Frame(main_paned_window)
        main_paned_window.add(top_frame, weight=3)
        bottom_frame = ttk.Frame(main_paned_window)
        main_paned_window.add(bottom_frame, weight=1)
        top_left_frame = ttk.Frame(top_frame)
        top_left_frame.pack(side='left', fill='both', expand=True, padx=(0, 5))
        top_right_frame = ttk.Frame(top_frame)
        top_right_frame.pack(side='left', fill='both', expand=False)

        id_frame = ttk.LabelFrame(top_left_frame, text="1. Project Loading")
        id_frame.pack(padx=0, pady=(0, 10), fill='x')
        id_frame.columnconfigure(1, weight=1)
        input_row = ttk.Frame(id_frame)
        input_row.grid(row=0, column=0, columnspan=2, sticky='ew', padx=5, pady=5)
        ttk.Label(input_row, text="Video ID:").pack(side='left')
        ttk.Entry(input_row, textvariable=self.video_id, width=10).pack(side='left', padx=5)
        self.load_plan_button = ttk.Button(input_row, text="Load Plan", command=self._load_project_plan)
        self.load_plan_button.pack(side='left', padx=5)

        info_row = ttk.Frame(id_frame)
        info_row.grid(row=1, column=0, columnspan=2, sticky='ew', padx=5, pady=(0, 5))
        ttk.Label(info_row, text="Project:").pack(side='left')
        ttk.Label(info_row, textvariable=self.project_name_var).pack(side='left', padx=5)
        url_label = tk.Label(info_row, textvariable=self.project_url_var, fg="blue", cursor="hand2")
        url_label.pack(side='left', padx=5)
        url_label.bind("<Button-1>", lambda e: webbrowser.open(
            self.project_url_var.get()) if "http" in self.project_url_var.get() else None)

        plan_frame = ttk.LabelFrame(top_left_frame, text="2. Montage Plan")
        plan_frame.pack(padx=0, pady=0, fill='both', expand=True)
        self.tree = ttk.Treeview(plan_frame, columns=('Block', 'Caption', 'SSID', 's'), show='headings', height=10)
        self.tree.heading('Block', text='Block')
        self.tree.heading('Caption', text='Caption')
        self.tree.heading('SSID', text='SSID')
        self.tree.heading('s', text='s')
        self.tree.column('Block', width=80, anchor='center')
        self.tree.column('Caption', width=300)
        self.tree.column('SSID', width=120, anchor='center')
        self.tree.column('s', width=30, anchor='center')
        scrollbar = ttk.Scrollbar(plan_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side='right', fill='y')
        self.tree.pack(side='left', fill='both', expand=True)
        self.tree.bind("<Double-1>", self._on_treeview_double_click)
        self.tree.bind('<<TreeviewSelect>>', self._on_tree_select)

        notebook = ttk.Notebook(top_right_frame)
        notebook.pack(fill='both', expand=True)
        self._build_prefixes_tab(notebook)
        self._build_manual_intro_tab(notebook)
        self._build_assets_tab(notebook)
        self._build_captions_tab(notebook)
        self._build_resources_tab(notebook)
        self._build_export_tab(notebook)
        self._build_queue_tab(notebook)

        control_frame = ttk.Frame(bottom_frame)
        control_frame.pack(fill='x', pady=5)
        style = ttk.Style()
        style.configure('Accent.TButton', font=('Helvetica', 12, 'bold'))
        self.add_to_queue_btn = ttk.Button(control_frame, text="Add to Queue", command=self._add_to_queue)
        self.add_to_queue_btn.pack(side='left', padx=5, ipady=5)
        self.start_queue_btn = ttk.Button(control_frame, text="Start Queue", command=self._start_queue,
                                          style='Accent.TButton')
        self.start_queue_btn.pack(side='left', padx=5, ipady=5)
        self.stop_queue_btn = ttk.Button(control_frame, text="Stop Queue", command=self._stop_queue,
                                         state='disabled')
        self.stop_queue_btn.pack(side='left', padx=5, ipady=5)
        log_frame = ttk.LabelFrame(bottom_frame, text="Execution Log")
        log_frame.pack(fill='both', expand=True, pady=(5, 0))
        self.log_text = scrolledtext.ScrolledText(log_frame, state='disabled', height=8, wrap='word')
        self.log_text.pack(fill='both', expand=True, padx=5, pady=5)

    def _build_assets_tab(self, notebook):
        tab = ttk.Frame(notebook)
        notebook.add(tab, text='Assets')
        self._create_asset_selector(tab, "Disclaimer:", 'prefix', [("Video", "*.mp4 *.mov *.jpg *.png")])
        self._create_asset_selector(tab, "Overlay:", 'overlay', [("Video/PNG", "*.mp4 *.mov *.png")])
        self._create_folder_selector(tab, "Audio Folder:", 'audio_folder', width=35)

        # File selector for glitch transition
        self._create_asset_selector(tab, "Glitch Transition:", 'glitch_effect', [("Video", "*.mp4 *.mov")])

        # Montage mode
        mode_frame = ttk.Frame(tab)
        mode_frame.pack(padx=10, pady=5, fill='x')
        ttk.Label(mode_frame, text="Montage Mode:", width=15).pack(side='left')
        ttk.Radiobutton(mode_frame, text="Zigzag Montage", variable=self.assets['montage_mode'], value='zigzag').pack(side='left', padx=5)
        ttk.Radiobutton(mode_frame, text="Seamless Combo", variable=self.assets['montage_mode'], value='seamless').pack(side='left', padx=5)
        ttk.Radiobutton(mode_frame, text="Third Mode", variable=self.assets['montage_mode'], value='third_mode').pack(side='left', padx=5)

        # Create frame for ad row
        ad_frame = ttk.Frame(tab)
        ad_frame.pack(padx=10, pady=5, fill='x')

        ttk.Label(ad_frame, text="Ad, after:", width=15).pack(side='left')

        # Position input window (5 chars)
        ad_pos_entry = ttk.Entry(ad_frame, textvariable=self.assets['ad_placement'], width=5)
        ad_pos_entry.pack(side='left', padx=5)

        # File path window
        ad_path_entry = ttk.Entry(ad_frame, width=35, state='readonly')
        ad_path_entry.pack(side='left', fill='x', expand=True, padx=5)

        # Path window text update function (similar to other selectors)
        def update_ad_path(*args):
            display_text = format_path_for_display(self.assets['ad_file'].get(), 35)
            ad_path_entry.config(state='normal')
            ad_path_entry.delete(0, tk.END)
            ad_path_entry.insert(0, display_text)
            ad_path_entry.config(state='readonly')

        self.assets['ad_file'].trace_add('write', update_ad_path)
        update_ad_path()

        # Select and clear buttons
        ttk.Button(ad_frame, text="...", command=lambda: self._select_asset('ad_file', [("Video", "*.mp4 *.mov")]), width=3).pack(side='left')
        ttk.Button(ad_frame, text="X", command=lambda: self.assets['ad_file'].set(""), width=3).pack(side='left', padx=5)

        # Fade transition controls
        fade_frame = ttk.Frame(tab)
        fade_frame.pack(padx=10, pady=5, fill='x')

        # Checkbutton for fade to black
        ttk.Checkbutton(fade_frame, text="Fade to black", variable=self.use_fade_transition, command=self._save_config).pack(side='left')

        # Spinbox for fade duration
        ttk.Label(fade_frame, text="Duration (sec):").pack(side='left', padx=(10, 2))
        ttk.Spinbox(fade_frame, from_=0.1, to=2.0, increment=0.1, textvariable=self.fade_duration, width=5, command=self._save_config).pack(side='left')

        music_frame = ttk.LabelFrame(tab, text="Background Music Playlist")
        music_frame.pack(padx=10, pady=5, fill='both', expand=True)
        listbox_frame = ttk.Frame(music_frame)
        listbox_frame.pack(fill='both', expand=True, padx=5, pady=5)
        self.music_listbox = tk.Listbox(listbox_frame, selectmode=tk.SINGLE, height=5)
        self.music_listbox.pack(side='left', fill='both', expand=True)
        scrollbar = ttk.Scrollbar(listbox_frame, orient='vertical', command=self.music_listbox.yview)
        scrollbar.pack(side='right', fill='y')
        self.music_listbox.config(yscrollcommand=scrollbar.set)

        for track in self.music_playlist:
            self.music_listbox.insert(tk.END, Path(track).name)

        btn_frame = ttk.Frame(music_frame)
        btn_frame.pack(pady=5)
        ttk.Button(btn_frame, text="Add", command=self._add_music_track).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="Delete", command=self._remove_music_track).pack(side='left', padx=5)

    def _build_captions_tab(self, notebook):
        tab = ttk.Frame(notebook)
        notebook.add(tab, text='Captions')
        
        # --- Block Settings ---
        blocks_frame = ttk.LabelFrame(tab, text="Block Settings")
        blocks_frame.pack(padx=10, pady=5, fill='x')
        ttk.Label(blocks_frame, text="Live Clip Duration (s):").pack(side='left', padx=5, pady=5)
        ttk.Spinbox(blocks_frame, from_=1.0, to=10.0, increment=0.1, textvariable=self.live_clip_duration, width=6).pack(side='left', padx=5)

        formula_frame = ttk.LabelFrame(tab, text="Assembly Formulas")
        formula_frame.pack(padx=10, pady=5, fill='x')
        for name, var in self.formulas.items():
            row = ttk.Frame(formula_frame)
            row.pack(fill='x', padx=5, pady=2)
            ttk.Label(row, text=f"{name}:", width=15).pack(side='left')
            ttk.Entry(row, textvariable=var).pack(side='left', fill='x', expand=True)
        
        legend_label = ttk.Label(formula_frame, text="💡 L - Live (native audio), C - Combo (voice + background)", font=('', 9, 'italic'))
        legend_label.pack(pady=2)

        self.captions_frame = ttk.LabelFrame(tab, text="Caption Settings")
        self.captions_frame.pack(padx=10, pady=5, fill='x')
        self._create_asset_selector(self.captions_frame, "Font:", 'font', [("Fonts", "*.ttf *.otf")])

        # Caption modes
        mode_frame = ttk.Frame(self.captions_frame)
        mode_frame.pack(fill='x', padx=5, pady=5)
        ttk.Label(mode_frame, text="Mode:").pack(side='left', padx=5)
        ttk.Radiobutton(mode_frame, text="duration", variable=self.captions['mode'], value='fixed').pack(side='left', padx=5)
        ttk.Radiobutton(mode_frame, text="before end", variable=self.captions['mode'], value='offset_end').pack(side='left', padx=5)

        self.caption_widgets_container = ttk.Frame(self.captions_frame)
        self.caption_widgets_container.pack(fill='x', expand=True, padx=5, pady=5)
        col1 = ttk.Frame(self.caption_widgets_container)
        col1.pack(side='left', fill='x', expand=True, padx=5)
        col2 = ttk.Frame(self.caption_widgets_container)
        col2.pack(side='left', fill='x', expand=True, padx=5)
        
        ttk.Label(col1, text="Size:").pack(anchor='w')
        ttk.Spinbox(col1, from_=10, to=500, textvariable=self.captions['font_size'], width=6).pack(fill='x', pady=2)
        
        ttk.Label(col1, text="from start / before end:").pack(anchor='w', pady=(5, 0))
        # Use one variable for display, but separate in logic
        ttk.Spinbox(col1, from_=1.0, to=300.0, increment=0.1, textvariable=self.captions['duration'], width=6).pack(
            fill='x', pady=2)
        
        ttk.Label(col1, text="Font Color:").pack(anchor='w', pady=(5, 0))
        self.font_color_btn = ttk.Button(col1, text=self.captions['font_color'].get(),
                                         command=lambda: self._choose_color(self.captions['font_color'],
                                                                            self.font_color_btn))
        self.font_color_btn.pack(fill='x', pady=2)
        
        outline_frame = ttk.Frame(col2)
        outline_frame.pack(fill='x')
        ttk.Label(outline_frame, text="Outline:").pack(anchor='w')
        self.outline_label = ttk.Label(outline_frame, text=f"{self.captions['outline_width'].get()} px", width=5)
        self.outline_label.pack(side='right')
        ttk.Scale(outline_frame, from_=0, to=10, orient='horizontal', variable=self.captions['outline_width'],
                  command=self._update_outline_label).pack(fill='x', pady=2, side='right', expand=True)
        
        ttk.Label(col2, text="Position:").pack(anchor='w', pady=(5, 0))
        ttk.Combobox(col2, textvariable=self.captions['position'], values=['Middle', 'Top', 'Bottom', 'Top Left', 'Bottom Left'],
                     state='readonly', width=10).pack(fill='x', pady=2)
        
        ttk.Label(col2, text="Outline Color:").pack(anchor='w', pady=(5, 0))
        self.outline_color_btn = ttk.Button(col2, text=self.captions['outline_color'].get(),
                                            command=lambda: self._choose_color(self.captions['outline_color'],
                                                                               self.outline_color_btn))
        self.outline_color_btn.pack(fill='x', pady=2)

    def _build_export_tab(self, notebook):
        tab = ttk.Frame(notebook)
        notebook.add(tab, text='Export')
        self._create_folder_selector(tab, "Output folder for finished videos:", 'output_folder', self._save_config, 40)
        self._create_folder_selector(tab, "Folder for temporary files:", 'temp_folder', self._save_config, 40)

    def _build_resources_tab(self, notebook):
        tab = ttk.Frame(notebook)
        notebook.add(tab, text='Resources')
        norm_frame = ttk.LabelFrame(tab, text="Resources for clip normalization")
        norm_frame.pack(padx=10, pady=10, fill='x')
        ttk.Label(norm_frame, text="Thread count:").pack(side='left', padx=5)
        ttk.Spinbox(norm_frame, from_=1, to=10, textvariable=self.resources['normalization_threads'], width=5).pack(
            side='left', padx=5)

        sound_frame = ttk.LabelFrame(tab, text="Audio Mixing Settings")
        sound_frame.pack(padx=10, pady=10, fill='x')
        settings_frame = ttk.Frame(sound_frame)
        settings_frame.pack(fill='x', padx=5, pady=5)

        row1 = ttk.Frame(settings_frame)
        row1.pack(fill='x', pady=2)
        ttk.Label(row1, text="Target source volume (LUFS):").pack(side='left')
        # <<<--- START OF FIX 2: AUTOSAVE ---
        lufs_entry = ttk.Entry(row1, textvariable=self.sound_settings['target_source_lufs'], width=10)
        lufs_entry.pack(side='left', padx=5)
        self.sound_settings['target_source_lufs'].trace_add('write', lambda *args: self._save_config())

        row2 = ttk.Frame(settings_frame)
        row2.pack(fill='x', pady=2)
        ttk.Label(row2, text="Background volume (relative to voice, dB):").pack(side='left')
        db_entry = ttk.Entry(row2, textvariable=self.sound_settings['bg_music_db'], width=10)
        db_entry.pack(side='left', padx=5)
        self.sound_settings['bg_music_db'].trace_add('write', lambda *args: self._save_config())

        row_cvl = ttk.Frame(settings_frame)
        row_cvl.pack(fill='x', pady=2)
        ttk.Label(row_cvl, text="Combo vs Live (dB):").pack(side='left')
        cvl_entry = ttk.Entry(row_cvl, textvariable=self.sound_settings['combo_vs_live_db'], width=10)
        cvl_entry.pack(side='left', padx=5)
        self.sound_settings['combo_vs_live_db'].trace_add('write', lambda *args: self._save_config())
        ttk.Label(row_cvl, text="💡 -3 = voice quieter than music, +3 = voice louder", font=('', 9, 'italic')).pack(side='left', padx=5)

        row3 = ttk.Frame(settings_frame)
        row3.pack(fill='x', pady=2)
        ttk.Label(row3, text="Final limiter (True Peak, dB):").pack(side='left')
        tp_entry = ttk.Entry(row3, textvariable=self.sound_settings['final_limiter_tp'], width=10)
        tp_entry.pack(side='left', padx=5)
        self.sound_settings['final_limiter_tp'].trace_add('write', lambda *args: self._save_config())
        # --- END OF FIX 2 ---

        ttk.Button(sound_frame, text="Test Audio Mixing", command=self._run_audio_test).pack(pady=5)

    def _build_queue_tab(self, notebook):
        tab = ttk.Frame(notebook)
        notebook.add(tab, text='Queue')
        self.queue_tree = ttk.Treeview(tab, columns=('ID', 'Project'), show='headings', height=5)
        self.queue_tree.heading('ID', text='Video ID')
        self.queue_tree.heading('Project', text='Project')
        self.queue_tree.column('ID', width=100)
        self.queue_tree.column('Project', width=250)
        self.queue_tree.pack(fill='both', expand=True, padx=10, pady=10)
        btn_frame = ttk.Frame(tab)
        btn_frame.pack(pady=5)
        ttk.Button(btn_frame, text="Remove Selected", command=self._remove_from_queue).pack(side='left')
        ttk.Button(btn_frame, text="Clear Queue", command=self._clear_queue).pack(side='left', padx=10)

    def _create_asset_selector(self, parent, label_text, asset_key, file_types, width=35):
        frame = ttk.Frame(parent)
        frame.pack(padx=10, pady=5, fill='x')
        ttk.Label(frame, text=label_text, width=15).pack(side='left')
        entry = ttk.Entry(frame, width=width, state='readonly')
        entry.pack(side='left', fill='x', expand=True, padx=5)

        def update_entry_text(*args):
            display_text = format_path_for_display(self.assets[asset_key].get(), width)
            entry.config(state='normal')
            entry.delete(0, tk.END)
            entry.insert(0, display_text)
            entry.config(state='readonly')

        self.assets[asset_key].trace_add('write', update_entry_text)
        update_entry_text()
        ttk.Button(frame, text="...", command=lambda: self._select_asset(asset_key, file_types), width=3).pack(
            side='left')
        ttk.Button(frame, text="X", command=lambda: self.assets[asset_key].set(""), width=3).pack(side='left', padx=5)

    def _create_folder_selector(self, parent, label_text, asset_key, save_callback=None, width=35):
        frame = ttk.Frame(parent)
        frame.pack(padx=10, pady=5, fill='x')
        ttk.Label(frame, text=label_text, width=22).pack(side='left')
        entry = ttk.Entry(frame, width=width, state='readonly')
        entry.pack(side='left', fill='x', expand=True, padx=5)

        def update_entry_text(*args):
            display_text = format_path_for_display(self.assets[asset_key].get(), width)
            entry.config(state='normal')
            entry.delete(0, tk.END)
            entry.insert(0, display_text)
            entry.config(state='readonly')

        self.assets[asset_key].trace_add('write', update_entry_text)
        update_entry_text()

        def select_and_save():
            path = filedialog.askdirectory(title=f"Select folder for '{label_text}'")
            if path: self.assets[asset_key].set(path); self.log(f"Montage2: Folder '{label_text}' set: {path}")
            if save_callback: save_callback()

        ttk.Button(frame, text="...", command=select_and_save, width=3).pack(side='left')

    def _build_entity_map(self):
        self.log("Montage2: Scanning archive to create Performer and Song map...")
        if not self.archive_path or not self.archive_path.is_dir():
            self.log(f"Montage2: ERROR: Archive not found or path not set: {self.archive_path}")
            return False
        
        self.entity_id_map.clear()
        self.ssid_map.clear()
        
        found_performers = 0
        found_songs = 0
        
        # Scan the entire archive tree
        for item in self.archive_path.rglob('*'):
            if not item.is_dir():
                continue
                
            # Search for Performers by EID (e.g., EID0082) in folder name
            eid_match = re.search(r'(EID\d+)', item.name)
            if eid_match:
                eid_code = eid_match.group(1).strip()
                # Save path to performer folder (if prefixes are there)
                if (item / 'interview_fragments').is_dir() or (item / 'entity_meta.json').exists():
                    self.entity_id_map[eid_code] = item
                    found_performers += 1
            
            # Search for songs by SSID (e.g., SSID0572) in folder name
            ssid_match = re.search(r'(SSID\d+)', item.name)
            if ssid_match:
                ssid_code = ssid_match.group(1).strip()
                self.ssid_map[ssid_code] = item
                found_songs += 1
                        
        self.log(f"Montage2: ✅ Map created. Performers: {found_performers}, Songs: {found_songs}.")
        return True

    def _enrich_plan(self, plan_df):
        if plan_df is None or plan_df.empty: return None
        if not self.entity_id_map and not self._build_entity_map(): return None
        return plan_df.copy()

    def _add_to_queue(self):
        # CRITICAL FIX 2: Check if button is locked during loading
        if self.plan_loading_thread and self.plan_loading_thread.is_alive():
            self.log("Montage2: ERROR: Plan is loading. Wait for completion.")
            return

        if self.montage_plan_df is None or self.montage_plan_df.empty:
            self.log("Montage2: ERROR: Plan not loaded. Nothing to add to queue.")
            return

        # Check video_id synchronization with loaded plan
        current_vid = self.video_id.get().strip()
        if not current_vid:
            self.log("Montage2: ERROR: Video ID not specified.")
            return

        # CRITICAL CHECK: Compare entered ID with actually loaded one
        if self.loaded_project_id is None:
            self.log("Montage2: ERROR: No confirmation of loaded project. Reload plan.")
            return

        if current_vid.lower() != self.loaded_project_id.lower():
            self.log(f"Montage2: ERROR: ID mismatch! Entered: '{current_vid}', Loaded: '{self.loaded_project_id}'. Reload plan.")
            return

        # Check project state
        project_name = self.project_name_var.get()
        if "Project not loaded" in project_name or "Error" in project_name or "Loading" in project_name:
            self.log("Montage2: ERROR: Project not loaded correctly. Cannot add to queue.")
            return

        self.log("Montage2: Preparing task for queue...")

        # Update archive map before adding to ensure everything is visible
        if not self._build_entity_map():
            self.log("Montage2: ERROR: Failed to scan archive.")
            return

        current_settings = {
            'video_id': self.video_id.get(),
            'project_name': self.project_name_var.get(),
            'montage_plan': self.montage_plan_df.copy(),
            'ssid_map': {k: str(v.absolute()) if isinstance(v, Path) else str(Path(v).absolute()) for k, v in self.ssid_map.items()},
            'entity_map': {k: str(v.absolute()) if isinstance(v, Path) else str(Path(v).absolute()) for k, v in self.entity_id_map.items()},
            # Assets with absolute paths
            'assets': {
                k: str(Path(v.get()).absolute()) if v.get() and ('/' in v.get() or '\\' in v.get()) else v.get()
                for k, v in self.assets.items()
            },
            'formulas': {k: v.get() for k, v in self.formulas.items()},
            'captions': {k: v.get() for k, v in self.captions.items()},
            'resources': {k: v.get() for k, v in self.resources.items()},
            'music_playlist': [str(Path(p).absolute()) for p in self.music_playlist],
            'general_prefixes': [str(Path(p).absolute()) for p in self.general_prefixes],
            'sound_settings': {k: v.get() for k, v in self.sound_settings.items()},
            'use_individual_prefixes': self.use_individual_var.get(),
            'live_clip_duration': self.live_clip_duration.get(),
            'live_duration': self.live_clip_duration.get(),
            'media_archive_path': str(self.archive_path.absolute()) if self.archive_path else '',
            # Transfer new processor settings with absolute paths
            'glitch_effect': str(Path(self.assets['glitch_effect'].get()).absolute()) if self.assets['glitch_effect'].get() else '',
            'use_fade_transition': self.use_fade_transition.get(),
            'fade_duration': self.fade_duration.get(),
            # Transfer advertisement block settings with absolute paths
            'ad_file': str(Path(self.assets['ad_file'].get()).absolute()) if self.assets['ad_file'].get() else '',
            'ad_placement': self.assets['ad_placement'].get(),
            'db_path': str(self.db_path.absolute()) if self.db_path else '',
            # --- MANUAL INTRO ---
            'manual_intro_enabled': self.manual_intro_radio_var.get() in ("manual", "third_mode", "auto"),
            'manual_intro_mode': self.manual_intro_radio_var.get(),
            'manual_intro_list1': [str(Path(p).absolute()) for p in self.manual_intro_list1],
            'manual_intro_list2': [str(Path(p).absolute()) for p in self.manual_intro_list2],
            'manual_intro_list3': [str(Path(p).absolute()) for p in self.manual_intro_list3],
            'manual_intro_bg_songs': self.manual_intro_bg_songs.copy(),
            'montage_mode': self.assets['montage_mode'].get()
        }
        self.montage_queue.append(current_settings)
        self.log(
            f"Montage2: Project '{current_settings['project_name']}' added to queue. Total in queue: {len(self.montage_queue)}.")
        self._update_queue_display()

    def _remove_from_queue(self):
        selected_items = self.queue_tree.selection()
        if not selected_items: return
        indices_to_remove = sorted([self.queue_tree.index(i) for i in selected_items], reverse=True)
        for index in indices_to_remove:
            removed = self.montage_queue.pop(index)
            self.log(f"Montage2: Project '{removed['project_name']}' removed from queue.")
        self._update_queue_display()

    def _clear_queue(self):
        self.montage_queue.clear()
        self.log("Montage2: Queue cleared.")
        self._update_queue_display()

    def _update_queue_display(self):
        for item in self.queue_tree.get_children(): self.queue_tree.delete(item)
        for i, task in enumerate(self.montage_queue):
            self.queue_tree.insert("", "end", iid=i, values=(task['video_id'], task['project_name']))

    def _start_queue(self):
        if not self.montage_queue: self.log("Montage2: Queue is empty."); return
        if self.is_queue_running.is_set(): self.log("Montage2: Queue already running."); return
        self.is_queue_running.set()
        self.start_queue_btn.config(state='disabled')
        self.stop_queue_btn.config(state='normal')
        self.queue_thread = threading.Thread(target=self._queue_worker, daemon=True)
        self.queue_thread.start()

    def _stop_queue(self):
        if not self.is_queue_running.is_set(): return
        self.log("Montage2: Stop queue signal sent...")
        self.is_queue_running.clear()
        montage_processor_2.stop_execution()
        self.stop_queue_btn.config(state='disabled')

    def _queue_worker(self):
        montage_had_error = False
        while self.is_queue_running.is_set() and self.montage_queue:
            task = self.montage_queue.pop(0)
            self.root.after(0, self._update_queue_display)
            self.log("\n" + "=" * 50 + f"\nMontage2: Starting montage of project: {task['project_name']}\n" + "=" * 50)

            settings_for_processor = {
                'video_id': task['video_id'],
                'output_path': task['assets']['output_folder'],
                'temp_path': task['assets']['temp_folder'],
                'disclaimer_path': task['assets']['prefix'],
                'overlay_path': task['assets']['overlay'],
                'caption_font': task['assets']['font'],
                'audio_folder': task['assets']['audio_folder'],
                'ssid_map': task.get('ssid_map', {}),
                'entity_map': task.get('entity_map', {}),
                'media_archive_path': task.get('media_archive_path'),
                'formula_intro': task['formulas']['Intro'],
                'formula_main': task['formulas']['Main Part'],
                'music_playlist': task['music_playlist'],
                'general_prefixes': task['general_prefixes'],
                'sound_settings': task['sound_settings'],
                'use_individual_prefixes': task['use_individual_prefixes'],
                'font_size': task['captions']['font_size'],
                'font_color': task['captions']['font_color'],
                'outline_color': task['captions']['outline_color'],
                'outline_width': task['captions']['outline_width'],
                'caption_pos': task['captions']['position'],
                # Task B: Fix caption data transfer logic
                'caption_duration': task['captions']['duration'],  # This is the number from the single Spinbox
                'caption_mode': task['captions']['mode'],  # Radio button
                'caption_offset_end': task['captions'].get('offset_end', 5.0),
                'normalization_threads': task['resources']['normalization_threads'],
                'live_clip_duration': task.get('live_clip_duration', 3.9),
                'db_path': task['db_path'],
                # --- START OF ADDED BLOCK ---
                'glitch_effect': task['assets']['glitch_effect'],
                'use_fade_transition': task['use_fade_transition'],
                'fade_duration': task['fade_duration'],
                # Advertisement block settings
                'ad_file': task['ad_file'],
                'ad_placement': task['ad_placement'],
                # --- MANUAL INTRO ---
                'manual_intro_enabled': task.get('manual_intro_enabled', False),
                'manual_intro_list1': task.get('manual_intro_list1', []),
                'manual_intro_list2': task.get('manual_intro_list2', []),
                'manual_intro_list3': task.get('manual_intro_list3', []),
                'manual_intro_bg_songs': task.get('manual_intro_bg_songs', {}),
                'montage_mode': task.get('montage_mode', 'zigzag')
                # --- END OF ADDED BLOCK ---
            }

            process_finished_event = threading.Event()
            success_status = [None]

            def run_and_get_status():
                try:
                    montage_processor_2.run_montage_process(task['montage_plan'], settings_for_processor, self)
                    if montage_processor_2.stop_flag.is_set():
                        success_status[0] = 'stopped'
                    else:
                        success_status[0] = True
                except Exception as e:
                    self.log(f"--- TRACEBACK --- \n{traceback.format_exc()}--- END TRACEBACK ---")
                    success_status[0] = False
                finally:
                    process_finished_event.set()

            montage_thread = threading.Thread(target=run_and_get_status, daemon=True)
            montage_thread.start()
            process_finished_event.wait()

            if success_status[0] is True:
                self.log(f"Montage2: Project {task['project_name']} successfully completed.")
                self._play_sound('success')
            else:
                self.log(f"Montage2: Project {task['project_name']} finished with an error or was stopped.")
                self._play_sound('error')
                montage_had_error = True
                break

        self.root.after(0, self._on_queue_finish, montage_had_error)

    def _on_queue_finish(self, was_error):
        if was_error:
            self.log("Montage2: Queue operation interrupted due to an error.")
        else:
            self.log("Montage2: Queue operation completed.")
        self.is_queue_running.clear()
        self.start_queue_btn.config(state='normal')
        self.stop_queue_btn.config(state='disabled')
        self.add_to_queue_btn.config(state='normal')

    def _load_project_plan_worker(self):
        # --- START OF ADDED BLOCK ---
        # CRITICAL FIX 1: Instant clearing of old data at start of loading
        def clear_all_project_data():
            self.montage_plan_df = None  # Forced clearing of old plan
            self.loaded_project_id = None  # Reset loaded project ID
            self.general_prefixes.clear()
            self.general_prefixes_listbox.delete(0, tk.END)
            # Clearing Treeview
            self.tree.delete(*self.tree.get_children())
            # Resetting project variables
            self.project_name_var.set("Loading...")
            self.project_url_var.set("")
            self.assets['prefix'].set("")
            self.assets['audio_folder'].set("")
            # Blocking "Add to Queue" button
            self.add_to_queue_btn.config(state='disabled')
            self.log("Montage2: All old project data cleared. Loading new one...")
        self.root.after(0, clear_all_project_data)
        # --- END OF ADDED BLOCK ---
        self.root.after(0, self.load_plan_button.config, {'state': 'disabled'})
        vid_input = self.video_id.get().strip()
        try:
            if not self.db_path or not self.db_path.exists():
                self.log(f"Montage2: ERROR: DB not found or path not set: {self.db_path}")
                return
            xls = pd.ExcelFile(self.db_path)
            sheet_map = {name.lower(): name for name in xls.sheet_names}
            actual_sheet_name = sheet_map.get(vid_input.lower())
            if not actual_sheet_name: raise ValueError(f"Sheet '{vid_input}' not found")
            vid = actual_sheet_name

            # Project data preparation (atomic collection)
            project_data = {
                'name': "Project not found",
                'url': "",
                'audio_folder': "",
                'plan_list': None
            }

            found_projects = list(self.projects_root_path.glob(f"{vid}*"))
            if len(found_projects) == 1:
                project_path = found_projects[0]
                self.log(f"Montage2: ✅ Found project: {project_path.name}")
                project_data['name'] = project_path.name
                audio_path = project_path / 'voice'
                if audio_path.is_dir():
                    project_data['audio_folder'] = str(audio_path)
            else:
                self.log("Montage2: Project not found. Audio folder not found.")

            df = pd.read_excel(xls, sheet_name=actual_sheet_name, header=None)
            url = df.iloc[1, 4] if len(df.columns) > 4 and len(df) > 1 else None
            if pd.notna(url): project_data['url'] = url

            # SSID - index 2 (column C), EID - index 3 (column D), Caption - index 10 (column K)
            plan_data = df.iloc[2:, [10, 2, 3]]
            plan_data.columns = ['Caption', 'SSID', 'EID_Ref']
            plan_data.dropna(subset=['SSID'], inplace=True)
            plan_data = plan_data[plan_data['SSID'].astype(str).str.strip() != '']

            # Updating archive map to count files in slices
            self._build_entity_map()

            final_plan_list = [{'Block': 'B01', 'Caption': '', 'SSID': 'N/A (Intro)', 's': ''}]
            for i, row in enumerate(plan_data.itertuples(index=False), start=2):
                ssid = str(row.SSID).strip()
                # Counting files in slices for this SSID
                slices_count = ''
                if ssid in self.ssid_map:
                    slices_path = self.ssid_map[ssid] / 'slices'
                    if slices_path.is_dir():
                        # Counting video files
                        video_files = [f for f in slices_path.iterdir()
                                      if f.is_file() and f.suffix.lower() in ('.mp4', '.mov', '.mkv', '.avi')]
                        if video_files:
                            slices_count = len(video_files)

                final_plan_list.append(
                    {'Block': f'B{i:02d}', 'Caption': str(row.Caption).strip(), 'SSID': ssid, 's': slices_count})

            project_data['plan_list'] = final_plan_list

            # Atomic UI update in one block
            def update_ui_atomically():
                self.project_name_var.set(project_data['name'])
                self.project_url_var.set(project_data['url'])
                if project_data['audio_folder']:
                    self.assets['audio_folder'].set(project_data['audio_folder'])
                self._populate_treeview(final_plan_list)
                # Loading texts for manual intro
                self._load_manual_intro_texts(df)
                
                # --- AUTO-LOAD FROM SEED ---
                self._load_manual_intro_from_seed(actual_sheet_name)
                
                # Unlocking "Add to Queue" button only after successful load
                self.add_to_queue_btn.config(state='normal')
            self.root.after(0, update_ui_atomically)

            self.montage_plan_df = pd.DataFrame(final_plan_list)
            self.loaded_project_id = actual_sheet_name  # Record successfully loaded ID
            self.log(f"Montage2: ✅ Plan successfully built. Total blocks: {len(self.montage_plan_df)}")
        except Exception as e:
            self.log(f"Montage2: ❌ ERROR loading plan: {e}")
            self.root.after(0, lambda: self.project_name_var.set("Loading error"))
            # Button remains blocked on error
        finally:
            if self.root.winfo_exists(): self.root.after(0, self.load_plan_button.config, {'state': 'normal'})
            self.plan_loading_thread = None

    def _load_manual_intro_from_seed(self, vid):
        """Loads manual intro settings from seed.json."""
        pid = vid.replace('VID', 'PID')
        seed_path = self.controller.DATABASE_PATH / 'seed' / f"{pid}_seed.json"
        
        if not seed_path.exists(): return

        try:
            with open(seed_path, 'r', encoding='utf-8') as f:
                seed_data = json.load(f)
            
            settings = seed_data.get("montage_manual_intro_settings")
            if not settings: return

            m_mode = settings.get("manual_mode", "auto")
            if m_mode is True: m_mode = "manual"
            if m_mode is False: m_mode = "auto"
            self.manual_intro_radio_var.set(m_mode)
            self.manual_intro_enabled.set(m_mode in ("manual", "third_mode"))

            for i in range(1, 4):
                block_data = settings.get(f"block{i}", {})
                clips = block_data.get("clips", [])
                bg_ssid = block_data.get("bg_ssid")
                text = block_data.get("text", "")

                # Filling lists
                target_list = getattr(self, f"manual_intro_list{i}")
                target_list.clear()
                target_list.extend(clips)

                # Filling Listbox
                lb = self.manual_blocks[i]['listbox']
                lb.delete(0, tk.END)
                for c in clips: lb.insert(tk.END, Path(c).name)

                # Text
                txt_widget = getattr(self, f"manual_intro_text{i}_widget")
                txt_widget.delete('1.0', tk.END)
                txt_widget.insert('1.0', text)

                # Background song
                if bg_ssid:
                    self.manual_intro_bg_songs[i] = bg_ssid
                    lb.insert(tk.END, f"[Live+Combo] {bg_ssid}")

            self.log(f"Montage2: Manual intro settings loaded from seed.")
        except Exception as e:
            self.log(f"Montage2: Auto-load from seed error: {e}")

    def _load_project_plan(self):
        if self.plan_loading_thread and self.plan_loading_thread.is_alive(): self.log(
            "Montage2: Loading already in progress..."); return
        if not self.video_id.get().strip(): self.log("Montage2: ERROR: Enter Video ID."); return
        self.plan_loading_thread = threading.Thread(target=self._load_project_plan_worker, daemon=True)
        self.plan_loading_thread.start()

    def _populate_treeview(self, plan_list):
        for item in self.tree.get_children(): self.tree.delete(item)
        for item in plan_list: self.tree.insert("", "end", values=(item['Block'], item['Caption'], item['SSID'], item.get('s', '')))

    def log(self, message):
        if hasattr(self, 'root') and self.root.winfo_exists():
            timestamp = datetime.now().strftime("[%H:%M:%S] ")
            self.root.after(0, self._log, timestamp + str(message))

    def _log(self, message):
        if not self.root.winfo_exists(): return
        self.log_text.config(state='normal')
        self.log_text.insert('end', message + '\n')
        self.log_text.see('end')
        self.log_text.config(state='disabled')
        self.root.update_idletasks()

    def _select_asset(self, asset_key, file_types):
        initial_dir = self.controller.COMMON_ASSETS_PATH
        file_path = filedialog.askopenfilename(title=f"Select file for '{asset_key}'", filetypes=file_types,
                                               initialdir=initial_dir)
        if file_path: self.assets[asset_key].set(file_path); self.log(
            f"Montage2: Asset '{asset_key}' set: {Path(file_path).name}")

    def _update_caption_widgets_state(self, *args):
        state = 'normal' if self.assets['font'].get() else 'disabled'
        for child in self.caption_widgets_container.winfo_children():
            for sub_child in child.winfo_children():
                try:
                    sub_child.config(state=state)
                except tk.TclError:
                    pass

    def _update_outline_label(self, value):
        self.outline_label.config(text=f"{int(float(value))} px")

    def _choose_color(self, color_var, button):
        color_code = colorchooser.askcolor(title="Select Color", initialcolor=color_var.get())
        if color_code and color_code[1]: color_var.set(color_code[1]); button.config(text=color_code[1])

    def _on_treeview_double_click(self, event):
        region = self.tree.identify("region", event.x, event.y)
        column_id_str = self.tree.identify_column(event.x)
        if region != "cell" or column_id_str != "#2": return
        selected_item_id = self.tree.focus()
        if not selected_item_id: return
        x, y, width, height = self.tree.bbox(selected_item_id, column_id_str)
        entry_var = tk.StringVar(value=self.tree.item(selected_item_id, "values")[1])
        entry = ttk.Entry(self.tree, textvariable=entry_var)
        entry.place(x=x, y=y, width=width, height=height)
        entry.focus_set()

        def on_focus_out(event):
            new_value = entry_var.get()
            self.tree.set(selected_item_id, column='Caption', value=new_value)
            if self.montage_plan_df is not None: self.montage_plan_df.loc[
                self.tree.index(selected_item_id), 'Caption'] = new_value
            self.log(f"Montage2: Caption changed: '{new_value}'")
            entry.destroy()

        entry.bind("<FocusOut>", on_focus_out)
        entry.bind("<Return>", on_focus_out)
        entry.bind("<Escape>", lambda e: entry.destroy())

    def _on_tree_select(self, event):
        """Saves SSID of the selected treeview row"""
        selected_items = self.tree.selection()
        if selected_items:
            values = self.tree.item(selected_items[0], "values")
            if values and len(values) >= 3:
                ssid = str(values[2]).strip()
                self.active_ssid = ssid
                if ssid and 'SSID' in ssid:
                    self.log(f"Montage2: Context set to SSID: {ssid} (Slices folder ready)")
            else:
                self.active_ssid = None
        else:
            self.active_ssid = None

    def _build_prefixes_tab(self, notebook):
        tab = ttk.Frame(notebook)
        notebook.add(tab, text='Prefixes')

        scan_btn = ttk.Button(tab, text="Scan archive for prefixes", command=self._scan_for_prefixes)
        scan_btn.pack(pady=5, padx=10)

        tree_frame = ttk.LabelFrame(tab, text="Available Prefixes")
        tree_frame.pack(fill='both', expand=True, padx=10, pady=5)

        ttk.Label(tree_frame, text="💡 Double click to find file").pack(anchor='w', padx=5, pady=(0, 2))

        self.prefix_treeview = ttk.Treeview(tree_frame,
                                            columns=('Performer', 'Type', 'Filename', 'Path'),
                                            displaycolumns=('Performer', 'Type', 'Filename'),
                                            show='headings', height=8)
        self.prefix_treeview.heading('Performer', text='Performer')
        self.prefix_treeview.heading('Type', text='Type')
        self.prefix_treeview.heading('Filename', text='Filename')
        self.prefix_treeview.column('Performer', width=150, anchor='w')
        self.prefix_treeview.column('Type', width=120, anchor='w')
        self.prefix_treeview.column('Filename', width=300, anchor='w')

        tree_scroll_y = ttk.Scrollbar(tree_frame, orient='vertical', command=self.prefix_treeview.yview)
        tree_scroll_x = ttk.Scrollbar(tree_frame, orient='horizontal', command=self.prefix_treeview.xview)
        self.prefix_treeview.configure(yscrollcommand=tree_scroll_y.set, xscrollcommand=tree_scroll_x.set)

        tree_scroll_y.pack(side='right', fill='y')
        tree_scroll_x.pack(side='bottom', fill='x')
        self.prefix_treeview.pack(side='left', fill='both', expand=True)
        self.prefix_treeview.bind('<Double-1>', self._preview_prefix)

        preview_btn = ttk.Button(tab, text="Show file in Finder", command=self._preview_prefix)
        preview_btn.pack(pady=5, padx=10)

        individual_frame = ttk.Frame(tab)
        individual_frame.pack(fill='x', padx=10, pady=(0, 5))
        self.use_individual_cb = ttk.Checkbutton(individual_frame, text="Use individual prefixes for blocks", variable=self.use_individual_var, command=self._save_config)
        self.use_individual_cb.pack(side='left', padx=5)

        list_frame = ttk.LabelFrame(tab, text="General Prefixes (for Intro)")
        list_frame.pack(fill='x', expand=False, padx=10, pady=5)

        listbox_container = ttk.Frame(list_frame)
        listbox_container.pack(fill='x', expand=True, padx=5, pady=5)
        self.general_prefixes_listbox = tk.Listbox(listbox_container, selectmode=tk.SINGLE, height=3)
        list_scroll = ttk.Scrollbar(listbox_container, orient='vertical', command=self.general_prefixes_listbox.yview)
        self.general_prefixes_listbox.config(yscrollcommand=list_scroll.set)
        list_scroll.pack(side='right', fill='y')
        self.general_prefixes_listbox.pack(side='left', fill='both', expand=True)

        for prefix in self.general_prefixes:
            self.general_prefixes_listbox.insert(tk.END, Path(prefix).name)

        btn_frame = ttk.Frame(list_frame)
        btn_frame.pack(pady=(0, 5), padx=5, fill='x')
        ttk.Button(btn_frame, text="Add", command=self._add_to_general_prefixes).pack(side='left')
        ttk.Button(btn_frame, text="Delete", command=self._remove_from_general_prefixes).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="↑", command=self._move_up, width=3).pack(side='left')
        ttk.Button(btn_frame, text="↓", command=self._move_down, width=3).pack(side='left', padx=(5, 0))

    def _build_manual_intro_tab(self, notebook):
        tab = ttk.Frame(notebook)
        notebook.add(tab, text='Manual Intro')

        # --- Top Control Panel ---
        ctrl_panel = ttk.Frame(tab)
        ctrl_panel.pack(padx=10, pady=5, fill='x')
        
        # Left part with radio buttons (two rows)
        radio_container = ttk.Frame(ctrl_panel)
        radio_container.pack(side='left')
        
        row1 = ttk.Frame(radio_container)
        row1.pack(side='top', anchor='w')
        
        row2 = ttk.Frame(radio_container)
        row2.pack(side='top', anchor='w')
        
        self.manual_intro_radio_var = tk.StringVar(value="auto")
        ttk.Radiobutton(row1, text="Zigzag Intro", variable=self.manual_intro_radio_var,
                       value="auto", command=self._on_manual_intro_radio_change).pack(side='left', padx=5)
        ttk.Radiobutton(row1, text="Enable manual intro", variable=self.manual_intro_radio_var,
                       value="manual", command=self._on_manual_intro_radio_change).pack(side='left', padx=5)
        
        ttk.Radiobutton(row2, text="Third intro mode",
                       variable=self.manual_intro_radio_var, value='third_mode',
                       command=self._on_manual_intro_radio_change).pack(side='left', padx=5)
        
        ttk.Separator(ctrl_panel, orient='vertical').pack(side='left', fill='y', padx=10)
        
        ttk.Button(ctrl_panel, text="Save block order", command=self._save_manual_intro_to_seed).pack(side='left', padx=5)
        ttk.Button(ctrl_panel, text="Clear fields", command=self._clear_manual_intro_fields).pack(side='left', padx=5)

        # Container for 3 blocks (vertical)
        blocks_container = ttk.Frame(tab)
        blocks_container.pack(fill='both', expand=True, padx=5, pady=5)

        # Create 3 blocks
        self.manual_blocks = {}
        for i in range(1, 4):
            block_id = f"B01-0{i}"
            frame = ttk.LabelFrame(blocks_container, text="")
            frame.pack(side='top', fill='x', expand=False, padx=2, pady=5)
            
            # Block header
            header = ttk.Frame(frame)
            header.pack(fill='x', padx=2, pady=2)
            ttk.Label(header, text=f"Block {block_id}", font=('Helvetica', 10, 'bold')).pack(side='left')
            
            # Save text button
            save_cmd = getattr(self, f"_save_text{i}_to_excel")
            ttk.Button(header, text="save text", command=save_cmd).pack(side='left', padx=10)
            
            # List management
            ttk.Button(header, text="+", width=2, command=lambda idx=i: self._manual_intro_add(idx)).pack(side='left', padx=1)
            ttk.Button(header, text="-", width=2, command=lambda idx=i: self._manual_intro_remove(idx)).pack(side='left', padx=1)
            ttk.Button(header, text="↑", width=2, command=lambda idx=i: self._manual_intro_move(idx, -1)).pack(side='left', padx=1)
            ttk.Button(header, text="↓", width=2, command=lambda idx=i: self._manual_intro_move(idx, 1)).pack(side='left', padx=1)
            ttk.Button(header, text="Set combo", command=lambda idx=i: self._manual_intro_set_bg(idx)).pack(side='left', padx=10)

            # Text field
            txt = scrolledtext.ScrolledText(frame, height=4, wrap='word', font=('Helvetica', 10))
            txt.pack(fill='x', padx=2, pady=2)
            setattr(self, f"manual_intro_text{i}_widget", txt)

            # Clips list
            lb_frame = ttk.Frame(frame)
            lb_frame.pack(fill='x', expand=False, padx=2, pady=2)
            lb = tk.Listbox(lb_frame, height=5, font=('Helvetica', 9), selectmode=tk.SINGLE)
            lb.pack(side='left', fill='x', expand=True)
            sb = ttk.Scrollbar(lb_frame, orient='vertical', command=lb.yview)
            sb.pack(side='right', fill='y')
            lb.config(yscrollcommand=sb.set)
            
            self.manual_blocks[i] = {'listbox': lb, 'data_list': getattr(self, f"manual_intro_list{i}")}

    def _save_manual_intro_to_seed(self):
        """Saves manual intro settings to project seed.json."""
        if not self.loaded_project_id:
            messagebox.showerror("Error", "Project not loaded.")
            return

        # VID0043v4 -> PID0043v4
        vid = self.loaded_project_id
        pid = vid.replace('VID', 'PID')
        
        seed_path = self.controller.DATABASE_PATH / 'seed' / f"{pid}_seed.json"
        
        if not seed_path.exists():
            messagebox.showerror("Error", f"Seed file not found: {seed_path}")
            return

        try:
            with open(seed_path, 'r', encoding='utf-8') as f:
                seed_data = json.load(f)
            
            settings = {
                "manual_mode": self.manual_intro_radio_var.get()
            }
            
            for i in range(1, 4):
                block_key = f"block{i}"
                lb = self.manual_blocks[i]['listbox']
                txt_widget = getattr(self, f"manual_intro_text{i}_widget")
                
                settings[block_key] = {
                    "clips": getattr(self, f"manual_intro_list{i}"),
                    "bg_ssid": self.manual_intro_bg_songs.get(i),
                    "text": txt_widget.get('1.0', tk.END).strip()
                }
            
            seed_data["montage_manual_intro_settings"] = settings
            
            with open(seed_path, 'w', encoding='utf-8') as f:
                json.dump(seed_data, f, indent=4, ensure_ascii=False)
            
            self.log(f"Montage2: Manual intro settings saved to {seed_path.name}")
            messagebox.showinfo("Success", "Intro settings saved to seed.json")
            
        except Exception as e:
            self.log(f"Montage2: ERROR saving to seed: {e}")
            messagebox.showerror("Error", f"Failed to save settings: {e}")

    def _clear_manual_intro_fields(self):
        """Fully clears manual intro fields."""
        if not messagebox.askyesno("Confirmation", "Are you sure you want to completely clear all manual intro fields?"):
            return
            
        for i in range(1, 4):
            # Clear arrays
            getattr(self, f"manual_intro_list{i}").clear()
            # Clear Listbox
            self.manual_blocks[i]['listbox'].delete(0, tk.END)
            # Clear Text
            getattr(self, f"manual_intro_text{i}_widget").delete('1.0', tk.END)
            # Clear songs
            self.manual_intro_bg_songs[i] = None
            
        self.log("Montage2: All manual intro fields cleared.")

    def _manual_intro_add(self, block_idx):
        """Adds files to block list"""
        lb = self.manual_blocks[block_idx]['listbox']
        
        # 1. Check for [Live+Combo] presence before file selection
        if "[Live+Combo]" in lb.get(0, tk.END):
            messagebox.showwarning("Error", "Cannot add clips after Live+Combo line. Remove it or move the clip higher.")
            return

        slices_folder = self._get_selected_ssid_slices_folder()
        initial_dir = str(slices_folder) if slices_folder else self.manual_intro_default_folder.get()
        
        files = filedialog.askopenfilenames(title="Add clips", initialdir=initial_dir)
        if not files: return
        
        target_list = getattr(self, f"manual_intro_list{block_idx}")
        
        for f in files:
            if f not in target_list:
                target_list.append(f)
                # 2. Using [Solo-Live] format
                lb.insert(tk.END, f"[Solo-Live] {Path(f).name}")
        self.log(f"Montage2: Added {len(files)} files to block B01-0{block_idx}.")

    def _manual_intro_remove(self, block_idx):
        """Removes selected element"""
        lb = self.manual_blocks[block_idx]['listbox']
        sel = lb.curselection()
        if not sel: return
        
        idx = sel[0]
        val = lb.get(idx)
        
        if "[Live+Combo]" in val:
            self.manual_intro_bg_songs[block_idx] = None
        else:
            target_list = getattr(self, f"manual_intro_list{block_idx}")
            # Try to find path by filename in the list
            # Since the format is now [Solo-Live] filename, we need to extract the name
            filename = val.replace("[Solo-Live] ", "").strip()
            for i, p in enumerate(target_list):
                if Path(p).name == filename:
                    target_list.pop(i)
                    break
        
        lb.delete(idx)

    def _manual_intro_move(self, block_idx, direction):
        """Moves element up/down"""
        lb = self.manual_blocks[block_idx]['listbox']
        sel = lb.curselection()
        if not sel: return
        
        idx = sel[0]
        new_idx = idx + direction
        if not (0 <= new_idx < lb.size()): return
        
        val = lb.get(idx)
        if "[Live+Combo]" in val or "[Live+Combo]" in lb.get(new_idx):
            # Don't move background song, it's always at the bottom
            return

        target_list = getattr(self, f"manual_intro_list{block_idx}")
        target_list[idx], target_list[new_idx] = target_list[new_idx], target_list[idx]
        
        lb.delete(idx)
        lb.insert(new_idx, val)
        lb.selection_set(new_idx)
        lb.activate(new_idx)
        lb.see(new_idx)

    def _manual_intro_set_bg(self, block_idx):
        """Fixes the selected song from the main table for the intro block"""
        sel = self.tree.selection()
        if not sel:
            messagebox.showerror("Error", "Select a row in the main montage plan.")
            return
        
        vals = self.tree.item(sel[0], 'values')
        block_name = vals[0]
        ssid = vals[2]
        caption = vals[1]
        
        if "B01" in block_name or "Intro" in ssid:
            messagebox.showerror("Error", "Select a song block row (B02+).")
            return
            
        display_str = f"[Live+Combo] {caption} ({ssid})"
        lb = self.manual_blocks[block_idx]['listbox']
        
        # Remove old link if it existed
        for i in range(lb.size()):
            if "[Live+Combo]" in lb.get(i):
                lb.delete(i)
                break
        
        self.manual_intro_bg_songs[block_idx] = ssid
        lb.insert(tk.END, display_str)
        self.log(f"Montage2: For B01-0{block_idx} background set: {ssid}")

    def _on_manual_intro_radio_change(self):
        """Radio-button change handler"""
        if self.manual_intro_radio_var.get() in ("manual", "third_mode", "auto"):
            self.manual_intro_enabled.set(True)
        else:
            self.manual_intro_enabled.set(False)
        self._save_config()

    def _get_selected_ssid_slices_folder(self):
        """Returns path to slices folder of the SSID selected in treeview, or None"""
        ssid = self.active_ssid
        if not ssid or ssid == 'N/A (Intro)' or 'SSID' not in ssid:
            return None
        if ssid not in self.ssid_map:
            return None
        slices_path = Path(self.ssid_map[ssid]) / 'slices'
        if not slices_path.is_dir():
            return None
        return slices_path

    def _load_manual_intro_texts(self, df):
        """Loads B01-01, B01-02, B01-03 texts from Excel into text fields"""
        if df is None or df.empty:
            return

        # Reset indices
        self.manual_intro_b01_01_row_idx = None
        self.manual_intro_b01_02_row_idx = None
        self.manual_intro_b01_03_row_idx = None # Added

        for idx, row in df.iterrows():
            block_id = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ''
            text = str(row.iloc[1]) if pd.notna(row.iloc[1]) else ''
            
            if 'B01-01' in block_id:
                self.manual_intro_b01_01_row_idx = idx
                if self.manual_intro_text1_widget:
                    self.manual_intro_text1_widget.delete('1.0', tk.END); self.manual_intro_text1_widget.insert('1.0', text)
            elif 'B01-02' in block_id:
                self.manual_intro_b01_02_row_idx = idx
                if self.manual_intro_text2_widget:
                    self.manual_intro_text2_widget.delete('1.0', tk.END); self.manual_intro_text2_widget.insert('1.0', text)
            elif 'B01-03' in block_id:
                self.manual_intro_b01_03_row_idx = idx
                if self.manual_intro_text3_widget:
                    self.manual_intro_text3_widget.delete('1.0', tk.END); self.manual_intro_text3_widget.insert('1.0', text)

    def _save_text_generic(self, row_idx, widget, label):
        if not self.db_path or not self.db_path.exists(): return
        if not self.loaded_project_id: return
        if row_idx is None:
            self.log(f"Montage2: ERROR: Row {label} not found in plan.")
            return

        try:
            from openpyxl import load_workbook
            wb = load_workbook(self.db_path)
            ws = wb[self.loaded_project_id]
            text = widget.get('1.0', tk.END).strip()
            # row_idx in pandas is 0-based, in openpyxl it's 1-based. Plus header?
            # In _load_manual_intro_texts we use df.iterrows(), so idx matches 0-based row.
            # In Excel this will be idx + 1.
            ws.cell(row=row_idx + 1, column=2, value=text)
            wb.save(self.db_path)
            self.log(f"Montage2: Text {label} saved to Excel.")
        except Exception as e:
            self.log(f"Montage2: ERROR saving {label}: {e}")

    def _save_text1_to_excel(self): self._save_text_generic(self.manual_intro_b01_01_row_idx, self.manual_intro_text1_widget, "B01-01")
    def _save_text2_to_excel(self): self._save_text_generic(self.manual_intro_b01_02_row_idx, self.manual_intro_text2_widget, "B01-02")
    def _save_text3_to_excel(self): self._save_text_generic(self.manual_intro_b01_03_row_idx, self.manual_intro_text3_widget, "B01-03")

    def _scan_for_prefixes(self):
        if self.montage_plan_df is None:
            self.log("Montage2: ERROR: Plan not loaded for prefix scanning.")
            return
        thread = threading.Thread(target=self._scan_for_prefixes_worker, daemon=True)
        thread.start()

    def _scan_for_prefixes_worker(self):
        self.root.after(0, lambda: self.log("Montage2: Starting archive scan for prefixes..."))

        if not self.entity_id_map:
            if not self._build_entity_map():
                self.root.after(0, lambda: self.log(
                    "Montage2: ERROR: Failed to create entity map. Scanning aborted."))
                return

        eids = self.montage_plan_df['EID'].unique()
        result_list = []

        for eid in eids:
            if stop_flag.is_set():
                self.root.after(0, lambda: self.log("Montage2: Scanning aborted by user."))
                break

            eid_str = str(eid).strip()
            if eid_str in ['N/A (Intro)', 'N/A (Outro)']:
                continue

            eid_path = self.entity_id_map.get(eid_str)
            if not eid_path or not eid_path.exists():
                continue

            entity_name = eid_path.name
            meta_file = eid_path / 'entity_meta.json'
            if meta_file.exists():
                try:
                    with open(meta_file, 'r', encoding='utf-8') as f:
                        meta_data = json.load(f)
                    entity_name = meta_data.get('primary_name', entity_name)
                except Exception:
                    pass

            for subdir in ['interview_fragments', 'music_fragments', 'news_about']:
                sub_path = eid_path / subdir
                if sub_path.is_dir():
                    for file in sub_path.iterdir():
                        if file.is_file() and file.suffix.lower() in ('.mp4', '.mov', '.mkv', '.avi'):
                            result_list.append({
                                'entity_name': entity_name,
                                'name': file.name,
                                'type': subdir,
                                'path': str(file)
                            })

        self.root.after(0, self._populate_prefix_treeview, result_list)

    def _populate_prefix_treeview(self, result_list):
        for item in self.prefix_treeview.get_children():
            self.prefix_treeview.delete(item)
        for item in result_list:
            self.prefix_treeview.insert('', 'end',
                                        values=(item['entity_name'], item['type'], item['name'], item['path']))
        self.log(f"Montage2: Scanning completed. Found {len(result_list)} prefixes.")

    def _preview_prefix(self, event=None):
        selected_item_id = self.prefix_treeview.focus()
        if not selected_item_id:
            self.log("Montage2: Select a prefix from the list to find the file.")
            return

        item = self.prefix_treeview.item(selected_item_id)
        path = item['values'][3]

        if not Path(path).exists():
            self.log(f"Montage2: ERROR: File not found at path: {path}")
            return

        import subprocess
        import sys

        try:
            if sys.platform == "darwin":
                subprocess.Popen(['open', '-R', path])
            elif sys.platform == "win32":
                subprocess.Popen(f'explorer /select,"{path}"')
            else:
                subprocess.Popen(['xdg-open', str(Path(path).parent)])
        except Exception as e:
            self.log(f"Montage2: ERROR: Failed to show file. {e}")

    def _add_to_general_prefixes(self):
        selected = self.prefix_treeview.focus()
        if not selected: return
        item = self.prefix_treeview.item(selected)
        path = item['values'][3]
        if path not in self.general_prefixes:
            self.general_prefixes.append(path)
            self.general_prefixes_listbox.insert(tk.END, Path(path).name)
            self._save_config()
            self.log(f"Montage2: Prefix added: {Path(path).name}")

    def _remove_from_general_prefixes(self):
        selected_indices = self.general_prefixes_listbox.curselection()
        if not selected_indices:
            return

        index_to_remove = selected_indices[0]

        del self.general_prefixes[index_to_remove]
        self.general_prefixes_listbox.delete(index_to_remove)

        self._save_config()
        self.log("Montage2: General prefix removed.")

    def _move_up(self):
        selected_indices = self.general_prefixes_listbox.curselection()
        if not selected_indices or selected_indices[0] == 0:
            return
        index = selected_indices[0]

        self.general_prefixes[index], self.general_prefixes[index - 1] = self.general_prefixes[index - 1], \
        self.general_prefixes[index]

        text = self.general_prefixes_listbox.get(index)
        self.general_prefixes_listbox.delete(index)
        self.general_prefixes_listbox.insert(index - 1, text)
        self.general_prefixes_listbox.selection_set(index - 1)

        self._save_config()
        self.log("Montage2: Prefix moved up.")

    def _move_down(self):
        selected_indices = self.general_prefixes_listbox.curselection()
        if not selected_indices or selected_indices[0] >= self.general_prefixes_listbox.size() - 1:
            return
        index = selected_indices[0]

        self.general_prefixes[index], self.general_prefixes[index + 1] = self.general_prefixes[index + 1], \
        self.general_prefixes[index]

        text = self.general_prefixes_listbox.get(index)
        self.general_prefixes_listbox.delete(index)
        self.general_prefixes_listbox.insert(index + 1, text)
        self.general_prefixes_listbox.selection_set(index + 1)

        self._save_config()
        self.log("Montage2: Prefix moved down.")

    def _add_music_track(self):
        file_types = [("Audio", "*.mp3 *.wav *.m4a")]
        files = filedialog.askopenfilenames(title="Add tracks to playlist", filetypes=file_types)
        added = []
        for file in files:
            if file not in self.music_playlist:
                self.music_playlist.append(file)
                self.music_listbox.insert(tk.END, Path(file).name)
                added.append(Path(file).name)
        if added:
            self._save_config()
            self.log(f"Montage2: Tracks added: {', '.join(added)}")

    def _remove_music_track(self):
        selected = self.music_listbox.curselection()
        if not selected: return
        index = selected[0]

        del self.music_playlist[index]
        self.music_listbox.delete(index)

        self._save_config()
        self.log(f"Montage2: Track removed.")

    def _run_audio_test(self):
        # Checking for project plan
        if self.montage_plan_df is None or self.montage_plan_df.empty:
            self.log("Montage2: ERROR: Plan not loaded. Load a project for the test.")
            return

        # Updating archive map before test to ensure SSIDs are found
        self.log("Montage2: Scanning archive for sound test...")
        self._build_entity_map()

        # Checking audio folder
        if not self.assets['audio_folder'].get():
            self.log("Montage2: No audio folder for test.")
            return

        # 1. Find data for test (song1, song2, narration1, narration2)
        valid_blocks = []
        for _, row in self.montage_plan_df.iterrows():
            ssid_val = str(row.get('SSID', '')).strip()
            if ssid_val.startswith('SSID'):
                valid_blocks.append(row)

        if not valid_blocks:
            self.log("Montage2: ERROR: No blocks with SSID in plan for test.")
            return

        # We need 2 blocks.
        r1 = valid_blocks[0]
        r2 = None
        # Trying to find a second block with a DIFFERENT SSID
        for b in valid_blocks[1:]:
            if str(b.get('SSID')).strip() != str(r1.get('SSID')).strip():
                r2 = b
                break
        # If no other SSID found, just take the second block (if any)
        if r2 is None and len(valid_blocks) > 1:
            r2 = valid_blocks[1]
        # If there's only one block in the plan
        if r2 is None:
            r2 = r1

        test_rows = [r1, r2]
        sample_files = {
            'disclaimer': self.assets['prefix'].get(),
            'glitch': self.assets['glitch_effect'].get(),
            'ad': self.assets['ad_file'].get()
        }

        audio_folder = Path(self.assets['audio_folder'].get())

        for i, row in enumerate(test_rows, 1):
            ssid = str(row.get('SSID', '')).strip()
            block_name = str(row.get('Block', '')).strip()

            # Path to song
            song_folder = self.ssid_map.get(ssid)
            if not song_folder:
                self.log(f"Montage2: ERROR: Song {ssid} not found in archive.")
                return
            raw_v_dir = Path(song_folder) / "raw_videos"
            v_files = [f for f in raw_v_dir.iterdir() if f.is_file() and f.suffix.lower() in ('.mp4', '.mov', '.mkv', '.avi')]
            if not v_files:
                self.log(f"Montage2: ERROR: No videos in {raw_v_dir}.")
                return
            sample_files[f'song{i}'] = str(v_files[0])

            # Narration voice
            block_num_match = re.search(r'\d+', block_name)
            if not block_num_match:
                self.log(f"Montage2: ERROR: Could not determine block number from {block_name}.")
                return
            block_num = block_num_match.group()
            narration_files = [
                f for f in audio_folder.glob('*')
                if f.suffix.lower() in self.AUDIO_EXTENSIONS and
                (f.name.startswith(f"B{block_num}-") or f.name.startswith(f"{int(block_num)}."))
            ]
            if not narration_files:
                self.log(f"Montage2: ERROR: No narration audio files found for block {block_name}.")
                return
            narration_files.sort(key=lambda f: _parse_audio_filename(f.name) or float('inf'))
            sample_files[f'narration{i}'] = str(narration_files[0])

        settings = {
            'live_clip_duration': self.live_clip_duration.get(),
            'montage_mode': self.assets['montage_mode'].get(),
            'sound_settings': {
                'target_source_lufs': self.sound_settings['target_source_lufs'].get(),
                'bg_music_db': self.sound_settings['bg_music_db'].get(),
                'combo_vs_live_db': self.sound_settings['combo_vs_live_db'].get(),
                'final_limiter_tp': self.sound_settings['final_limiter_tp'].get()
            }
        }

        self.log(f"Montage2: Starting extended audio mixing test...")
        thread = threading.Thread(target=self._run_audio_test_worker, args=(sample_files, settings), daemon=True)
        thread.start()

    def _run_audio_test_worker(self, sample_files, settings):
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = Path(self.assets['output_folder'].get()) / f"audio_test_{timestamp}.mp3"
            montage_processor_2.generate_test_mix(sample_files, settings, str(output_path), self)
            self.root.after(0, lambda: self.log(f"Montage2: Test saved: {output_path}"))
        except Exception as e:
            self.root.after(0, lambda: self.log(f"Montage2: Test error: {e}"))
            self.root.after(0, lambda: self.log(f"--- TRACEBACK ---\n{traceback.format_exc()}--- END TRACEBACK ---"))

# --- END OF FILE montage_interface_2.py ---
#