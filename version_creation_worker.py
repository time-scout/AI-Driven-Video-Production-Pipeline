#
# --- START OF FILE version_creation_worker.py ---
#

import pandas as pd
import openpyxl
import random
import time
from pathlib import Path
from typing import Callable, Optional, Tuple, List, Dict
import re
import threading
import json
from concurrent.futures import ThreadPoolExecutor, as_completed

# --- Audio imports ---
import sys
from pydub import AudioSegment
from pydub.playback import play

# Import local modules
import versions_rewriter

def _play_alert_async(sound_type: str, sound_alerts_path: Path):
    """
    Asynchronously plays a sound file using pydub.
    Runs in a separate thread to avoid blocking the main process.
    """

    def task():
        try:
            sound_filename = {
                'fragment': 'version_fragment_complete.mp3',
                'success': 'version_complete.mp3',
                'error': 'version_error.mp3'
            }.get(sound_type)
            if not sound_filename:
                print(f"Unknown sound type: {sound_type}", file=sys.stderr)
                return
            sound_path = sound_alerts_path / sound_filename
            if not sound_path.exists():
                print(f"Sound file not found: {sound_path}", file=sys.stderr)
                return

            audio = AudioSegment.from_file(sound_path)
            play(audio)
        except Exception as e:
            # Print error to console but don't abort main process.
            print(f"Error playing sound {sound_type}: {e}", file=sys.stderr)
            print("Ensure ffmpeg is installed (e.g., 'brew install ffmpeg')", file=sys.stderr)

    # Start playing task in a separate thread
    threading.Thread(target=task, daemon=True).start()


def _strip_leading_enumeration(text: str) -> str:
    """Removes leading enumeration from text."""
    pattern = re.compile(
        r"^\s*("
        r"\d+[\.\)]?\s*|"
        r"\b(first|second|third|fourth|fifth|sixth|seventh|eighth|ninth|tenth|eleventh|twelfth|thirteenth|fourteenth|fifteenth|sixteenth|seventeenth|eighteenth|nineteenth|twentieth|one|two|three|four|five|six|seven|eight|nine|ten|finally)\b,?\s*|"
        r"[IVXLCDM]+\.\s*"
        r")+",
        re.IGNORECASE
    )
    return pattern.sub('', text).lstrip()


def _prepare_available_songs_data(all_blocks: list) -> Tuple[List[dict], dict]:
    """
    Extracts available songs from B02+ blocks and creates internal mapping.

    Returns:
        - available_songs_data: List of {"id": int, "song_by_artist": str} for AI
        - internal_mapping: Dict mapping id to full metadata
    """
    available_songs_data = []
    internal_mapping = {}
    unique_id = 1

    for block in all_blocks:
        if block['block_id'] == 'B01':
            continue

        caption = block.get('caption', '')
        # Parse "Song Name by Artist" format using the LAST occurrence of " by "
        by_matches = list(re.finditer(r'\s+by\s+', caption, flags=re.IGNORECASE))
        if by_matches:
            last_match = by_matches[-1]
            song_name = caption[:last_match.start()].strip()
            artist = caption[last_match.end():].strip()
        else:
            song_name = caption
            artist = "Unknown Artist"

        if song_name:
            song_by_artist = f"{song_name} by {artist}"
            available_songs_data.append({
                "id": unique_id,
                "song_by_artist": song_by_artist
            })

            internal_mapping[unique_id] = {
                "block_id": block['block_id'],
                "ssid": block.get('ssid', ''),
                "entity_id": block.get('entity_id', ''),
                "song_name": song_name,
                "artist_name": artist
            }
            unique_id += 1

    return available_songs_data, internal_mapping


def _split_text_programmatically(text: str, min_words: int = 14) -> List[str]:
    """
    Improved version of text splitting from script_creation_worker.py.
    Accounts for quotes, abbreviations, and sentence boundaries.
    """
    abbreviations = ["Mr.", "Mrs.", "Ms.", "Dr.", "St.", "Jr.", "Sr.",
                    "vs.", "approx.", "incl.", "etc.",
                    "Jan.", "Feb.", "Mar.", "Apr.", "Jun.", "Jul.", "Aug."
                    "Sep.", "Oct.", "Nov.", "Dec."]

    # 1. Protect abbreviations
    protected_text = text
    for abbr in abbreviations:
        prefix = abbr.rstrip('.')
        pattern_abbr = rf"\b({re.escape(prefix)})\."
        protected_text = re.sub(pattern_abbr, r"\1|DOT|", protected_text,
                                flags=re.IGNORECASE)

    # 2. Improved sentence splitting pattern (handles quotes)
    pattern = r'(?<=[.!?])\s+(?=[A-Z"“])|(?<=[.!?]["”])\s+(?=[A-Z"“])'
    sentences = re.split(pattern, protected_text.strip())

    # 3. Restore dots
    final_sentences = []
    for s in sentences:
        if s.strip():
            final_sentences.append(s.replace("|DOT|", ".").strip())

    fragments = []
    buffer = []
    current_word_count = 0

    for sentence in final_sentences:
        words = sentence.split()
        buffer.append(sentence)
        current_word_count += len(words)

        if current_word_count >= min_words:
            fragments.append(" ".join(buffer))
            buffer = []
            current_word_count = 0

    if buffer:
        if fragments:
            fragments[-1] = fragments[-1] + " " + " ".join(buffer)
        else:
            fragments.append(" ".join(buffer))

    return fragments


def _process_standard_block(block: dict, settings: dict) -> Optional[List[dict]]:
    """
    Processes standard blocks (B02+) with 2-step processing (rewrite → split).
    For v1: uses direct programmatic splitting without rewrite.
    For v2+: uses rewrite with AI-placed [PAUSE] markers.
    """
    ai_manager = settings['ai_manager']
    base_prompt_path = settings['prompt_path']
    target_version = settings.get('target_version', 'v2')
    log_callback = settings['log_callback_for_thread']

    source_text = _strip_leading_enumeration(block['source_text'])
    fragments = []

    if target_version in ['v1', 'v3']:
        # For v3 use limit from settings, for v1 use hardcoded 14
        limit = int(settings.get('main_limit', 14)) if target_version == 'v3' else 14
        log_callback(f"{target_version} mode: splitting block {block['block_id']} programmatically (limit: {limit}).")
        fragments = _split_text_programmatically(source_text, min_words=limit)
    else:
        # --- v2+ MODE: AI REWRITE + SPLIT ---
        # Use standard prompt from settings
        rewriter_prompt_path = base_prompt_path
        rewritten_text = versions_rewriter.rewrite_text_with_ai_manager(source_text, rewriter_prompt_path, ai_manager)

        if rewritten_text is None:
            log_callback(f"Critical rewrite error for block {block['block_id']}.")
            return None

        # Split text by [PAUSE] marker
        fragments = [f.strip() for f in rewritten_text.split('[PAUSE]') if f.strip()]
        log_callback(f"Block {block['block_id']} split into {len(fragments)} parts using [PAUSE] markers.")

    if not fragments:
        log_callback(f"Error: no fragments found after processing block {block['block_id']}.")
        return None

    # --- Result formation ---
    output_fragments = []
    for i, fragment_text in enumerate(fragments):
        new_block_id = f"{block['block_id']}-{i + 1:02d}"

        # Copy metadata from original block
        fragment_dict = {
            'block_id': new_block_id,
            'text': fragment_text.strip(),
            'ssid': block.get('ssid', ''),
            'entity_id': block.get('entity_id', ''),
            'caption': block.get('caption', '')
        }
        output_fragments.append(fragment_dict)

    return output_fragments


def _identify_teasers_with_ai(intro_text: str, available_songs: List[dict],
                             prompt_path: Path, ai_manager) -> Optional[List[int]]:
    """
    Uses AI to identify which songs are being teased in the intro.
    """
    try:
        # Load and format prompt
        prompt_template = prompt_path.read_text(encoding='utf-8')
        full_prompt = f"{prompt_template}\n\nOriginal Intro Text: {intro_text}\n\nAvailable Songs: {json.dumps(available_songs, ensure_ascii=False)}"

        # Call AI
        response = ai_manager.execute_ai_task(
            task_category="version_creation",
            input_data={"prompt": full_prompt}
        )

        if response.get("error"):
            return None

        # Parse JSON response
        result_text = response.get("text", "")
        json_match = re.search(r'\{.*\}', result_text, re.DOTALL)
        if not json_match:
            return None

        teaser_result = json.loads(json_match.group(0))
        return teaser_result.get("identified_teaser_ids_in_original_order", [])

    except Exception as e:
        print(f"Error identifying teasers: {e}")
        return None


def _compose_intro_with_ai(teasers: List[dict], original_intro_text: str, prompt_path: Path, ai_manager) -> Optional[str]:
    """
    Uses AI to compose a new intro based on identified teasers.
    """
    try:
        # Shuffle teasers for variety
        shuffled_teasers = teasers.copy()
        random.shuffle(shuffled_teasers)

        # Prepare input data
        input_json = json.dumps({"reordered_teasers_full_data": shuffled_teasers}, ensure_ascii=False)

        # Load and format prompt
        prompt_template = prompt_path.read_text(encoding='utf-8')
        full_prompt = f"{prompt_template}\n\nOriginal Intro Text: {original_intro_text}\n\nInput Metadata:\n{input_json}"

        # Call AI
        response = ai_manager.execute_ai_task(
            task_category="version_creation",
            input_data={"prompt": full_prompt}
        )

        if response.get("error"):
            return None

        return response.get("text", "").strip()

    except Exception as e:
        print(f"Error composing intro: {e}")
        return None


def _rewrite_fallback_intro_with_ai(original_text: str, prompt_path: Path, ai_manager) -> Optional[str]:
    """
    Performs intro rewrite in fallback mode (without teasers).
    """
    try:
        prompt_template = prompt_path.read_text(encoding='utf-8')
        full_prompt = f"{prompt_template}\n\nOriginal Intro Text: {original_text}"

        response = ai_manager.execute_ai_task(
            task_category="version_creation",
            input_data={"prompt": full_prompt}
        )

        if response.get("error"):
            return None

        return response.get("text", "").strip()
    except Exception as e:
        print(f"Error in fallback intro rewrite: {e}")
        return None


def _process_intro_block(intro_block: dict, all_blocks: list, settings: dict) -> Optional[Tuple[List[dict], List[dict]]]:
    """
    Performs processing for intro blocks (B01).
    For v1: direct programmatic splitting of source text, but AI identifies mentioned songs.
    For v2+: 3-step AI processing (identify -> compose -> split).
    """
    ai_manager = settings['ai_manager']
    log_callback = settings['log_callback_for_thread']
    prompts_base_path = settings['prompt_path'].parent
    target_version = settings.get('target_version', 'v2')

    # Step A: Prepare available songs data (MANDATORY for all versions)
    available_songs_data, internal_mapping = _prepare_available_songs_data(all_blocks)

    if not available_songs_data:
        log_callback("No available songs found for intro processing")
        return None

    # Step B: Identify songs mentioned in intro (MANDATORY for all versions)
    teaser_identifier_prompt = prompts_base_path / 'version_intro_teaser_identifier.txt'
    identified_ids = _identify_teasers_with_ai(
        intro_block['source_text'],
        available_songs_data,
        teaser_identifier_prompt,
        ai_manager
    )

    identification_source = ""
    reordered_teasers_full_data = []
    composed_intro = ""

    if target_version in ['v1', 'v3']:
        limit = int(settings.get('intro_limit', 14)) if target_version == 'v3' else 14
        log_callback(f"{target_version} mode: preserving original text for B01 (limit: {limit}).")

        identification_source = "Genuine Match" if identified_ids else f"{target_version}-original"
        composed_intro = _strip_leading_enumeration(intro_block['source_text'])

        if identified_ids:
            for teaser_id in identified_ids:
                if teaser_id in internal_mapping:
                    reordered_teasers_full_data.append(internal_mapping[teaser_id])

        fragments = _split_text_programmatically(composed_intro, min_words=limit)

    else:
        # --- v2+ MODE: AI REWRITE + COMPOSE ---
        if not identified_ids:
            # FALLBACK MODE: Rewrite without song binding
            log_callback("No teasers identified - using General Fallback Rewrite")
            identification_source = "Fallback"
            fallback_prompt = prompts_base_path / 'version_short_intro_fallback_splitter_prompt.txt'
            composed_intro = _rewrite_fallback_intro_with_ai(
                intro_block['source_text'],
                fallback_prompt,
                ai_manager
            )
        else:
            # GENUINE MATCH: AI successfully identified teasers
            log_callback(f"AI identified {len(identified_ids)} teasers - using Genuine Match")
            identification_source = "Genuine Match"

            identified_teasers = []
            for teaser_id in identified_ids:
                if teaser_id in internal_mapping:
                    identified_teasers.append(internal_mapping[teaser_id])

            if not identified_teasers:
                log_callback("No valid teaser data found for identified IDs")
                return None

            reordered_teasers_full_data = identified_teasers.copy()
            random.shuffle(reordered_teasers_full_data)

            composer_prompt = prompts_base_path / 'version_intro_composer.txt'
            composed_intro = _compose_intro_with_ai(
                reordered_teasers_full_data,
                intro_block['source_text'],
                composer_prompt,
                ai_manager
            )

        if not composed_intro:
            log_callback("Failed to compose intro")
            return None

        # AI Split for v2+ (via [PAUSE] markers)
        fragments = [f.strip() for f in composed_intro.split('[PAUSE]') if f.strip()]
        log_callback(f"Intro split into {len(fragments)} parts using [PAUSE] markers.")

    if not fragments:
        log_callback("Failed to split intro into fragments")
        return None

    # Add identification_source to each teaser for Excel output
    for teaser in reordered_teasers_full_data:
        teaser['identification_source'] = identification_source

    # Form final fragments for main table (A-K columns)
    final_fragments = []
    for i, fragment_text in enumerate(fragments):
        new_block_id = f"{intro_block['block_id']}-{i + 1:02d}"
        final_fragments.append({
            'block_id': new_block_id,
            'text': fragment_text.strip(),
            'ssid': intro_block.get('ssid', ''),
            'entity_id': intro_block.get('entity_id', ''),
            'caption': intro_block.get('caption', '')
        })

    return final_fragments, reordered_teasers_full_data

    if not available_songs_data:
        log_callback("No available songs found for intro processing")
        return None

    # Step B: Identify teasers
    teaser_identifier_prompt = prompts_base_path / 'version_intro_teaser_identifier.txt'
    identified_ids = _identify_teasers_with_ai(
        intro_block['source_text'],
        available_songs_data,
        teaser_identifier_prompt,
        ai_manager
    )

    # Handle both Genuine Match and Fallback scenarios
    identification_source = ""
    reordered_teasers_full_data = []
    composed_intro = ""

    if not identified_ids:
        # FALLBACK MODE: Rewrite without song binding
        log_callback("No teasers identified - using General Fallback Rewrite")
        identification_source = "Fallback"
        reordered_teasers_full_data = []

        fallback_prompt = prompts_base_path / 'version_short_intro_fallback_splitter_prompt.txt'
        composed_intro = _rewrite_fallback_intro_with_ai(
            intro_block['source_text'],
            fallback_prompt,
            ai_manager
        )

    else:
        # GENUINE MATCH: AI successfully identified teasers
        log_callback(f"AI identified {len(identified_ids)} teasers - using Genuine Match")
        identification_source = "Genuine Match"

        # Step C: Collect full data for identified teasers and shuffle
        identified_teasers = []
        for teaser_id in identified_ids:
            if teaser_id in internal_mapping:
                identified_teasers.append(internal_mapping[teaser_id])

        if not identified_teasers:
            log_callback("No valid teaser data found for identified IDs")
            return None

        # Shuffle for variety
        reordered_teasers_full_data = identified_teasers.copy()
        random.shuffle(reordered_teasers_full_data)

        # Step D: Compose intro using AI composer
        composer_prompt = prompts_base_path / 'version_intro_composer.txt'
        composed_intro = _compose_intro_with_ai(
            reordered_teasers_full_data,
            intro_block['source_text'],
            composer_prompt,
            ai_manager
        )

    if not composed_intro:
        log_callback("Failed to compose intro")
        return None

    # Step E: Split into fragments (NEW LOGIC)
    fragments = [f.strip() for f in composed_intro.split('[PAUSE]') if f.strip()]

    if not fragments:
        log_callback("Failed to split intro into fragments (no markers found)")
        return None

    log_callback(f"Intro split into {len(fragments)} parts using [PAUSE] markers.")

    # Add identification_source to each teaser for Excel output
    for teaser in reordered_teasers_full_data:
        teaser['identification_source'] = identification_source

    # Step F: Create final fragments for main table (A-K columns)
    final_fragments = []
    for i, fragment_text in enumerate(fragments):
        new_block_id = f"{intro_block['block_id']}-{i + 1:02d}"

        fragment_dict = {
            'block_id': new_block_id,
            'text': fragment_text.strip(),
            'ssid': intro_block.get('ssid', ''),
            'entity_id': intro_block.get('entity_id', ''),
            'caption': intro_block.get('caption', '')
        }
        final_fragments.append(fragment_dict)

    return final_fragments, reordered_teasers_full_data


def _process_block(block: dict, all_blocks: list, settings: dict) -> Optional[Tuple[List[dict], Optional[List[dict]]]]:
    """
    Routes block to appropriate processor based on block type.
    B01 gets special 3-step intro processing, B02+ get standard 2-step.

    Returns:
        - final_fragments: List of fragment dictionaries for main table (A-K)
        - intro_teasers: List of teaser metadata for intro metadata table (P-U), or None for non-intro blocks
    """
    if block['block_id'] == 'B01':
        result = _process_intro_block(block, all_blocks, settings)
        if result:
            return result[0], result[1]  # final_fragments, reordered_teasers_full_data
        return None, None
    else:
        fragments = _process_standard_block(block, settings)
        return fragments, None


def _rewrite_and_split_block(block: dict, settings: dict) -> Optional[List[dict]]:
    """
    Legacy function - kept for backward compatibility.
    Use _process_block instead.
    """
    fragments, _ = _process_block(block, [], settings)  # Pass empty all_blocks for backward compatibility
    return fragments


def _validate_blocks_structure(blocks: List[Dict], log_callback: Callable) -> bool:
    """
    Checks block structure consistency before processing.
    """
    log_callback("  - Checking block structure consistency...")

    if not blocks or len(blocks) < 3:
        log_callback(
            f"VALIDATION ERROR: Found only {len(blocks)} blocks. At least 3 are required (intro, body, outro).")
        return False

    for i, block in enumerate(blocks):
        text = block.get('source_text', '').strip()
        if not text:
            log_callback(
                f"VALIDATION ERROR: Block #{i + 1} (ID: {block.get('block_id', 'N/A')}) contains no text. Block chain is broken.")
            return False

    log_callback("  - Block structure is valid.")
    return True


def _read_source_sheet(video_id: str, db_path: Path, log_callback: Callable) -> Optional[Tuple[List[Dict], str]]:
    """Reads data from the source Excel sheet."""
    try:
        df = pd.read_excel(db_path, sheet_name=video_id, header=None)
        df = df.fillna('')

        # URL is now in column 7 (index 6)
        url = df.iloc[1, 6] if len(df.columns) > 6 and len(df) > 1 else ""

        log_callback("  - Using column B for text reading.")

        # New column structure:
        # A (0) - Block ID, B (1) - Text, C (2) - SSID, D (3) - EID, E (4) - Clean Caption
        COL_BLOCK_ID, COL_TEXT, COL_SSID, COL_EID, COL_CAPTION = 0, 1, 2, 3, 4

        blocks = []
        for index, row in df.iloc[2:].iterrows():
            block_id = str(row.get(COL_BLOCK_ID, '')).strip()
            source_text = str(row.get(COL_TEXT, '')).strip()

            # Add block only if there is a block ID or text
            if block_id or source_text:
                blocks.append({
                    'block_id': block_id,
                    'source_text': source_text,
                    'entity_name': '',  # To be filled later if needed
                    'entity_id': str(row.get(COL_EID, '')).strip(),
                    'ssid': str(row.get(COL_SSID, '')).strip(),
                    'caption': str(row.get(COL_CAPTION, '')).strip() if len(df.columns) > COL_CAPTION else ''
                })

        if not blocks:
            log_callback("WARNING: No blocks found for processing in the source sheet.")
            return None

        log_callback(f"  - Found {len(blocks)} potential blocks for processing.")
        return blocks, url

    except ValueError:
        log_callback(f"ERROR: Sheet with name '{video_id}' not found.")
        return None
    except Exception as e:
        log_callback(f"ERROR reading Excel: {e}")
        return None


def _process_blocks_concurrently(blocks: list, settings: dict, log_callback: Callable, stop_event) -> Optional[Tuple[List[dict], Optional[List[dict]]]]:
    """Processes blocks (rewrite + split) using a pool of 10 threads via AI Manager."""
    ai_manager = settings.get('ai_manager')
    sound_alerts_path = settings.get('sound_alerts_path')

    if not ai_manager:
        log_callback("ERROR: AI Manager not passed to settings")
        return None

    total_blocks = len(blocks)
    MAX_WORKERS = 10

    log_callback(f"  - Starting rewrite and split in {MAX_WORKERS} threads via AI Manager...")

    # Add callback to settings to pass into thread
    settings['log_callback_for_thread'] = log_callback

    final_fragments_list = []
    intro_teasers_data = None  # Will hold the reordered_teasers_full_data from intro processing
    processed_count = 0

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        future_to_block = {executor.submit(_process_block, block, blocks, settings): block for block in blocks}

        for future in as_completed(future_to_block):
            if stop_event.is_set():
                log_callback("  - Stop signal received. Aborting processing...")
                executor.shutdown(wait=False, cancel_futures=True)
                return None

            original_block = future_to_block[future]
            try:
                fragments_result, teasers_result = future.result()

                if fragments_result is None:
                    log_callback(f"❌ CRITICAL ERROR: Failed to process block {original_block.get('block_id', 'N/A')}. Process aborted.")
                    _play_alert_async('error', sound_alerts_path)
                    executor.shutdown(wait=False, cancel_futures=True)
                    return None

                final_fragments_list.extend(fragments_result)

                # Store intro teasers data if this was the intro block
                if original_block['block_id'] == 'B01' and teasers_result is not None:
                    intro_teasers_data = teasers_result

                processed_count += 1
                log_callback(f"✅ Block {original_block.get('block_id', 'N/A')} ready. ({processed_count}/{total_blocks})")

            except Exception as e:
                log_callback(f"  - ❌ CRITICAL ERROR during block {original_block.get('block_id', 'N/A')} processing: {e}. Process aborted.")
                _play_alert_async('error', sound_alerts_path)
                executor.shutdown(wait=False, cancel_futures=True)
                return None

    # Important! Need to sort `final_fragments_list` by `block_id`
    # so fragments follow in the correct order (B01-01, B01-02..., B02-01...)
    final_fragments_list.sort(key=lambda x: x['block_id'])

    log_callback("All blocks successfully processed and split.")
    return final_fragments_list, intro_teasers_data


def _write_new_sheet(final_fragments: list, intro_teasers_data: Optional[List[dict]], settings: dict, source_url: str, log_callback: Callable) -> bool:
    """Writes final result to a new Excel sheet with two independent tables."""
    try:
        db_path, new_sheet_name, source_sheet_name = settings['db_path'], settings['new_sheet_name'], settings[
            'video_id']
        workbook = openpyxl.load_workbook(db_path)

        if new_sheet_name in workbook.sheetnames:
            log_callback(f"  - Sheet '{new_sheet_name}' already exists. Deleting old version.")
            del workbook[new_sheet_name]

        source_sheet_index = workbook.sheetnames.index(source_sheet_name)
        sheet = workbook.create_sheet(title=new_sheet_name, index=source_sheet_index + 1)

        # --- Table 1: Main Table Headers (Columns A-K) ---
        sheet['A1'] = "Block ID"
        sheet['B1'] = "Text"
        sheet['C1'] = "SSID"
        sheet['D1'] = "EID"
        sheet['F1'] = "url"
        sheet['K1'] = "Caption"

        # --- Table 2: Intro Teaser Metadata Table Headers (Columns P-V) ---
        sheet['P1'] = "Intro Song Number"
        sheet['Q1'] = "Intro Song Name"
        sheet['R1'] = "SSID"
        sheet['S1'] = "Intro Artist"
        sheet['T1'] = "EID"
        sheet['U1'] = "Original Block ID"
        sheet['V1'] = "Identification Source"

        # Write URL to the correct place
        sheet['F2'] = source_url

        # --- Table 1: Write Main Table Data (Columns A-K) ---
        for row_idx, fragment_data in enumerate(final_fragments, start=3):
            # Mandatory fields for each row
            sheet[f'A{row_idx}'] = fragment_data['block_id']
            sheet[f'B{row_idx}'] = fragment_data.get('text', '')

            # Fields written ONLY for the first fragment (-01)
            if fragment_data['block_id'].endswith('-01'):
                sheet[f'C{row_idx}'] = fragment_data.get('ssid', '')
                sheet[f'D{row_idx}'] = fragment_data.get('entity_id', '')
                sheet[f'K{row_idx}'] = fragment_data.get('caption', '')
            else:
                # Leave cells empty for other fragments
                sheet[f'C{row_idx}'] = ''
                sheet[f'D{row_idx}'] = ''
                sheet[f'K{row_idx}'] = ''

        # --- Table 2: Write Intro Teaser Metadata (Columns P-V) ---
        if intro_teasers_data:
            for row_idx, teaser_data in enumerate(intro_teasers_data, start=3):
                # Sequential story order within intro (1, 2, 3...)
                sheet[f'P{row_idx}'] = row_idx - 2  # Row 3 -> 1, Row 4 -> 2, etc.
                sheet[f'Q{row_idx}'] = teaser_data.get('song_name', '')
                sheet[f'R{row_idx}'] = teaser_data.get('ssid', '')
                sheet[f'S{row_idx}'] = teaser_data.get('artist_name', '')
                sheet[f'T{row_idx}'] = teaser_data.get('entity_id', '')
                sheet[f'U{row_idx}'] = teaser_data.get('block_id', '')
                sheet[f'V{row_idx}'] = teaser_data.get('identification_source', '')

        workbook.save(db_path)
        return True
    except Exception as e:
        log_callback(f"ERROR writing to Excel: {e}")
        return False


def run_version_creation(settings: dict, log_callback: Callable, stop_event: "threading.Event", sound_alerts_path: Path):
    """Main orchestrator function for the entire version creation process."""
    try:
        video_id = settings['video_id']
        log_callback(f"Starting process for Video ID: {video_id}")

        if not settings['db_path'].exists():
            log_callback(f"CRITICAL ERROR: Database file not found: {settings['db_path']}")
            _play_alert_async('error', sound_alerts_path)
            return

        # --- STEP 1: Reading data ---
        log_callback("Step 1/4: Reading source sheet...")
        source_data = _read_source_sheet(video_id, settings['db_path'], log_callback)
        if source_data is None:
            log_callback("Process aborted due to reading error.")
            _play_alert_async('error', sound_alerts_path)
            return
        blocks_to_process, source_url = source_data

        # --- STEP 2: Primary validation ---
        log_callback("Step 2/4: Data structure validation...")
        if not _validate_blocks_structure(blocks_to_process, log_callback):
            _play_alert_async('error', sound_alerts_path)
            log_callback("Process aborted due to inconsistent data.")
            return

        if stop_event.is_set(): log_callback("Process stopped by user."); return

        # --- STEP 3: Multi-threaded processing (rewrite + split) ---
        log_callback(f"Step 3/4: Processing {len(blocks_to_process)} text blocks (rewrite + split)...")
        processing_result = _process_blocks_concurrently(blocks_to_process, settings, log_callback, stop_event)

        if processing_result is None:
            log_callback("Processing failed or was aborted. Shutting down.")
            return

        final_fragments, intro_teasers_data = processing_result

        if stop_event.is_set(): log_callback("Process stopped by user."); return

        # --- STEP 4: Result writing ---
        log_callback("Step 4/4: Writing result to new sheet...")
        success = _write_new_sheet(final_fragments, intro_teasers_data, settings, source_url, log_callback)
        if success:
            log_callback(f"✅ SUCCESS! New version successfully created in sheet '{settings['new_sheet_name']}'.")
            _play_alert_async('success', sound_alerts_path)
        else:
            log_callback("❌ ERROR: Failed to write result to Excel.")
            _play_alert_async('error', sound_alerts_path)

    except Exception as e:
        log_callback(f"CRITICAL UNEXPECTED ERROR in process: {e}")
        _play_alert_async('error', sound_alerts_path)

#
# --- END OF FILE version_creation_worker.py ---
#