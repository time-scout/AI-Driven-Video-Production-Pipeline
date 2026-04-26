# process_visuals_interface.py

import tkinter as tk
from tkinter import ttk, scrolledtext, simpledialog
import subprocess
import sys
import webbrowser
import urllib.parse
import urllib.request
import queue
import threading
from pathlib import Path
from typing import List, Dict, Optional, Tuple
import re
import shutil
import datetime

import entity_manager_v2
import download_orchestrator_v2
import video_slicer_v2
import zoom_creator_v2
import video_parser_v2
import json
import pandas as pd
from openpyxl import load_workbook

class ToolTip:
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tooltip_window = None
        self.widget.bind("<Enter>", self.show_tooltip)
        self.widget.bind("<Leave>", self.hide_tooltip)

    def show_tooltip(self, event):
        x, y, _, _ = self.widget.bbox("insert")
        x += self.widget.winfo_rootx() + 25
        y += self.widget.winfo_rooty() + 25

        self.tooltip_window = tk.Toplevel(self.widget)
        self.tooltip_window.wm_overrideredirect(True)
        self.tooltip_window.wm_geometry(f"+{x}+{y}")

        label = tk.Label(self.tooltip_window, text=self.text, justify='left',
                         background="#ffffe0", relief='solid', borderwidth=1,
                         wraplength=300,
                         font=("tahoma", "8", "normal"))
        label.pack(ipadx=1)

    def hide_tooltip(self, event):
        if self.tooltip_window:
            self.tooltip_window.destroy()
        self.tooltip_window = None



class OverwriteDialog(tk.Toplevel):
    def __init__(self, parent, folder_path, entity_name):
        super().__init__(parent)
        self.transient(parent)
        self.title("Folder not empty")
        self.parent = parent
        self.result = "skip"
        message = f"Folder '{folder_path.name}' for entity '{entity_name}' already contains files.\n\nChoose action:"
        ttk.Label(self, text=message, justify='left').pack(padx=20, pady=(20, 10))
        btn_frame = ttk.Frame(self)
        btn_frame.pack(padx=20, pady=(10, 20))
        ttk.Button(btn_frame, text="Overwrite (delete old)", command=lambda: self.set_result("overwrite")).pack(
            side="left", padx=5)
        ttk.Button(btn_frame, text="Add (keep old)", command=lambda: self.set_result("append")).pack(
            side="left", padx=5)
        ttk.Button(btn_frame, text="Skip this task", command=lambda: self.set_result("skip")).pack(side="left",
                                                                                                          padx=5)
        self.protocol("WM_DELETE_WINDOW", lambda: self.set_result("skip"))
        self.grab_set()
        self.wait_window(self)

    def set_result(self, result):
        self.result = result
        self.destroy()


class OverwriteInterviewDialog(tk.Toplevel):
    def __init__(self, parent, folder_path, entity_name):
        super().__init__(parent)
        self.transient(parent)
        self.title("Folder interview_fragments not empty")
        self.parent = parent
        self.result = "skip"
        message = f"Folder 'interview_fragments' for '{entity_name}' already contains clips.\n\nSelect action:"
        ttk.Label(self, text=message, justify='left').pack(padx=20, pady=(20, 10))
        btn_frame = ttk.Frame(self)
        btn_frame.pack(padx=20, pady=(10, 20))
        ttk.Button(btn_frame, text="Overwrite (delete all except 'h_')", command=lambda: self.set_result("overwrite")).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="Add (keep old)", command=lambda: self.set_result("append")).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="Skip interview slicing", command=lambda: self.set_result("skip")).pack(side="left", padx=5)
        self.protocol("WM_DELETE_WINDOW", lambda: self.set_result("skip"))
        self.grab_set()
        self.wait_window(self)

    def set_result(self, result):
        self.result = result
        self.destroy()


class DurationCheckDialog(tk.Toplevel):
    def __init__(self, parent, long_videos_info):
        super().__init__(parent)
        self.transient(parent)
        self.title("Long videos detected")
        self.parent = parent
        self.result = "cancel"
        self.geometry("600x400")
        self.minsize(500, 300)
        main_frame = ttk.Frame(self, padding=10)
        main_frame.pack(fill='both', expand=True)
        message = "The following videos are longer than 30 minutes. Select action:"
        ttk.Label(main_frame, text=message, wraplength=580).pack(pady=(0, 10), anchor='w')
        text_frame = ttk.Frame(main_frame)
        text_frame.pack(fill='both', expand=True, pady=5)
        info_text = scrolledtext.ScrolledText(text_frame, wrap='word', height=10)
        info_text.pack(side='left', fill='both', expand=True)
        scrollbar = ttk.Scrollbar(text_frame, orient='vertical', command=info_text.yview)
        scrollbar.pack(side='right', fill='y')
        info_text['yscrollcommand'] = scrollbar.set
        for info in long_videos_info:
            duration_str = video_parser_v2.format_duration(info['duration'])
            line = f"Duration: {duration_str} — {info.get('title', 'Untitled')}\n{info['url']}\n\n"
            info_text.insert(tk.END, line)
        info_text.config(state='disabled')
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill='x', pady=(10, 0))
        ttk.Button(btn_frame, text="Cancel (edit list)", command=lambda: self.set_result("cancel")).pack(
            side="right", padx=5)
        ttk.Button(btn_frame, text="Download only short", command=lambda: self.set_result("only_short")).pack(
            side="right", padx=5)
        ttk.Button(btn_frame, text="Download anyway", command=lambda: self.set_result("all")).pack(side="right",
                                                                                                     padx=5)
        self.protocol("WM_DELETE_WINDOW", lambda: self.set_result("cancel"))
        self._load_semantic_settings()
        self.grab_set()
        self.wait_window(self)

    def set_result(self, result):
        self.result = result
        self.destroy()


class ProcessingTab(ttk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent, padding=10)
        self.app = app
        self.active_entity_data = {}
        self.stop_event = threading.Event()
        self.on_top_var = tk.BooleanVar(value=False)
        self._last_resize_width = 0
        self.loaded_source = None      # 'excel', 'json', or None
        self.loaded_files = []          # List of loaded files
        self.order_modified = False     # Order change flag
        self.photo_dl_queue = queue.Queue();
        self.video_dl_queue = queue.Queue()
        self.slicer_queue = queue.Queue();
        self.slices_queue = queue.Queue()
        self.zoom_queue = queue.Queue()
        self.semantic_slicer_queue = queue.Queue()
        self.conversion_queue = queue.Queue()
        self.composite_queue = queue.Queue()
        self.photo_worker_thread = None;
        self.video_worker_thread = None
        self.interview_video_links_text = None
        self.slicer_worker_thread = None;
        self.slices_worker_thread = None
        self.zoom_worker_thread = None
        self.semantic_slicer_worker_thread = None # <<< ADD THIS LINE
        self.conversion_worker_thread = None
        self.composite_worker_thread = None
        self.refresh_thread = None
        self.video_check_thread = None
        self.app.update_processing_tab_row = self.update_treeview_row

        # Database paths configuration
        if self.app.DATABASE_PATH:
            self.DB_PATH = self.app.DATABASE_PATH / 'main_database.xlsx'
        else:
            self.DB_PATH = None
            
        self.master_db = None
        self.songs_db = None

        self._build_header_controls(self)
        self._load_semantic_settings()
        self._load_slicer_settings()
        main_container = ttk.Frame(self)
        main_container.pack(fill='both', expand=True, pady=(0, 10))
        main_container.rowconfigure(0, weight=1)
        main_container.columnconfigure(0, weight=1)
        main_container.columnconfigure(1, weight=0)
        self._build_table_container(main_container)
        self._build_action_panel_with_tabs(main_container)
        self._build_log_panels(self)
        self.action_panel.update_idletasks()
        min_width = self.action_panel.winfo_reqwidth()
        main_container.columnconfigure(1, minsize=min_width)

        # yt-dlp tab status initialization
        self._update_tab_status()

    def _build_header_controls(self, parent):
        header_frame = ttk.Frame(parent)
        header_frame.pack(fill='x', side='top', pady=(0, 5))

        # Left side: project input and loading
        load_frame = ttk.Frame(header_frame)
        load_frame.pack(side='left', padx=(0, 10))

        ttk.Label(load_frame, text="Video ID:").pack(side='left', padx=(0, 5))
        self.video_id_entry = ttk.Entry(load_frame, width=12)
        self.video_id_entry.pack(side='left', padx=(0, 5))
        self.video_id_entry.insert(0, "1") # Default 1

        self.load_button = ttk.Button(load_frame, text="Load from Excel", command=self._on_load_click)
        self.load_button.pack(side='left', padx=(0, 5))

        self.load_json_button = ttk.Button(load_frame, text="Load from JSON", command=self._on_load_from_json_click)
        self.load_json_button.pack(side='left', padx=(0, 10))

        # Statistics update button
        self.refresh_button = ttk.Button(header_frame, text="Update Statistics",
                                         command=self.refresh_current_list_stats)
        self.refresh_button.pack(side='left', padx=(0, 10))

        self.scan_status_label = ttk.Label(header_frame, text="", foreground="#007bff")
        self.scan_status_label.pack(side='left', pady=(2, 0))

        on_top_check = ttk.Checkbutton(header_frame, text="Always on top", variable=self.on_top_var,
                                       command=self._toggle_always_on_top)
        on_top_check.pack(side='right')

    def _toggle_always_on_top(self):
        self.app.root.attributes('-topmost', self.on_top_var.get())

    def _build_table_container(self, parent):
        table_container = ttk.Frame(parent)
        table_container.grid(row=0, column=0, sticky='nsew', padx=(0, 10))
        table_container.rowconfigure(0, weight=1)
        table_container.columnconfigure(0, weight=1)
        
        # New columns according to task
        self.column_configs = [
            ("artist", 200, 20),
            ("song", 200, 20),
            ("ssid", 100, 20),
            ("clip", 50, 20),
            ("slices", 30, 10),
            ("visuals", 30, 10)
        ]
        
        columns = [c[0] for c in self.column_configs]
        self.tree = ttk.Treeview(table_container, columns=columns, show="headings", height=20)
        self.tree.grid(row=0, column=0, sticky='nsew')
        
        self.tree.heading("artist", text="Artist")
        self.tree.column("artist", width=200, stretch=tk.YES, minwidth=120)
        
        self.tree.heading("song", text="Song")
        self.tree.column("song", width=200, stretch=tk.YES, minwidth=120)
        
        self.tree.heading("ssid", text="SSID")
        self.tree.column("ssid", width=100, anchor='center', stretch=tk.NO, minwidth=80)
        
        self.tree.heading("clip", text="Clip")
        self.tree.column("clip", width=50, anchor='center', stretch=tk.NO, minwidth=50)

        self.tree.heading("slices", text="Slices")
        self.tree.column("slices", width=30, anchor='center', stretch=tk.NO, minwidth=30)

        self.tree.heading("visuals", text="Visuals")
        self.tree.column("visuals", width=30, anchor='center', stretch=tk.NO, minwidth=30)

        self.tree.bind('<<TreeviewSelect>>', self.on_entity_select)
        scrollbar = ttk.Scrollbar(table_container, orient="vertical", command=self.tree.yview)
        scrollbar.grid(row=0, column=1, sticky='ns')
        self.tree.configure(yscrollcommand=scrollbar.set)
        # Disable complex resize logic as there are few columns now
        # table_container.bind('<Configure>', self._on_resize_columns)

    def _build_action_panel_with_tabs(self, parent):
        self.action_panel = ttk.Notebook(parent)
        self.action_panel.grid(row=0, column=1, sticky='nsew')
        self.order_tab_frame = ttk.Frame(self.action_panel, padding=10)
        self.photo_tab_frame = ttk.Frame(self.action_panel, padding=10)
        self.video_tab_frame = ttk.Frame(self.action_panel, padding=10)
        self.action_panel.add(self.order_tab_frame, text="Block Order")
        self.action_panel.add(self.photo_tab_frame, text="Photo Work")
        self.action_panel.add(self.video_tab_frame, text="Video Work")
        self.action_panel.bind("<<NotebookTabChanged>>", self._on_tab_changed)
        self._populate_order_tab(self.order_tab_frame)
        self._populate_photo_tab(self.photo_tab_frame)
        self._populate_video_tab(self.video_tab_frame)
        self.set_action_panel_state('disabled')

    def _populate_photo_tab(self, parent_frame):
        nav_panel = self._create_common_nav_panel(parent_frame, tab_type='photo')
        nav_panel.pack(fill='x', pady=(0, 5), anchor='w')
        download_frame = ttk.LabelFrame(parent_frame, text="Download Photos", padding=5)
        download_frame.pack(fill='both', expand=True, pady=5)
        photo_actions_frame = ttk.Frame(download_frame)
        photo_actions_frame.pack(fill='x', pady=2, side='top')
        photo_count_frame = ttk.Frame(photo_actions_frame)
        photo_count_frame.pack(side='left', anchor='w', pady=(0, 10))
        ttk.Label(photo_count_frame, text="Photo count:").pack(side='left')
        self.photo_count_spinbox = ttk.Spinbox(photo_count_frame, from_=1, to=999, width=5)
        self.photo_count_spinbox.set(60)
        self.photo_count_spinbox.pack(side='left', padx=5)
        buttons_frame = ttk.Frame(download_frame)
        buttons_frame.pack(fill='x', pady=2)
        self.add_photo_task_btn = ttk.Button(buttons_frame, text="Download for selected", command=self.add_photo_task)
        self.add_photo_task_btn.pack(pady=2)
        self.batch_photo_task_btn = ttk.Button(buttons_frame, text="Download for all empty",
                                               command=self.add_batch_photo_task)
        self.batch_photo_task_btn.pack(pady=2)
        self.manual_query_photo_task_btn = ttk.Button(buttons_frame, text="Download with manual query...",
                                                      command=self.add_manual_query_photo_task)
        self.manual_query_photo_task_btn.pack(pady=2)
        zoom_frame = ttk.LabelFrame(parent_frame, text="Create Zooms from Photos", padding=5)
        zoom_frame.pack(fill='x', pady=5)
        zoom_grid = ttk.Frame(zoom_frame)
        zoom_grid.pack()
        ttk.Label(zoom_grid, text="Duration (sec):").grid(row=0, column=0, sticky='w', pady=2)
        self.zoom_duration = ttk.Spinbox(zoom_grid, from_=1, to=99, width=5);
        self.zoom_duration.set(6);
        self.zoom_duration.grid(row=0, column=1, sticky='ew', padx=5)
        ttk.Label(zoom_grid, text="Direction:").grid(row=1, column=0, sticky='w', pady=2)
        self.zoom_direction = ttk.Combobox(zoom_grid, values=["Zoom In", "Zoom Out", "Pan Left", "Pan Right", "Pan Up",
                                                              "Pan Down"], state="readonly", width=15);
        self.zoom_direction.set("Zoom In");
        self.zoom_direction.grid(row=1, column=1, sticky='ew', padx=5)
        ttk.Label(zoom_grid, text="Speed:").grid(row=2, column=0, sticky='w', pady=2)
        self.zoom_speed = ttk.Combobox(zoom_grid, values=["Slow", "Medium", "Fast"], state="readonly",
                                       width=10)
        self.zoom_speed.set("Medium");
        self.zoom_speed.grid(row=2, column=1, sticky='ew', padx=5)
        ttk.Label(zoom_grid, text="Background Blur:").grid(row=3, column=0, sticky='w', pady=2)
        self.zoom_blur = ttk.Scale(zoom_grid, from_=20, to=200, orient='horizontal');
        self.zoom_blur.set(100);
        self.zoom_blur.grid(row=3, column=1, sticky='ew', padx=5)
        self.zoom_start_btn = ttk.Button(zoom_frame, text="Start Creation Zooms", command=self.add_zoom_task);
        self.zoom_start_btn.pack(pady=(5, 0))

    def _populate_order_tab(self, parent_frame):
        # Loading status
        status_frame = ttk.Frame(parent_frame)
        status_frame.pack(fill='x', pady=(0, 10))

        ttk.Label(status_frame, text="Data Source:").pack(side='left', padx=5)
        self.order_source_label = ttk.Label(status_frame, text="Not loaded", foreground="gray")
        self.order_source_label.pack(side='left', padx=5)

        # Order control buttons
        buttons_frame = ttk.Frame(parent_frame)
        buttons_frame.pack(fill='x', pady=5)

        self.order_up_btn = ttk.Button(buttons_frame, text="↑ Up",
                                         command=self._move_entity_up, width=12)
        self.order_up_btn.pack(side='left', padx=5)

        self.order_down_btn = ttk.Button(buttons_frame, text="↓ Down",
                                           command=self._move_entity_down, width=12)
        self.order_down_btn.pack(side='left', padx=5)

        self.order_save_btn = ttk.Button(buttons_frame, text="Save Order",
                                           command=self._save_order_to_json)
        self.order_save_btn.pack(side='left', padx=20)

        # Functionality info
        info_frame = ttk.LabelFrame(parent_frame, text="Information", padding=10)
        info_frame.pack(fill='both', expand=True, pady=10)

        info_text = scrolledtext.ScrolledText(info_frame, height=15, wrap='word',
                                              state='disabled', font=("Arial", 10))
        info_text.pack(fill='both', expand=True)

        info_text.configure(state='normal')
        info_text.insert('1.0', "")
        info_text.configure(state='disabled')

        # Buttons are initially disabled
        self._update_order_buttons_state()

    def _populate_video_tab(self, parent_frame):
        # 1. Create main notebook widget for sub-tabs
        video_sub_notebook = ttk.Notebook(parent_frame)
        self.video_sub_notebook = video_sub_notebook # <<< ADD THIS LINE
        video_sub_notebook.pack(fill='both', expand=True)

        # 2. Create frames for each sub-tab
        download_slice_frame = ttk.Frame(video_sub_notebook, padding=5)
        self.semantic_slice_frame = ttk.Frame(video_sub_notebook, padding=10)
        composite_frame = ttk.Frame(video_sub_notebook, padding=10)
        updater_frame = ttk.Frame(video_sub_notebook, padding=10)

        video_sub_notebook.add(download_slice_frame, text="1. Download and Slice")
        video_sub_notebook.add(self.semantic_slice_frame, text="2. Semantic Slicing")
        video_sub_notebook.add(composite_frame, text="3. Composite Video")
        video_sub_notebook.add(updater_frame, text="yt-dlp Update")

        # 3. Fill first sub-tab ("Download and Slice") with OLD content
        nav_panel = self._create_common_nav_panel(download_slice_frame, tab_type='video') # <-- Note the tab_type
        nav_panel.pack(fill='x', pady=(0, 5), anchor='w')

        video_links_frame = ttk.LabelFrame(download_slice_frame, text="Video Download", padding=5)
        video_links_frame.pack(fill='both', expand=True, pady=5)

        ttk.Label(video_links_frame, text="Regular videos:").pack(anchor='w')
        self.regular_video_links_text = scrolledtext.ScrolledText(video_links_frame, height=5, width=40)
        self.regular_video_links_text.pack(fill='both', expand=True, pady=(0, 5))

        ttk.Label(video_links_frame, text="Interviews:").pack(anchor='w', pady=(5, 0))
        self.interview_video_links_text = scrolledtext.ScrolledText(video_links_frame, height=5, width=40)
        self.interview_video_links_text.pack(fill='both', expand=True, pady=(0, 5))

        # Download buttons
        download_buttons_frame = ttk.Frame(video_links_frame)
        download_buttons_frame.pack(anchor='center', pady=(5, 0))

        self.add_video_task_btn = ttk.Button(download_buttons_frame, text="Download without slicing", command=self.add_video_task)
        self.add_video_task_btn.pack(side='left', padx=(0, 5))

        self.add_video_with_slice_task_btn = ttk.Button(download_buttons_frame, text="Download with slicing", command=self.add_video_with_slice_task)
        self.add_video_with_slice_task_btn.pack(side='left')

        slicer_frame = ttk.LabelFrame(download_slice_frame, text="Slice full video into parts", padding=5)
        slicer_frame.pack(fill='x', pady=5)

        slicer_params_frame = ttk.Frame(slicer_frame)
        slicer_params_frame.pack(fill='x')

        ttk.Label(slicer_params_frame, text="Duration (sec):").pack(side='left', padx=(0, 5))
        self.slice_duration = tk.DoubleVar(value=4.8)
        self.slicer_cut_duration = ttk.Spinbox(slicer_params_frame, from_=0.1, to=99.0, increment=0.1, textvariable=self.slice_duration, width=6, command=self._save_slicer_settings)
        self.slicer_cut_duration.pack(side='left')
        self.slicer_cut_duration.bind("<FocusOut>", self._save_slicer_settings)

        self.slicer_manual_start_btn = ttk.Button(slicer_frame, text="Slice", command=self.add_slicer_task)
        self.slicer_manual_start_btn.pack(pady=(5, 0))

        self._make_text_widget_context_menu()
        self.regular_video_links_text.bind("<Button-2>", self._show_context_menu)
        self.regular_video_links_text.bind("<Button-3>", self._show_context_menu)
        self.interview_video_links_text.bind("<Button-2>", self._show_context_menu)
        self.interview_video_links_text.bind("<Button-3>", self._show_context_menu)

        # 4. Fill second sub-tab ("Semantic Slicing") with NEW content
        self._populate_semantic_slice_tab(self.semantic_slice_frame)

        # 5. Fill third sub-tab ("Composite Video")
        self._populate_composite_tab(composite_frame)

        # 6. Fill fourth sub-tab ("Update yt-dlp")
        self._populate_updater_tab(updater_frame)

    def _create_common_nav_panel(self, parent, tab_type):
        nav_panel = ttk.Frame(parent)
        self.youtube_btn = ttk.Button(nav_panel, text="YouTube Search", command=self.open_youtube)
        self.youtube_btn.pack(side='left', padx=2)
        self.clips_btn = ttk.Button(nav_panel, text="Open Slices", command=self.open_slices_folder)
        self.clips_btn.pack(side='left', padx=2)

        if tab_type == 'photo':
            self.bridge_btn = ttk.Button(nav_panel, text="Open in Bridge", command=self.open_in_bridge)
            self.bridge_btn.pack(side='left', padx=2)

        # ONLY for video - adding visuals buttons (right)
        if tab_type == 'video':
            self.visuals_clear_btn = ttk.Button(nav_panel, text="x", width=3, command=self._clear_visuals)
            self.visuals_clear_btn.pack(side='right', padx=2)

            self.visuals_minus_btn = ttk.Button(nav_panel, text="-1", width=4, command=self._set_visuals_minus)
            self.visuals_minus_btn.pack(side='right', padx=2)

            self.visuals_plus_btn = ttk.Button(nav_panel, text="+1", width=4, command=self._set_visuals_plus)
            self.visuals_plus_btn.pack(side='right', padx=2)

        return nav_panel

    def _on_resize_columns(self, event):
        width = self.tree.winfo_width()
        if width <= 1 or width == self._last_resize_width: return
        self._last_resize_width = width
        for col_id, col_width, _ in self.column_configs: self.tree.column(col_id, width=col_width)
        hide_order = [('photos', 850), ('unchecked', 780), ('zoom', 700), ('h_zoom', 640), ('raw_video', 590),
                      ('clip', 540), ('h_clip', 490), ('projects', 430), ('eid', 300), ('q', 250)]
        for col_id, threshold in hide_order:
            if width < threshold: self.tree.column(col_id, width=0)
        total_visible_width = sum(self.tree.column(c, 'width') for c in self.tree['columns'])
        name_col_width = self.tree.column('name', 'width')
        if name_col_width > 0:
            extra_space = width - total_visible_width - 20
            if extra_space > 0: self.tree.column('name', width=name_col_width + extra_space)

    def _build_log_panels(self, parent):
        self.bottom_container = ttk.Frame(self);
        self.bottom_container.pack(fill='both', expand=True, side='top')
        self.bottom_container.rowconfigure(0, weight=1);
        self.bottom_container.columnconfigure(0, weight=1)
        self.photo_log_container = ttk.Frame(self.bottom_container);
        self.photo_log_container.grid(row=0, column=0, sticky='nsew')
        self.photo_log_container.rowconfigure(0, weight=1);
        self.photo_log_container.columnconfigure(0, weight=1);
        self.photo_log_container.columnconfigure(1, weight=1)
        self.video_log_container = ttk.Frame(self.bottom_container);
        self.video_log_container.grid(row=0, column=0, sticky='nsew')
        self.video_log_container.rowconfigure(0, weight=1);
        self.video_log_container.columnconfigure(0, weight=1);
        self.video_log_container.columnconfigure(1, weight=1)

        def create_log_panel(p, text, row, col, padx, pady):
            frame = ttk.LabelFrame(p, text=text);
            frame.grid(row=row, column=col, rowspan=2, sticky='nsew', padx=padx, pady=pady)
            log_widget = scrolledtext.ScrolledText(frame, height=5, wrap='word');
            log_widget.pack(fill='both', expand=True, padx=5, pady=5)
            log_widget.configure(state='normal');
            # --- CHANGE: Removed line blocking copying ---
            # log_widget.bind("<KeyPress>", lambda e: "break")
            return log_widget

        self.photo_dl_log = create_log_panel(self.photo_log_container, "Photo Download Log", 0, 0, (0, 5), (0, 0))
        self.zoom_log = create_log_panel(self.photo_log_container, "Zoom Creation Log (from Photos)", 0, 1, (5, 0), (0, 0))
        self.video_dl_log = create_log_panel(self.video_log_container, "Video Download / Sem. Slicing Log", 0, 0, (0, 5), (0, 0))
        self.slicer_log = create_log_panel(self.video_log_container, "Video Slicing Log", 0, 1, (5, 0), (0, 0))
        self.photo_log_container.tkraise()

    def _on_tab_changed(self, event):
        selected_tab = self.action_panel.select()
        if selected_tab == str(self.photo_tab_frame):
            self.photo_log_container.tkraise()
        elif selected_tab == str(self.video_tab_frame):
            self.video_log_container.tkraise()

    def _make_text_widget_context_menu(self):
        self.context_menu = tk.Menu(self, tearoff=0)
        self.context_menu.add_command(label="Cut", command=lambda: self.focus_get().event_generate('<<Cut>>'))
        self.context_menu.add_command(label="Copy", command=lambda: self.focus_get().event_generate('<<Copy>>'))
        self.context_menu.add_command(label="Paste", command=lambda: self.focus_get().event_generate('<<Paste>>'))
        self.context_menu.add_separator()
        self.context_menu.add_command(label="Select All", command=self._select_all_in_focused)

    def _select_all_in_focused(self, event=None):
        widget = self.focus_get();
        if isinstance(widget, (tk.Text, scrolledtext.ScrolledText)): widget.tag_add('sel', '1.0', 'end-1c')
        return "break"

    def _show_context_menu(self, event):
        widget = event.widget
        try:
            selection = widget.get(tk.SEL_FIRST, tk.SEL_LAST);
            self.context_menu.entryconfig("Cut", state="normal");
            self.context_menu.entryconfig("Copy", state="normal")
        except tk.TclError:
            self.context_menu.entryconfig("Cut", state="disabled");
            self.context_menu.entryconfig("Copy", state="disabled")
        try:
            self.clipboard_get();
            self.context_menu.entryconfig("Paste", state="normal")
        except tk.TclError:
            self.context_menu.entryconfig("Paste", state="disabled")
        self.context_menu.tk_popup(event.x_root, event.y_root)

    def log_to_panel(self, panel, log_key, message):
        def _log(): panel.insert(tk.END, message + '\n'); panel.see(tk.END)

        self.app.root.after(0, _log)
        if hasattr(self.app, 'write_to_log_file'): self.app.write_to_log_file(log_key, message)

    def check_folder_before_processing(self, entity_path: Path, target_folder_name: str, entity_name: str, log_panel,
                                       log_key) -> str:
        target_path = entity_path / target_folder_name
        if not target_path.exists() or not any(target_path.iterdir()): return "append"
        self.log_to_panel(log_panel, log_key,
                          f"⚠️ Folder '{target_folder_name}' for '{entity_name}' is not empty. Waiting for decision...")
        dialog = OverwriteDialog(self, target_path, entity_name)
        decision = dialog.result
        if decision == "skip":
            self.log_to_panel(log_panel, log_key, f"⏹️ User chose to skip task for '{entity_name}'.")
        return decision

    def _check_and_prepare_interview_folder(self, entity_path: Path, entity_name: str) -> str:
        target_path = entity_path / "interview_fragments"
        if not target_path.exists():
            target_path.mkdir(parents=True, exist_ok=True)
            (target_path / "transcriptions").mkdir(exist_ok=True)
            return "proceed"

        highlight_pattern = re.compile(r'^h\d*_.*')
        non_highlight_files = [f for f in target_path.iterdir() if f.is_file() and not highlight_pattern.match(f.name)]

        if not non_highlight_files:
            (target_path / "transcriptions").mkdir(exist_ok=True)
            return "proceed"

        dialog = OverwriteInterviewDialog(self, target_path, entity_name)
        decision = dialog.result

        if decision == "overwrite":
            self.log_to_panel(self.video_dl_log, 'video_dl', f"⚠️ Clearing 'interview_fragments' for '{entity_name}' ('h...' files preserved)...")
            for file_to_delete in non_highlight_files:
                file_to_delete.unlink()
        elif decision == "skip":
            self.log_to_panel(self.video_dl_log, 'video_dl', f"⏹️ User skipped semantic slicing for '{entity_name}'.")
            return "skip"

        (target_path / "transcriptions").mkdir(exist_ok=True)
        return "proceed"

    def refresh_current_list_stats(self):
        all_items = self.tree.get_children()
        if not all_items:
            self.scan_status_label.config(text="Table is empty. Nothing to update.")
            return

        # Collect data from tags: (eid, role, ssid, song_name)
        tasks_to_refresh = []
        for item_id in all_items:
            tags = self.tree.item(item_id, 'tags')
            if tags and len(tags) > 3:
                tasks_to_refresh.append({
                    'eid': tags[0],
                    'role': tags[1],
                    'ssid': tags[2],
                    'song_name': tags[3],
                    'name': self.tree.item(item_id, 'values')[0] # Artist Name
                })

        if not tasks_to_refresh:
            self.scan_status_label.config(text="No data to update.")
            return

        self.refresh_button.config(state='disabled')
        self.scan_status_label.config(text=f"Updating {len(tasks_to_refresh)} rows...")

        if self.refresh_thread and self.refresh_thread.is_alive():
            self.scan_status_label.config(text="Update already running.")
            self.refresh_button.config(state='normal')
            return

        self.refresh_thread = threading.Thread(target=self._run_stats_refresh, args=(tasks_to_refresh,), daemon=True)
        self.refresh_thread.start()

    def _run_stats_refresh(self, tasks_to_refresh):
        """Worker function to run in a separate thread."""
        total = len(tasks_to_refresh)
        archive_path = Path(self.app.get_setting('media_archive_path', ""))
        
        def sanitize(t): return t.lower().replace(' ', '_')
        def clean(t): return re.sub(r'[^a-z0-9_-]', '', sanitize(str(t)))

        for i, task in enumerate(tasks_to_refresh, 1):
            if self.stop_event.is_set():
                self.app.root.after(0, self.scan_status_label.config, {'text': 'Update interrupted.'})
                break

            ssid = task['ssid']
            self.app.root.after(0, self.scan_status_label.config, {'text': f'Updating... {i}/{total} ({ssid})'})

            try:
                artist_folder = f"{clean(task['name'])}_{task['eid']}"
                song_folder = f"{ssid}_{clean(task['song_name'])}_by_{clean(task['name'])}_{task['eid']}"
                raw_path = archive_path / artist_folder / song_folder / "raw_videos"
                slices_path = archive_path / artist_folder / song_folder / "slices"

                clip_count = 0
                if raw_path.exists():
                    clip_count = len([f for f in raw_path.iterdir() if f.is_file() and not f.name.startswith('.')])

                slices_count = 0
                if slices_path.exists():
                    slices_count = len([f for f in slices_path.iterdir() if f.is_file() and f.suffix == '.mp4'])

                self.app.root.after(0, self.update_treeview_row, ssid, {'clip': clip_count, 'slices': slices_count})
            except Exception as e:
                print(f"Error refreshing {ssid}: {e}")

        def on_complete():
            self.scan_status_label.config(text="Statistics updated.")
            self.refresh_button.config(state='normal')
            self.app.root.after(3000, lambda: self.scan_status_label.config(text=""))

        self.app.root.after(0, on_complete)

    def add_photo_task(self, task_info: dict = None, manual_query: str = None):
        if not task_info:
            if not self.active_entity_data: return
            task_info = self.active_entity_data
        name = task_info.get('name');
        eid = task_info.get('eid');
        role = task_info.get('role')
        # For photos use only artist folder (ssid=None)
        entity_path, _ = entity_manager_v2.get_or_create_entity_path(Path(self.app.get_setting('media_archive_path', "")), eid, name, role)
        if not entity_path:
            self.log_to_panel(self.photo_dl_log, 'photo_dl', f"Could not get path for {name}. Task canceled.");
            return
        (entity_path / "photos").mkdir(exist_ok=True)
        decision = self.check_folder_before_processing(entity_path, "unchecked_photos", name, self.photo_dl_log,
                                                       'photo_dl')
        if decision == "skip": return
        try:
            count = int(self.photo_count_spinbox.get());
            assert count > 0
        except (ValueError, AssertionError):
            self.app.log_message_to_selection_tab("Error: Photo count must be a number > 0.");
            return
        task = {'eid': eid, 'name': name, 'role': role, 'count': count, 'overwrite_decision': decision}
        if manual_query: task['manual_query'] = manual_query
        self.photo_dl_queue.put(task)
        log_msg_extra = f" (manual query: '{manual_query}')" if manual_query else ""
        self.log_to_panel(self.photo_dl_log, 'photo_dl',
                          f"Task added for: {name}{log_msg_extra}. In queue: {self.photo_dl_queue.qsize()}")
        self.ensure_worker_running('photo')

    def add_batch_photo_task(self):
        tasks_to_add = []
        all_items = self.tree.get_children()
        if not all_items:
            self.app.log_message_to_selection_tab("No entities in table for batch processing.");
            return
        for item_id in all_items:
            values = self.tree.item(item_id, 'values');
            tags = self.tree.item(item_id, 'tags')
            is_unchecked_empty = values[9] == '0';
            is_photos_empty = values[10] == '0'
            if is_unchecked_empty and is_photos_empty:
                task_info = {'name': values[0], 'eid': tags[0], 'role': tags[1]};
                tasks_to_add.append(task_info)
        if not tasks_to_add:
            self.log_to_panel(self.photo_dl_log, 'photo_dl', "No entities without photos found for download.");
            return
        self.log_to_panel(self.photo_dl_log, 'photo_dl',
                          f"Found {len(tasks_to_add)} entities for batch download. Adding to queue...")
        for task_info in tasks_to_add:
            self.add_photo_task(task_info=task_info)

    def add_manual_query_photo_task(self):
        if not self.active_entity_data:
            self.app.log_message_to_selection_tab("First select an entity in the table.");
            return
        query = simpledialog.askstring("Manual query",
                                       f"Enter search query for photo '{self.active_entity_data.get('name')}':",
                                       parent=self)
        if query and query.strip():
            self.add_photo_task(manual_query=query.strip())
        else:
            self.log_to_panel(self.photo_dl_log, 'photo_dl', "Canceled: manual query not entered.")

    def add_video_task(self):
        # Automatic daily check for yt-dlp
        self._run_daily_check_if_needed()

        if not self.active_entity_data:
            self.app.log_message_to_selection_tab("Error: First select an entity.")
            return

        # Check interview_fragments folder if semantic slicing is active
        is_semantic_active = self.semantic_active_var.get()
        interview_text_content = self.interview_video_links_text.get('1.0', tk.END).strip()

        if is_semantic_active and interview_text_content:
            name = self.active_entity_data.get('name')
            eid = self.active_entity_data.get('eid')
            role = self.active_entity_data.get('role')
            ssid = self.active_entity_data.get('ssid')
            song_name = self.active_entity_data.get('song_name')
            
            # For interview-slicing we need artist path (where interview_fragments are located)
            entity_path, _ = entity_manager_v2.get_or_create_entity_path(Path(self.app.get_setting('media_archive_path', "")), eid, name, role)

            if not entity_path:
                self.log_to_panel(self.video_dl_log, 'video_dl', f"❌ Could not get path for {name}. Task canceled.")
                return

            decision = self._check_and_prepare_interview_folder(entity_path, name)
            if decision == "skip":
                return  # Completely cancel task if user decided so

        regular_text = self.regular_video_links_text.get('1.0', tk.END).strip()
        interview_text = self.interview_video_links_text.get('1.0', tk.END).strip()

        if not regular_text and not interview_text:
            self.app.log_message_to_selection_tab("Error: Enter video links.")
            return

        # 1. Individual deduplication
        unique_regular = sorted(list(set(["https" + s.strip() for s in regular_text.split("https") if s.strip()])))
        unique_interview = sorted(list(set(["https" + s.strip() for s in interview_text.split("https") if s.strip()])))

        # 2. Applying "Priority Rule"
        final_regular_links = [link for link in unique_regular if link not in unique_interview]
        final_interview_links = unique_interview # All interviews remain

        all_links_for_metadata = final_regular_links + final_interview_links
        if not all_links_for_metadata:
            self.app.log_message_to_selection_tab("Error: No valid links found.")
            return

        links_map = {'regular': final_regular_links, 'interview': final_interview_links}

        if self.video_check_thread and self.video_check_thread.is_alive():
            self.log_to_panel(self.video_dl_log, 'video_dl', "Please wait, previous check is not yet complete.")
            return

        self.add_video_task_btn.config(state='disabled')
        self.log_to_panel(self.video_dl_log, 'video_dl',
                          f"Starting processing of {len(all_links_for_metadata)} unique links...")

        entity_data_copy = self.active_entity_data.copy()

        self.video_check_thread = threading.Thread(
            target=self._add_video_task_worker,
            args=(all_links_for_metadata, entity_data_copy, links_map),
            daemon=True
        )
        self.video_check_thread.start()

    def add_video_with_slice_task(self):
        """Download video AND slice into parts after download."""
        # Automatic daily check for yt-dlp
        self._run_daily_check_if_needed()

        if not self.active_entity_data:
            self.app.log_message_to_selection_tab("Error: First select an entity.")
            return

        # Check slices folder BEFORE queuing task
        name = self.active_entity_data.get('name')
        eid = self.active_entity_data.get('eid')
        role = self.active_entity_data.get('role')
        ssid = self.active_entity_data.get('ssid')
        song_name = self.active_entity_data.get('song_name')

        entity_path, _ = entity_manager_v2.get_or_create_entity_path(
            Path(self.app.get_setting('media_archive_path', "")), eid, name, role, ssid=ssid, song_name=song_name
        )

        if not entity_path:
            self.log_to_panel(self.video_dl_log, 'video_dl', f"❌ Could not get path for {name}. Task canceled.")
            return

        # Check slices folder
        slices_path = entity_path / "slices"
        if slices_path.exists() and any(slices_path.iterdir()):
            # Folder is not empty - ask user
            decision = self._check_slices_folder_dialog(slices_path, name)
            if decision == "skip":
                return

        # Check interview_fragments folder if semantic slicing is active
        is_semantic_active = self.semantic_active_var.get()
        interview_text_content = self.interview_video_links_text.get('1.0', tk.END).strip()

        if is_semantic_active and interview_text_content:
            # For interview-slicing we need artist path (where interview_fragments are located)
            entity_path_artist, _ = entity_manager_v2.get_or_create_entity_path(
                Path(self.app.get_setting('media_archive_path', "")), eid, name, role
            )

            if not entity_path_artist:
                self.log_to_panel(self.video_dl_log, 'video_dl', f"❌ Could not get path for {name}. Task canceled.")
                return

            decision = self._check_and_prepare_interview_folder(entity_path_artist, name)
            if decision == "skip":
                return

        # 2. Applying "Priority Rule"
        final_regular_links = [link for link in unique_regular if link not in unique_interview]
        final_interview_links = unique_interview # All interviews remain

        all_links_for_metadata = final_regular_links + final_interview_links
        if not all_links_for_metadata:
            self.app.log_message_to_selection_tab("Error: No valid links found.")
            return

        links_map = {'regular': final_regular_links, 'interview': final_interview_links}

        if self.video_check_thread and self.video_check_thread.is_alive():
            self.log_to_panel(self.video_dl_log, 'video_dl', "Please wait, previous check is not yet complete.")
            return

        self.add_video_with_slice_task_btn.config(state='disabled')
        self.log_to_panel(self.video_dl_log, 'video_dl',
                          f"Starting processing of {len(all_links_for_metadata)} unique links (with slicing)...")

        entity_data_copy = self.active_entity_data.copy()

        self.video_check_thread = threading.Thread(
            target=self._add_video_task_worker,
            args=(all_links_for_metadata, entity_data_copy, links_map, True),  # True = with_slicing
            daemon=True
        )
        self.video_check_thread.start()

    def _check_slices_folder_dialog(self, slices_path, entity_name):
        """Shows a dialog if the slices folder is not empty."""
        dialog = tk.Toplevel(self)
        dialog.transient(self)
        dialog.title("Slices folder not empty")
        dialog.result = "skip"
        message = f"Folder 'slices' for '{entity_name}' already contains files.\n\nSelect action:"
        ttk.Label(dialog, text=message, justify='left').pack(padx=20, pady=(20, 10))
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(padx=20, pady=(10, 20))
        ttk.Button(btn_frame, text="Overwrite (delete old)", command=lambda: self._set_dialog_result(dialog, "overwrite")).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="Add (keep old)", command=lambda: self._set_dialog_result(dialog, "append")).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="Skip task", command=lambda: self._set_dialog_result(dialog, "skip")).pack(side='left', padx=5)
        dialog.protocol("WM_DELETE_WINDOW", lambda: self._set_dialog_result(dialog, "skip"))
        dialog.grab_set()
        dialog.wait_window()
        return dialog.result

    def _set_dialog_result(self, dialog, result):
        dialog.result = result
        dialog.destroy()

    def _add_video_task_worker(self, links, entity_data, links_map, with_slicing=False):
        try:
            name = entity_data.get('name')
            eid = entity_data.get('eid')
            role = entity_data.get('role')
            ssid = entity_data.get('ssid')
            song_name = entity_data.get('song_name')

            # For video use SSID to get song folder
            entity_path, _ = entity_manager_v2.get_or_create_entity_path(Path(self.app.get_setting('media_archive_path', "")), eid, name, role, ssid=ssid, song_name=song_name)
            if not entity_path:
                self.app.root.after(0, self.log_to_panel, self.video_dl_log, 'video_dl', f"Could not get path for {name}. Task canceled.")
                return

            try:
                self.app.root.after(0, self.log_to_panel, self.video_dl_log, 'video_dl', "Getting video metadata...")
                metadata_list = video_parser_v2.get_videos_metadata(links, lambda msg: self.log_to_panel(self.video_dl_log, 'video_dl', msg))
                self.app.root.after(0, self.log_to_panel, self.video_dl_log, 'video_dl', "Metadata successfully retrieved.")
            except Exception as e:
                log_msg = f"CRITICAL ERROR retrieving metadata: {e}. Ensure yt-dlp is installed in the project environment."
                self.app.root.after(0, self.log_to_panel, self.video_dl_log, 'video_dl', log_msg)
                return

            self.app.root.after(0, self._process_metadata_and_queue_task, metadata_list, entity_data, entity_path, links_map, with_slicing)

        except Exception as e:
            error_message = f"Critical error in background video handler: {e}"
            self.app.root.after(0, self.log_to_panel, self.video_dl_log, 'video_dl', error_message)
        finally:
            self.app.root.after(0, self.add_video_task_btn.config, {'state': 'normal'})
            if hasattr(self, 'add_video_with_slice_task_btn'):
                self.app.root.after(0, self.add_video_with_slice_task_btn.config, {'state': 'normal'})

    def _process_metadata_and_queue_task(self, metadata_list, entity_data, entity_path, links_map, with_slicing=False):
        name = entity_data.get('name')
        eid = entity_data.get('eid')
        role = entity_data.get('role')
        ssid = entity_data.get('ssid')
        song_name = entity_data.get('song_name')

        long_videos = []
        short_videos_links = []
        unavailable_links = []

        for meta in metadata_list:
            if meta['error']:
                unavailable_links.append(meta['url'])
                self.log_to_panel(self.video_dl_log, 'video_dl',
                                  f"⚠️ Verification error: {meta['error']} for {meta['url']}")
            elif meta['duration'] and meta['duration'] > 1800:
                long_videos.append(meta)
            else:
                short_videos_links.append(meta['url'])

        if unavailable_links:
            self.app.log_message_to_selection_tab(f"Could not verify {len(unavailable_links)} links. See log.")

        final_links_to_process = []
        if long_videos:
            dialog = DurationCheckDialog(self, long_videos)
            decision = dialog.result
            if decision == 'all':
                final_links_to_process = short_videos_links + [v['url'] for v in long_videos]
            elif decision == 'only_short':
                final_links_to_process = short_videos_links
            else:
                self.log_to_panel(self.video_dl_log, 'video_dl', "Operation canceled by user for list editing.")
                updated_text = ""
                for meta in metadata_list:
                    if meta['error']: updated_text += f"{meta['url']} [ERROR: {meta['error']}]\n"
                    elif meta['duration'] > 1800:
                        duration_str = video_parser_v2.format_duration(meta['duration'])
                        updated_text += f"{meta['url']} [Duration: {duration_str}]\n"
                    else: updated_text += f"{meta['url']}\n"
                self.regular_video_links_text.delete('1.0', tk.END)
                self.regular_video_links_text.insert('1.0', updated_text)
                self.interview_video_links_text.delete('1.0', tk.END)
                return
        else:
            final_links_to_process = short_videos_links

        if not final_links_to_process:
            self.log_to_panel(self.video_dl_log, 'video_dl', "No suitable videos to add to the queue.")
            return

        decision = self.check_folder_before_processing(entity_path, "raw_videos", name, self.video_dl_log, 'video_dl')
        if decision == "skip":
            return

        final_regular_links = [link for link in final_links_to_process if link in links_map['regular']]
        final_interview_links = [link for link in final_links_to_process if link in links_map['interview']]

        task = {'eid': eid, 'name': name, 'role': role,
                'ssid': ssid, 'song_name': song_name,
                'regular_links': final_regular_links,
                'interview_links': final_interview_links,
                'overwrite_decision': decision,
                'with_slicing': with_slicing}
        print(f"DEBUG: Putting task into queue: {task}")
        self.video_dl_queue.put(task)
        self.log_to_panel(self.video_dl_log, 'video_dl',
                           f"Task added for: {name} ({len(final_links_to_process)} videos). In queue: {self.video_dl_queue.qsize()}")
        self.regular_video_links_text.delete('1.0', tk.END)
        self.interview_video_links_text.delete('1.0', tk.END)
        self.ensure_worker_running('video')

    def add_slicer_task(self):
        """Slice all golden masters from raw_videos into slices."""
        if not self.active_entity_data: return
        name = self.active_entity_data.get('name')
        eid = self.active_entity_data.get('eid')
        role = self.active_entity_data.get('role')
        ssid = self.active_entity_data.get('ssid')
        song_name = self.active_entity_data.get('song_name')

        # Working with SONG folder
        entity_path, _ = entity_manager_v2.get_or_create_entity_path(
            Path(self.app.get_setting('media_archive_path', "")), eid, name, role, ssid=ssid, song_name=song_name
        )
        if not entity_path: return

        # Check slices folder BEFORE queuing task
        slices_path = entity_path / "slices"
        if slices_path.exists() and any(slices_path.iterdir()):
            decision = self._check_slices_folder_dialog(slices_path, name)
            if decision == "skip":
                return
            if decision == "overwrite":
                shutil.rmtree(slices_path)
                slices_path.mkdir(exist_ok=True)
        else:
            slices_path.mkdir(exist_ok=True)

        # Checking that there are files to slice
        raw_videos_path = entity_path / "raw_videos"
        if not raw_videos_path.exists():
            self.log_to_panel(self.slicer_log, 'slicer', f"❌ raw_videos folder not found for {name}.")
            return

        source_files = list(raw_videos_path.glob('raw_video_*.mp4'))
        if not source_files:
            self.log_to_panel(self.slicer_log, 'slicer', f"❌ No files in raw_videos for {name}.")
            return

        # Form task for slicing
        try:
            slice_duration = float(self.slice_duration.get())
        except (ValueError, TypeError):
            self.log_to_panel(self.slicer_log, 'slicer', f"❌ Incorrect duration value: {self.slice_duration.get()}")
            return

        task = {
            'eid': eid, 'name': name, 'role': role, 'ssid': ssid, 'song_name': song_name,
            'slice_duration': slice_duration,
            'entity_path': entity_path,
            'downloaded_files': None  # None means slice all files from raw_videos
        }

        self.slices_queue.put(task)
        self.log_to_panel(self.slicer_log, 'slicer',
                          f"Task added to slice for parts for: {name} ({len(source_files)} files, duration: {slice_duration} sec). "
                          f"In queue: {self.slices_queue.qsize()}")
        self.ensure_worker_running('slices')
        self.ensure_worker_running('slicer')

    def add_zoom_task(self):
        if not self.active_entity_data: return
        name = self.active_entity_data.get('name');
        eid = self.active_entity_data.get('eid');
        role = self.active_entity_data.get('role')
        # Zooms are made from artist photos (ssid=None)
        entity_path, _ = entity_manager_v2.get_or_create_entity_path(Path(self.app.get_setting('media_archive_path', "")), eid, name, role)
        if not entity_path: return
        decision = self.check_folder_before_processing(entity_path, "zoomed_videos", name, self.zoom_log, 'zoom')
        if decision == "skip": return
        speed_map = {"Slow": 0.7, "Medium": 1.0, "Fast": 1.5}
        settings = {'duration': int(self.zoom_duration.get()), 'effect': self.zoom_direction.get(),
                    'effect_speed': speed_map.get(self.zoom_speed.get(), 1.0), 'blur_radius': int(self.zoom_blur.get())}
        task = {'eid': eid, 'name': name, 'role': role, 'settings': settings, 'overwrite_decision': decision}
        self.zoom_queue.put(task)
        self.log_to_panel(self.zoom_log, 'zoom',
                          f"Task added to create zooms for: {name}. In queue: {self.zoom_queue.qsize()}")
        self.ensure_worker_running('zoom')

    def add_composite_task(self):
        """Adds task to create composite video (Sandwich)."""
        if not self.active_entity_data:
            return

        name = self.active_entity_data.get('name')
        eid = self.active_entity_data.get('eid')
        role = self.active_entity_data.get('role')
        ssid = self.active_entity_data.get('ssid')
        song_name = self.active_entity_data.get('song_name', '')

        # Get entity paths
        entity_path, _ = entity_manager_v2.get_or_create_entity_path(
            Path(self.app.get_setting('media_archive_path', "")), eid, name, role, ssid=ssid, song_name=song_name
        )
        if not entity_path:
            self.log_to_panel(self.video_dl_log, 'video_dl', "❌ Could not get entity path.")
            return

        # Parse links from text fields
        import video_song_sandwich_worker

        audio_text = self.composite_audio_text.get("1.0", tk.END).strip()
        video_text = self.composite_video_text.get("1.0", tk.END).strip()

        audio_urls = video_song_sandwich_worker.parse_urls_from_text(audio_text)
        video_urls = video_song_sandwich_worker.parse_urls_from_text(video_text)

        # Validation
        if not audio_urls:
            self.log_to_panel(self.video_dl_log, 'video_dl', "❌ AUDIO link not specified!")
            return

        if not video_urls:
            self.log_to_panel(self.video_dl_log, 'video_dl', "❌ VIDEO links not specified!")
            return

        # Take first audio link
        audio_url = audio_urls[0]

        # Check raw_videos folder
        decision = self.check_folder_before_processing(entity_path, "raw_videos", name, self.video_dl_log, 'video_dl')
        if decision == "skip":
            return

        # Form task
        task = {
            'type': 'composite',
            'eid': eid,
            'name': name,
            'role': role,
            'ssid': ssid,
            'song_name': song_name,
            'audio_url': audio_url,
            'video_urls': video_urls,
            'overwrite_decision': decision
        }

        self.composite_queue.put(task)
        self.log_to_panel(self.video_dl_log, 'video_dl',
                          f"Task 'Composite Video' added for: {name}. Audio: 1, Video: {len(video_urls)}. "
                          f"In queue: {self.composite_queue.qsize()}")
        self.ensure_worker_running('composite')

    def add_composite_with_slice_task(self):
        """Adds task to create composite video (Sandwich) WITH SLICING into parts."""
        if not self.active_entity_data:
            return

        name = self.active_entity_data.get('name')
        eid = self.active_entity_data.get('eid')
        role = self.active_entity_data.get('role')
        ssid = self.active_entity_data.get('ssid')
        song_name = self.active_entity_data.get('song_name', '')

        # Get entity paths
        entity_path, _ = entity_manager_v2.get_or_create_entity_path(
            Path(self.app.get_setting('media_archive_path', "")), eid, name, role, ssid=ssid, song_name=song_name
        )
        if not entity_path:
            self.log_to_panel(self.video_dl_log, 'video_dl', "❌ Could not get entity path.")
            return

        # Check slices folder BEFORE queuing task
        slices_path = entity_path / "slices"
        if slices_path.exists() and any(slices_path.iterdir()):
            decision = self._check_slices_folder_dialog(slices_path, name)
            if decision == "skip":
                return

        # Parse links from text fields
        import video_song_sandwich_worker

        audio_text = self.composite_audio_text.get("1.0", tk.END).strip()
        video_text = self.composite_video_text.get("1.0", tk.END).strip()

        audio_urls = video_song_sandwich_worker.parse_urls_from_text(audio_text)
        video_urls = video_song_sandwich_worker.parse_urls_from_text(video_text)

        # Validation
        if not audio_urls:
            self.log_to_panel(self.video_dl_log, 'video_dl', "❌ AUDIO link not specified!")
            return

        if not video_urls:
            self.log_to_panel(self.video_dl_log, 'video_dl', "❌ VIDEO links not specified!")
            return

        # Take first audio link
        audio_url = audio_urls[0]

        # Check raw_videos folder
        decision = self.check_folder_before_processing(entity_path, "raw_videos", name, self.video_dl_log, 'video_dl')
        if decision == "skip":
            return

        # Form task
        try:
            slice_duration = float(self.slice_duration.get())
        except (ValueError, TypeError):
            self.log_to_panel(self.video_dl_log, 'video_dl', f"❌ Incorrect duration value: {self.slice_duration.get()}")
            return

        task = {
            'type': 'composite',
            'eid': eid,
            'name': name,
            'role': role,
            'ssid': ssid,
            'song_name': song_name,
            'audio_url': audio_url,
            'video_urls': video_urls,
            'overwrite_decision': decision,
            'with_slicing': True,
            'is_sandwich': True,
            'slice_duration': slice_duration
        }

        self.composite_queue.put(task)
        self.log_to_panel(self.video_dl_log, 'video_dl',
                          f"Task 'Composite Video with slicing' added for: {name}. Audio: 1, Video: {len(video_urls)}. "
                          f"In queue: {self.composite_queue.qsize()}")
        self.ensure_worker_running('composite')

    def _load_semantic_settings(self):
        """Loads settings from config.json or sets default values."""
        self.semantic_settings = {}
        if not self.app.CONFIG_FILE_PATH:
            # If config path does not exist, exit silently.
            # Settings will load during next update once path is set.
            return

        self.semantic_settings = {
            "is_active": False,
            "slicing_type": "full",
            "language": "en",
            "whisper_model": "small",
            "min_speakers": 2,
            "max_speakers": 2,
            "min_clip_duration": "6.0",
            "max_clip_duration": "12.0",
            "video_crf": 23
        }
        try:
            with open(self.app.CONFIG_FILE_PATH, 'r', encoding='utf-8') as f:
                config_data = json.load(f)
            if "semantic_slicer_settings" in config_data:
                self.semantic_settings.update(config_data["semantic_slicer_settings"])
        except (FileNotFoundError, json.JSONDecodeError, KeyError) as e:
            print(f"Info: Could not load semantic slicer settings, using defaults. Reason: {e}")

    def _save_semantic_settings(self, *args):
        """Collects current values from widgets and saves them to config.json."""
        if not self.app.CONFIG_FILE_PATH:
            return
        try:
            settings_to_save = {
                "is_active": self.semantic_active_var.get(),
                "slicing_type": self.slicing_type_var.get(),
                "language": self.language_combo.get(),
                "whisper_model": self.whisper_model_combo.get(),
                "min_speakers": int(self.min_speakers_spinbox.get()),
                "max_speakers": int(self.max_speakers_spinbox.get()),
                "min_clip_duration": self.min_duration_entry.get(),
                "max_clip_duration": self.max_duration_entry.get(),
                "video_crf": int(self.crf_spinbox.get())
            }
        except AttributeError:
            return

        try:
            with open(self.app.CONFIG_FILE_PATH, 'r', encoding='utf-8') as f:
                config_data = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError):
            config_data = {}

        config_data["semantic_slicer_settings"] = settings_to_save

        with open(self.app.CONFIG_FILE_PATH, 'w', encoding='utf-8') as f:
            json.dump(config_data, f, indent=4, ensure_ascii=False)

    def _load_slicer_settings(self):
        """Loads slicer settings from config.json or sets default values."""
        default_duration = 4.8
        if not self.app.CONFIG_FILE_PATH:
            if hasattr(self, 'slice_duration'):
                self.slice_duration.set(default_duration)
            return

        try:
            with open(self.app.CONFIG_FILE_PATH, 'r', encoding='utf-8') as f:
                config_data = json.load(f)
            if "slicer_settings" in config_data and "slice_duration" in config_data["slicer_settings"]:
                saved_duration = config_data["slicer_settings"]["slice_duration"]
                if hasattr(self, 'slice_duration'):
                    self.slice_duration.set(float(saved_duration))
            else:
                if hasattr(self, 'slice_duration'):
                    self.slice_duration.set(default_duration)
        except (FileNotFoundError, json.JSONDecodeError, KeyError) as e:
            if hasattr(self, 'slice_duration'):
                self.slice_duration.set(default_duration)

    def _save_slicer_settings(self, *args):
        """Saves current slice duration value to config.json."""
        if not self.app.CONFIG_FILE_PATH or not hasattr(self, 'slice_duration'):
            return

        try:
            current_duration = float(self.slice_duration.get())
        except (ValueError, TypeError):
            return

        try:
            with open(self.app.CONFIG_FILE_PATH, 'r', encoding='utf-8') as f:
                config_data = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError):
            config_data = {}

        if "slicer_settings" not in config_data:
            config_data["slicer_settings"] = {}
        config_data["slicer_settings"]["slice_duration"] = current_duration

        with open(self.app.CONFIG_FILE_PATH, 'w', encoding='utf-8') as f:
            json.dump(config_data, f, indent=4, ensure_ascii=False)

    def _check_yt_dlp_versions(self):
        """Checks yt-dlp versions in a separate thread."""
        threading.Thread(target=self._check_yt_dlp_versions_worker, daemon=True).start()

    def _check_yt_dlp_versions_worker(self):
        """Worker function to check yt-dlp versions."""
        def log(message):
            self.app.root.after(0, lambda: self.updater_log.insert(tk.END, message + '\n'))
            self.app.root.after(0, lambda: self.updater_log.see(tk.END))

        log("▶️ STEP 1: Determining path to Python interpreter...")
        python_path = sys.executable
        log(f"Python path: {python_path}")

        log("▶️ STEP 2: Getting current yt-dlp version...")
        try:
            result = subprocess.run([python_path, '-m', 'pip', 'show', 'yt-dlp'],
                                  capture_output=True, text=True, encoding='utf-8')
            if result.returncode == 0:
                lines = result.stdout.split('\n')
                version_line = next((line for line in lines if line.startswith('Version:')), None)
                location_line = next((line for line in lines if line.startswith('Location:')), None)
                if version_line:
                    current_version = version_line.split(': ', 1)[1].strip()
                    log(f"Current yt-dlp version: {current_version}")
                else:
                    log("❌ Could not determine yt-dlp version")
                    return
                if location_line:
                    location = location_line.split(': ', 1)[1].strip()
                    log(f"Location: {location}")
            else:
                log(f"❌ Error getting yt-dlp info: {result.stderr}")
                return
        except Exception as e:
            log(f"❌ Error running pip show: {e}")
            return

        log("▶️ STEP 3: Getting latest version from PyPI...")
        try:
            with urllib.request.urlopen('https://pypi.org/pypi/yt-dlp/json', timeout=10) as response:
                data = json.loads(response.read().decode('utf-8'))
                latest_version = data['info']['version']
                log(f"Latest version on PyPI: {latest_version}")
        except Exception as e:
            log(f"❌ Error getting data from PyPI: {e}")
            return

        log("▶️ STEP 4: Comparing versions...")
        if current_version == latest_version:
            status = 'ok'
            log("✅ Latest yt-dlp version installed")
        else:
            status = 'update_available'
            log(f"❗ Update available: {current_version} → {latest_version}")

        # Save results
        today = datetime.date.today().isoformat()
        updates = {
            'last_yt_dlp_check_date': today,
            'yt_dlp_status': status
        }
        self._save_yt_dlp_config(updates)

        # Update tab status
        self.app.root.after(0, self._update_tab_status)

    def _run_daily_check_if_needed(self):
        """Starts daily yt-dlp version check if necessary."""
        config = self._load_yt_dlp_config()
        last_check_date = config.get('last_yt_dlp_check_date')
        today = datetime.date.today().isoformat()

        if last_check_date != today:
            # Run check in background
            threading.Thread(target=self._check_yt_dlp_versions_worker, daemon=True).start()
            # Update check date
            self._save_yt_dlp_config({'last_yt_dlp_check_date': today})

    def _update_tab_status(self):
        """Updates tab status and buttons based on saved data."""
        config = self._load_yt_dlp_config()
        status = config.get('yt_dlp_status', 'unknown')
        previous_version = config.get('previous_yt_dlp_version')

        # Control "Rollback" button
        if previous_version:
            self.revert_version_btn.config(state='normal')
        else:
            self.revert_version_btn.config(state='disabled')

        # Control tab title
        # Use saved reference to video_sub_notebook
        video_sub_notebook = self.video_sub_notebook
        tab_index = None
        for i in range(video_sub_notebook.index('end')):
            if video_sub_notebook.tab(i, 'text') in ['yt-dlp Update', 'yt-dlp Update (❗)', 'yt-dlp Update (✔)']:
                tab_index = i
                break

        if tab_index is not None:
            if status == 'update_available':
                video_sub_notebook.tab(tab_index, text='yt-dlp Update (❗)')
            elif status == 'ok':
                video_sub_notebook.tab(tab_index, text='yt-dlp Update (✔)')
            else:
                video_sub_notebook.tab(tab_index, text='yt-dlp Update')

    def _revert_yt_dlp(self):
        """Rolls back yt-dlp to the previous version."""
        threading.Thread(target=self._revert_yt_dlp_worker, daemon=True).start()

    def _revert_yt_dlp_worker(self):
        """Worker function for yt-dlp rollback."""
        def log(message):
            self.app.root.after(0, lambda: self.updater_log.insert(tk.END, message + '\n'))
            self.app.root.after(0, lambda: self.updater_log.see(tk.END))

        log("▶️ STEP 1: Reading saved previous version...")
        config = self._load_yt_dlp_config()
        previous_version = config.get('previous_yt_dlp_version')

        if not previous_version:
            log("❌ Previous version not found. Perform an update first.")
            return

        log(f"Previous version found: {previous_version}")

        log("▶️ STEP 2: Forming rollback command...")
        python_path = sys.executable
        command = [python_path, '-m', 'pip', 'install', f'yt-dlp=={previous_version}']
        log(f"Command: {' '.join(command)}")

        log("▶️ STEP 3: Executing rollback...")
        try:
            result = subprocess.run(command, capture_output=True, text=True, encoding='utf-8')
            log("pip output:")
            for line in result.stdout.split('\n'):
                if line.strip():
                    log(f"  {line}")
            if result.stderr:
                log("Errors:")
                for line in result.stderr.split('\n'):
                    if line.strip():
                        log(f"  {line}")

            if result.returncode == 0:
                log("✅ yt-dlp rollback successful")
                # Clear saved previous version after successful rollback
                self._save_yt_dlp_config({'previous_yt_dlp_version': None})
            else:
                log(f"❌ Rollback failed (code: {result.returncode})")
        except Exception as e:
            log(f"❌ Error during rollback: {e}")

        # Update tab status
        self.app.root.after(0, self._update_tab_status)

    def _update_yt_dlp(self):
        """Updates yt-dlp to the latest version."""
        threading.Thread(target=self._update_yt_dlp_worker, daemon=True).start()

    def _update_yt_dlp_worker(self):
        """Worker function for yt-dlp update."""
        def log(message):
            self.app.root.after(0, lambda: self.updater_log.insert(tk.END, message + '\n'))
            self.app.root.after(0, lambda: self.updater_log.see(tk.END))

        log("▶️ STEP 1: Retrieving current version for saving...")
        # First get the current version
        python_path = sys.executable
        try:
            result = subprocess.run([python_path, '-m', 'pip', 'show', 'yt-dlp'],
                                  capture_output=True, text=True, encoding='utf-8')
            if result.returncode == 0:
                lines = result.stdout.split('\n')
                version_line = next((line for line in lines if line.startswith('Version:')), None)
                if version_line:
                    current_version = version_line.split(': ', 1)[1].strip()
                    log(f"Current version: {current_version}")
                else:
                    log("❌ Could not determine current version")
                    return
            else:
                log(f"❌ Error getting current version: {result.stderr}")
                return
        except Exception as e:
            log(f"❌ Error getting version: {e}")
            return

        log("▶️ STEP 2: Forming update command...")
        command = [python_path, '-m', 'pip', 'install', '--upgrade', 'yt-dlp']
        log(f"Command: {' '.join(command)}")

        log("▶️ STEP 3: Executing update...")
        try:
            result = subprocess.run(command, capture_output=True, text=True, encoding='utf-8')
            log("pip output:")
            for line in result.stdout.split('\n'):
                if line.strip():
                    log(f"  {line}")
            if result.stderr:
                log("Errors:")
                for line in result.stderr.split('\n'):
                    if line.strip():
                        log(f"  {line}")

            if result.returncode == 0:
                log("✅ yt-dlp update successful")
                # Save old version for rollback ONLY AFTER SUCCESS
                self._save_yt_dlp_config({'previous_yt_dlp_version': current_version})
                log(f"Previous version saved for rollback: {current_version}")
            else:
                log(f"❌ Update failed (code: {result.returncode})")
        except Exception as e:
            log(f"❌ Error during update: {e}")

        # Update tab status
        self.app.root.after(0, self._update_tab_status)

    def _populate_semantic_slice_tab(self, parent_frame):
        """Populates semantic slicing tab with settings widgets."""
        # Variables for widget state
        self.semantic_active_var = tk.BooleanVar()
        self.slicing_type_var = tk.StringVar()

        # --- Main container and activation checkbox ---
        main_frame = ttk.Frame(parent_frame)
        main_frame.pack(fill='x')

        activation_cb = ttk.Checkbutton(main_frame, text="Activate semantic slicing",
                                        variable=self.semantic_active_var, command=self._toggle_semantic_settings_state)
        activation_cb.pack(anchor='w', pady=(0, 10))
        ToolTip(activation_cb, "Enables or disables the entire semantic slicing process that runs after video download.")

        self.settings_frame = ttk.LabelFrame(main_frame, text="Settings")
        self.settings_frame.pack(fill='x', expand=True)

        # --- Widget creation ---
        # Slicing type
        type_frame = ttk.Frame(self.settings_frame)
        type_frame.grid(row=0, column=0, columnspan=2, sticky='w', padx=10, pady=5)
        ttk.Label(type_frame, text="Slicing type:").pack(side='left', anchor='w')
        whole_only_rb = ttk.Radiobutton(type_frame, text="Whole clips only", variable=self.slicing_type_var, value="whole_only", command=self._save_semantic_settings)
        whole_only_rb.pack(side='left', padx=5)
        full_rb = ttk.Radiobutton(type_frame, text="Whole and stitched", variable=self.slicing_type_var, value="full", command=self._save_semantic_settings)
        full_rb.pack(side='left', padx=5)
        ToolTip(whole_only_rb, "Faster mode.\nLooks for continuous quotes without 'surgical' cuts.")
        ToolTip(full_rb, "Full analysis.\nLooks for both whole quotes and complex ones stitched from multiple parts.")

        # Whisper Model
        ttk.Label(self.settings_frame, text="Whisper Model:").grid(row=1, column=0, sticky='w', padx=10, pady=5)
        self.whisper_model_combo = ttk.Combobox(self.settings_frame, values=["tiny", "base", "small", "medium", "large"], state="readonly", width=10)
        self.whisper_model_combo.grid(row=1, column=1, sticky='w', padx=10, pady=5)
        self.whisper_model_combo.bind("<<ComboboxSelected>>", self._save_semantic_settings)
        ToolTip(self.whisper_model_combo, "Speech recognition model.\nbase: fast, low precision.\nsmall: great balance of speed and quality.\nmedium: high precision, slow.")

        # Language
        ttk.Label(self.settings_frame, text="Audio Language:").grid(row=2, column=0, sticky='w', padx=10, pady=5)
        self.language_combo = ttk.Combobox(self.settings_frame, values=["en", "ru", "es", "fr", "de"], width=10)
        self.language_combo.grid(row=2, column=1, sticky='w', padx=10, pady=5)
        self.language_combo.bind("<FocusOut>", self._save_semantic_settings)
        self.language_combo.bind("<<ComboboxSelected>>", self._save_semantic_settings)
        ToolTip(self.language_combo, "Specify the primary speech language in the interview (e.g., 'en' for English).")

        # Speaker count
        speakers_frame = ttk.Frame(self.settings_frame)
        speakers_frame.grid(row=3, column=1, sticky='w', padx=10, pady=5)
        ttk.Label(self.settings_frame, text="Number of speakers (Min/Max):").grid(row=3, column=0, sticky='w', padx=10, pady=5)
        self.min_speakers_spinbox = ttk.Spinbox(speakers_frame, from_=1, to=10, width=4, command=self._save_semantic_settings)
        self.min_speakers_spinbox.pack(side='left')
        self.max_speakers_spinbox = ttk.Spinbox(speakers_frame, from_=1, to=10, width=4, command=self._save_semantic_settings)
        self.max_speakers_spinbox.pack(side='left', padx=5)
        ToolTip(speakers_frame, "Approximate number of speakers. This is just a hint for the system.\nIf the actual number differs, the program will not error, but accuracy may decrease.\nFor a monologue, set 1 / 1.")

        # Clip duration
        duration_frame = ttk.Frame(self.settings_frame)
        duration_frame.grid(row=4, column=1, sticky='w', padx=10, pady=5)
        ttk.Label(self.settings_frame, text="Clip duration (Min/Max, sec):").grid(row=4, column=0, sticky='w', padx=10, pady=5)
        self.min_duration_entry = ttk.Entry(duration_frame, width=6)
        self.min_duration_entry.pack(side='left')
        self.max_duration_entry = ttk.Entry(duration_frame, width=6)
        self.max_duration_entry.pack(side='left', padx=5)
        self.min_duration_entry.bind("<FocusOut>", self._save_semantic_settings)
        self.max_duration_entry.bind("<FocusOut>", self._save_semantic_settings)
        ToolTip(duration_frame, "Specify the minimum and maximum desired duration for the final clips.\nDecimal values can be used (e.g., 6.5).")

        # Video Quality (CRF)
        ttk.Label(self.settings_frame, text="Video Quality (CRF):").grid(row=5, column=0, sticky='w', padx=10, pady=5)
        self.crf_spinbox = ttk.Spinbox(self.settings_frame, from_=0, to=51, width=10, command=self._save_semantic_settings)
        self.crf_spinbox.grid(row=5, column=1, sticky='w', padx=10, pady=5)
        crf_tooltip_text = (
            "Constant Rate Factor (CRF). Controls video quality.\n\n"
            "• 0: Lossless (huge file).\n"
            "• 17-18: Visually lossless. Perfect for archiving.\n"
            "• 23: Default value. Excellent balance of quality and size.\n"
            "• 28-30: Lower quality, high compression.\n\n"
            "Lower value = better quality and LARGER file size."
        )
        ToolTip(self.crf_spinbox, crf_tooltip_text)

        # Manual run button
        self.semantic_run_btn = ttk.Button(parent_frame, text="Slice Interview")
        self.semantic_run_btn.pack(pady=20, anchor='center')
        self.semantic_run_btn.config(width=len(self.semantic_run_btn.cget("text")))
        ToolTip(self.semantic_run_btn, "Starts the semantic slicing process only for videos\nfrom the 'raw_videos' folder of the selected entity in the table.\nThe main run occurs automatically after video download.")

        # Load saved values into created widgets
        self._apply_loaded_settings()
        # Set initial state (active/inactive)
        self._toggle_semantic_settings_state()

    def _apply_loaded_settings(self):
        """Applies values loaded from self.semantic_settings to the widgets."""
        if not hasattr(self, 'semantic_settings'):
            return
        self.semantic_active_var.set(self.semantic_settings.get("is_active", False))
        self.slicing_type_var.set(self.semantic_settings.get("slicing_type", "full"))
        self.language_combo.set(self.semantic_settings.get("language", "en"))
        self.whisper_model_combo.set(self.semantic_settings.get("whisper_model", "small"))
        self.min_speakers_spinbox.set(self.semantic_settings.get("min_speakers", 2))
        self.max_speakers_spinbox.set(self.semantic_settings.get("max_speakers", 5))
        self.min_duration_entry.delete(0, tk.END)
        self.min_duration_entry.insert(0, self.semantic_settings.get("min_clip_duration", "6.0"))
        self.max_duration_entry.delete(0, tk.END)
        self.max_duration_entry.insert(0, self.semantic_settings.get("max_clip_duration", "12.0"))
        self.crf_spinbox.set(self.semantic_settings.get("video_crf", 23))

    def _toggle_semantic_settings_state(self):
        """Enables or disables all settings widgets depending on the main checkbox."""
        new_state = 'normal' if self.semantic_active_var.get() else 'disabled'

        # Iterate through all child elements of settings frame
        for widget in self.settings_frame.winfo_children():
            # For frames with radiobuttons or spinboxes, recurse through children
            if isinstance(widget, (ttk.Frame, ttk.LabelFrame)):
                for child in widget.winfo_children():
                    try:
                        child.configure(state=new_state)
                    except tk.TclError:
                        pass
            else:
                try:
                    widget.configure(state=new_state)
                except tk.TclError:
                    pass

        self.semantic_run_btn.configure(state=new_state)
        # Save checkbox state
        self._save_semantic_settings()

    def _populate_updater_tab(self, parent_frame):
        """Populates yt-dlp update tab with widgets."""
        # Frame for buttons
        buttons_frame = ttk.Frame(parent_frame)
        buttons_frame.pack(fill='x', pady=(0, 10))

        # Three buttons
        self.check_version_btn = ttk.Button(buttons_frame, text="Check Version", command=self._check_yt_dlp_versions)
        self.check_version_btn.pack(side='left', padx=(0, 5))

        self.update_version_btn = ttk.Button(buttons_frame, text="Update to Latest", command=self._update_yt_dlp)
        self.update_version_btn.pack(side='left', padx=(0, 5))

        self.revert_version_btn = ttk.Button(buttons_frame, text="Revert to Previous", command=self._revert_yt_dlp, state='disabled')
        self.revert_version_btn.pack(side='left')

        # Log
        self.updater_log = scrolledtext.ScrolledText(parent_frame, height=15, wrap='word')
        self.updater_log.pack(fill='both', expand=True, padx=5, pady=5)
        self.updater_log.configure(state='normal')

    def _populate_composite_tab(self, parent_frame):
        """Populates composite video creation tab (Sandwich Worker)."""
        # Navigation panel
        nav_panel = self._create_common_nav_panel(parent_frame, tab_type='video')
        nav_panel.pack(fill='x', pady=(0, 5), anchor='w')

        # Frame for links entry
        composite_links_frame = ttk.LabelFrame(parent_frame, text="Composite Video (Sandwich)", padding=5)
        composite_links_frame.pack(fill='both', expand=True, pady=5)

        # Window 1: Audio
        ttk.Label(composite_links_frame, text="1. AUDIO Source (One link)").pack(anchor='w')
        self.composite_audio_text = scrolledtext.ScrolledText(composite_links_frame, height=3, width=40)
        self.composite_audio_text.pack(fill='both', expand=True, pady=(0, 10))

        # Window 2: Video
        ttk.Label(composite_links_frame, text="2. VIDEO Source (One or more links)").pack(anchor='w')
        self.composite_video_text = scrolledtext.ScrolledText(composite_links_frame, height=8, width=40)
        self.composite_video_text.pack(fill='both', expand=True, pady=(0, 10))

        # Composite creation buttons
        composite_buttons_frame = ttk.Frame(composite_links_frame)
        composite_buttons_frame.pack(anchor='center', pady=(5, 0))

        self.composite_start_btn = ttk.Button(
            composite_buttons_frame,
            text="Create Composite (Sandwich)",
            command=self.add_composite_task
        )
        self.composite_start_btn.pack(side='left', padx=(0, 5))

        self.composite_with_slice_btn = ttk.Button(
            composite_buttons_frame,
            text="Sandwich with slicing",
            command=self.add_composite_with_slice_task
        )
        self.composite_with_slice_btn.pack(side='left')

        # Context menu binding
        self.composite_audio_text.bind("<Button-2>", self._show_context_menu)
        self.composite_audio_text.bind("<Button-3>", self._show_context_menu)
        self.composite_video_text.bind("<Button-2>", self._show_context_menu)
        self.composite_video_text.bind("<Button-3>", self._show_context_menu)

    def _load_yt_dlp_config(self):
        """Loads yt-dlp settings from config.json."""
        if not self.app.CONFIG_FILE_PATH:
            return {
                'last_yt_dlp_check_date': None,
                'previous_yt_dlp_version': None,
                'yt_dlp_status': 'unknown'
            }
        try:
            with open(self.app.CONFIG_FILE_PATH, 'r', encoding='utf-8') as f:
                config_data = json.load(f)
            return {
                'last_yt_dlp_check_date': config_data.get('last_yt_dlp_check_date', None),
                'previous_yt_dlp_version': config_data.get('previous_yt_dlp_version', None),
                'yt_dlp_status': config_data.get('yt_dlp_status', 'unknown')
            }
        except (FileNotFoundError, json.JSONDecodeError):
            return {
                'last_yt_dlp_check_date': None,
                'previous_yt_dlp_version': None,
                'yt_dlp_status': 'unknown'
            }

    def _save_yt_dlp_config(self, updates):
        """Saves yt-dlp settings updates to config.json."""
        if not self.app.CONFIG_FILE_PATH:
            return
        try:
            with open(self.app.CONFIG_FILE_PATH, 'r', encoding='utf-8') as f:
                config_data = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError):
            config_data = {}

        config_data.update(updates)

        with open(self.app.CONFIG_FILE_PATH, 'w', encoding='utf-8') as f:
            json.dump(config_data, f, indent=4, ensure_ascii=False)

    def ensure_worker_running(self, worker_type):
        worker_map = {'photo': self._photo_worker_loop, 'video': self._video_worker_loop,
                      'slicer': self._slicer_worker_loop, 'slices': self._slices_worker_loop,
                      'zoom': self._zoom_worker_loop,
                      'semantic_slicer': self._semantic_slicer_worker_loop,
                      'conversion': self._conversion_worker_loop,
                      'composite': self._composite_worker_loop}
        log_map = {'photo': (self.photo_dl_log, 'photo_dl'), 'video': (self.video_dl_log, 'video_dl'),
                   'slicer': (self.slicer_log, 'slicer'), 'slices': (self.video_dl_log, 'video_dl'),
                   'zoom': (self.zoom_log, 'zoom'),
                   'semantic_slicer': (self.video_dl_log, 'video_dl'),
                   'conversion': (self.slicer_log, 'slicer'),
                   'composite': (self.video_dl_log, 'video_dl')}
        current_thread = getattr(self, f"{worker_type}_worker_thread")
        if current_thread is None or not current_thread.is_alive():
            log_panel, log_key = log_map[worker_type]
            self.log_to_panel(log_panel, log_key, f"Starting worker ({worker_type})...")
            self.stop_event.clear()
            new_thread = threading.Thread(target=worker_map[worker_type], daemon=True)
            new_thread.start()
            setattr(self, f"{worker_type}_worker_thread", new_thread)

    def _photo_worker_loop(self):
        self._generic_download_worker(self.photo_dl_queue, self.photo_dl_log, 'photo_dl',
                                      download_orchestrator_v2.process_single_photo_task, "unchecked_photos")
        self.photo_worker_thread = None

    def _video_worker_loop(self):
        # ... (start of function without changes) ...
        progress_callback = lambda msg: self.log_to_panel(self.video_dl_log, 'video_dl', msg)
        while True:
            if self.stop_event.is_set(): progress_callback("Worker (video download) stopped."); break
            try:
                task = self.video_dl_queue.get(timeout=1)
                overwrite_decision = task.get('overwrite_decision', 'append')
                downloaded_files = download_orchestrator_v2.process_single_video_task(
                    task=task, base_archive_path=Path(self.app.get_setting('media_archive_path', "")),
                    progress_callback=progress_callback, stop_event=self.stop_event,
                    overwrite_decision=overwrite_decision
                )
                if not self.stop_event.is_set():
                    eid = task.get('eid');
                    name = task.get('name');
                    role = task.get('role')
                    ssid = task.get('ssid')
                    song_name = task.get('song_name')
                    
                    entity_path, _ = entity_manager_v2.get_or_create_entity_path(
                        Path(self.app.get_setting('media_archive_path', "")), eid, name, role, ssid=ssid, song_name=song_name
                    )
                    if entity_path:
                        folder_path = entity_path / "raw_videos"
                        actual_count = len([f for f in folder_path.iterdir() if f.is_file() and not f.name.startswith('.')]) if folder_path.exists() else 0
                        self.app.root.after(0, self.update_treeview_row, ssid, {'clip': actual_count})
                if downloaded_files:
                    # Create path ourselves, as in old logic
                    entity_path, _ = entity_manager_v2.get_or_create_entity_path(
                        Path(self.app.get_setting('media_archive_path', "")),
                        task['eid'], task['name'], task['role'],
                        ssid=task.get('ssid'), song_name=task.get('song_name')
                    )

                    # Separate files
                    interview_files = [f for f in downloaded_files if "--INTERVIEW--" in f.name]

                    # --- AUTOMATIC SLICING DISABLED (at user request) ---
                    # Task for simple slicer (ALL files)
                    # slicer_task = task.copy()
                    # slicer_task['files_to_slice'] = downloaded_files
                    # slicer_task['entity_path'] = entity_path  # Add entity_path to task
                    # self.slicer_queue.put(slicer_task)
                    # self.log_to_panel(self.slicer_log, 'slicer', f"Added {len(downloaded_files)} videos to the simple slicing queue.")
                    # self.ensure_worker_running('slicer')

                    # Task for semantic slicer (ONLY interviews), if present
                    # if interview_files and self.semantic_active_var.get():
                    #     semantic_task = task.copy()
                    #     semantic_task['interview_files'] = interview_files
                    #     semantic_task['entity_path'] = entity_path  # Add entity_path to task
                    #     current_semantic_settings = {
                    #         "slicing_type": self.slicing_type_var.get(),
                    #         "language": self.language_combo.get(),
                    #         "whisper_model": self.whisper_model_combo.get(),
                    #         "min_speakers": int(self.min_speakers_spinbox.get()),
                    #         "max_speakers": int(self.max_speakers_spinbox.get()),
                    #         "min_clip_duration": self.min_duration_entry.get(),
                    #         "max_clip_duration": self.max_duration_entry.get(),
                    #         "video_crf": int(self.crf_spinbox.get())
                    #     }
                    #     semantic_task['settings'] = current_semantic_settings
                    #     google_key = self.app.get_setting('GOOGLE_API_KEY')
                    #     hf_token = self.app.get_setting('HF_TOKEN')
                    #     semantic_task['api_keys'] = {'google_api_key': google_key, 'hf_token': hf_token}
                    #     self.semantic_slicer_queue.put(semantic_task)
                    #     self.log_to_panel(self.video_dl_log, 'video_dl', f"Added {len(interview_files)} interviews to the semantic slicing queue.")
                    #     self.ensure_worker_running('semantic_slicer')

                    # --- SLICING INTO SLICES (if enabled) ---
                    if task.get('with_slicing', False):
                        slice_duration = float(self.slice_duration.get())
                        slices_task = task.copy()
                        slices_task['slice_duration'] = slice_duration
                        slices_task['entity_path'] = entity_path
                        slices_task['downloaded_files'] = downloaded_files
                        self.slices_queue.put(slices_task)
                        self.log_to_panel(self.video_dl_log, 'video_dl', f"➔ Task added to slice for parts (duration: {slice_duration} sec).")
                        self.ensure_worker_running('slices')

                self.video_dl_queue.task_done()
            except queue.Empty:
                progress_callback("Queue (video download) is empty. Worker is finishing.");
                break
            except Exception as e:
                progress_callback(f"Critical error in worker (video download): {e}");
                break
        self.video_worker_thread = None

    def _slicer_worker_loop(self):
        progress_callback = lambda msg: self.log_to_panel(self.slicer_log, 'slicer', msg)
        while True:
            if self.stop_event.is_set(): progress_callback("Worker (slice/conversion) stopped."); break
            try:
                task = self.slicer_queue.get(timeout=1)
                files_to_slice = task.get('files_to_slice', [])
                progress_callback(f"▶️ Starting simple slicing of {len(files_to_slice)} videos for '{task.get('name')}'")
                for video_file in files_to_slice:
                    if self.stop_event.is_set(): break
                    # CALL THE CORRECT FUNCTION
                    video_slicer_v2.slice_clips_only(
                        source_video_path=video_file,
                        entity_path=task.get('entity_path'),
                        eid=task.get('eid'),
                        num_cuts=task.get('slicer_settings', {}).get('num_cuts', 5),
                        cut_duration=task.get('slicer_settings', {}).get('cut_duration', 6),
                        progress_callback=progress_callback
                    )
                self.slicer_queue.task_done()
            except queue.Empty:
                progress_callback("Queue (slice/conversion) is empty. Worker is finishing.");
                break
            except Exception as e:
                progress_callback(f"❌ Critical error in worker (slice/conversion): {e}");
                break
        self.slicer_worker_thread = None

    def _slices_worker_loop(self):
        """Worker for slicing golden master into parts."""
        progress_callback = lambda msg: self.log_to_panel(self.video_dl_log, 'video_dl', msg)
        while True:
            if self.stop_event.is_set():
                progress_callback("Worker (parts slicing) stopped.")
                break
            try:
                task = self.slices_queue.get(timeout=1)
                entity_path = task.get('entity_path')
                eid = task.get('eid')
                ssid = task.get('ssid')
                slice_duration = task.get('slice_duration', 4.8)
                downloaded_files = task.get('downloaded_files', [])

                # Find all golden master files in raw_videos
                raw_videos_path = entity_path / "raw_videos"
                if not raw_videos_path.exists():
                    progress_callback("❌ Folder raw_videos not found.")
                    self.slices_queue.task_done()
                    continue

                # Filter only downloaded files (if specified) or all mp4 files
                if downloaded_files:
                    source_files = [f for f in downloaded_files if f.exists() and f.suffix == '.mp4']
                else:
                    source_files = sorted(raw_videos_path.glob('raw_video_*.mp4'))

                if not source_files:
                    progress_callback("❌ No files to slice into parts.")
                    self.slices_queue.task_done()
                    continue

                progress_callback(f"▶️ Starting slicing of {len(source_files)} files into parts (duration: {slice_duration} sec)...")

                # Get data for naming
                name = task.get('name')
                song_name = task.get('song_name')

                # Create slices folder
                slices_path = entity_path / "slices"
                slices_path.mkdir(exist_ok=True)

                total_slices_created = 0

                for source_idx, source_file in enumerate(source_files, start=1):
                    if self.stop_event.is_set():
                        break

                    progress_callback(f"   → Processing file {source_idx}/{len(source_files)}: {source_file.name}")

                    # Get source file duration
                    import video_slicer_v2
                    total_duration = video_slicer_v2.get_video_duration(source_file, progress_callback)

                    if total_duration < slice_duration:
                        progress_callback(f"      Skipping (duration {total_duration:.1f} sec < {slice_duration} sec)")
                        continue

                    # Calculate number of full slices
                    num_slices = int(total_duration // slice_duration)

                    progress_callback(f"      {num_slices} slices will be created")

                    # Source name for prefix
                    source_name_stem = source_file.stem
                    source_index_match = re.search(r'raw_video_(\d+)_', source_name_stem)
                    source_index = int(source_index_match.group(1)) if source_index_match else source_idx

                    # Name sanitization
                    def sanitize(t): return t.lower().replace(' ', '_')
                    def clean(t): return re.sub(r'[^a-z0-9_-]', '', sanitize(str(t)))

                    artist_clean = clean(name)
                    song_clean = clean(song_name)
                    is_sandwich = task.get('is_sandwich', False)

                    # Slicing
                    for i in range(num_slices):
                        if self.stop_event.is_set():
                            break

                        start_time = i * slice_duration
                        slice_idx = i + 1

                        # Output filename (adding _sandwich if it's a composite)
                        sandwich_suffix = "_sandwich" if is_sandwich else ""
                        output_filename = f"{ssid}_{artist_clean}_{song_clean}{sandwich_suffix}_{source_index:03d}_{slice_idx:03d}.mp4"
                        output_path = slices_path / output_filename

                        # ffmpeg command for stream copy
                        slice_command = [
                            'ffmpeg', '-y',
                            '-ss', str(start_time),
                            '-i', str(source_file),
                            '-t', str(slice_duration),
                            '-c:v', 'copy',
                            '-c:a', 'copy',
                            '-movflags', '+faststart',
                            str(output_path)
                        ]

                        try:
                            subprocess.run(slice_command, check=True, capture_output=True, text=True, encoding='utf-8')
                            progress_callback(f"      ✅ {output_filename}")
                            total_slices_created += 1
                        except subprocess.CalledProcessError as e:
                            progress_callback(f"      ❌ Slicing error {output_filename}: {e.stderr.strip()}")
                            continue

                # Update statistics in TreeView
                if ssid and total_slices_created > 0:
                    slices_count = len([f for f in slices_path.iterdir() if f.is_file() and f.suffix == '.mp4'])
                    self.app.root.after(0, self.update_treeview_row, ssid, {'slices': slices_count})

                progress_callback(f"✅ Slicing into parts completed. {total_slices_created} slices created.")
                self.slices_queue.task_done()

            except queue.Empty:
                progress_callback("Queue (parts slicing) is empty. Worker is finishing.")
                break
            except Exception as e:
                progress_callback(f"❌ Critical error in worker (parts slicing): {e}")
                import traceback
                progress_callback(f"   {traceback.format_exc()}")
                break
        self.slices_worker_thread = None

    def _zoom_worker_loop(self):
        progress_callback = lambda msg: self.log_to_panel(self.zoom_log, 'zoom', msg)
        while True:
            if self.stop_event.is_set(): progress_callback("Worker (zooms) stopped."); break
            try:
                task = self.zoom_queue.get(timeout=1)
                eid = task.get('eid');
                name = task.get('name');
                role = task.get('role')
                settings = task.get('settings', {});
                entity_path, _ = entity_manager_v2.get_or_create_entity_path(Path(self.app.get_setting('media_archive_path', "")), eid, name,
                                                                             role)
                if not entity_path: self.zoom_queue.task_done(); continue
                photos_path = entity_path / "photos"
                if not photos_path.exists():
                    progress_callback(f"Folder 'photos' for {name} not found.");
                    self.zoom_queue.task_done();
                    continue
                photo_files = [p for p in photos_path.iterdir() if p.suffix.lower() in ['.jpg', '.jpeg', '.png']]
                if not photo_files:
                    progress_callback(f"No photos to process in folder 'photos' for {name}.");
                    self.zoom_queue.task_done();
                    continue
                overwrite_decision = task.get('overwrite_decision', 'append')
                target_path = entity_path / "zoomed_videos"
                if overwrite_decision == 'overwrite':
                    if target_path.exists():
                        progress_callback(f"⚠️ Clearing folder {target_path} as requested...");
                        shutil.rmtree(target_path)
                progress_callback(f"Starting creation of {len(photo_files)} zooms for {name}...")
                created_any = False
                for i, photo_file in enumerate(sorted(photo_files)):
                    if self.stop_event.is_set(): break
                    if zoom_creator_v2.create_zoom_from_photo(photo_file, entity_path, settings, index=i + 1, eid=eid,
                                                              progress_callback=progress_callback):
                        created_any = True
                zoom_creator_v2.cleanup_temp_folder(entity_path)
                if not self.stop_event.is_set():
                    progress_callback(f"✅ Zoom creation for {name} completed.")
                    if created_any:
                        zooms_path = entity_path / "zoomed_videos"
                        if zooms_path.exists():
                            all_zooms = [f.name for f in zooms_path.iterdir() if f.is_file() and f.suffix == '.mp4']
                            h_zooms_count = len([z for z in all_zooms if z.startswith('h_')])
                            zooms_count = len(all_zooms) - h_zooms_count
                            new_stats = {'zooms_total': len(all_zooms), 'zooms_highlights': h_zooms_count}
                            entity_manager_v2.update_readiness_stats(entity_path, new_stats, progress_callback)
                            self.app.root.after(0, self.update_treeview_row, eid,
                                                {'zoom': zooms_count, 'h_zoom': h_zooms_count})
                self.zoom_queue.task_done()
            except queue.Empty:
                progress_callback("Queue (zooms) is empty. Worker is finishing.");
                break
            except Exception as e:
                progress_callback(f"Critical error in worker (zooms): {e}");
                break
        self.zoom_worker_thread = None

    def _semantic_slicer_worker_loop(self):
        progress_callback = lambda msg: self.log_to_panel(self.video_dl_log, 'video_dl', msg)

        while not self.stop_event.is_set():
            try:
                task = self.semantic_slicer_queue.get(timeout=1)

                # Serialize task dictionary to JSON string
                task_json_str = json.dumps(task, default=str)

                # Path to Python interpreter in current venv
                python_executable = sys.executable
                # Path to worker script
                script_path = Path(__file__).parent / "interview_semantic_slicer.py"

                command = [python_executable, str(script_path)]

                progress_callback(f"▶️ Starting semantic slicing in isolated process...")

                # Wrap start and wait in try...finally for guaranteed cleanup
                process = None
                try:
                    # Start subprocess
                    process = subprocess.Popen(
                        command,
                        stdin=subprocess.PIPE,
                        stdout=subprocess.PIPE,
                        stderr=subprocess.PIPE,
                        text=True,
                        encoding='utf-8'
                    )

                    # Pass JSON to stdin and close it
                    process.stdin.write(task_json_str)
                    process.stdin.close()

                    # Read output (logs) from stdout in real time
                    while True:
                        output = process.stdout.readline()
                        if output == '' and process.poll() is not None:
                            break
                        if output:
                            progress_callback(output.strip())

                    # Check return code
                    return_code = process.poll()
                    if return_code != 0:
                        stderr_output = process.stderr.read()
                        progress_callback(f"❌ Slicing process finished with error (code: {return_code}).")
                        progress_callback(f"stderr: {stderr_output.strip()}")
                    else:
                        progress_callback(f"✅ Isolated slicing process completed successfully.")

                finally:
                    # <<< START OF NEW BLOCK: SANITARY CLEANUP >>>
                    progress_callback("🧹 Performing sanitary cleanup of orphaned processes...")
                    try:
                        # This command will find and kill any processes whose command line
                        # contains both the path to our venv python and the name of our worker script.
                        # This is safe and will not affect the main GUI.
                        cleanup_command = (
                            f"pkill -f \"{sys.executable}.*interview_semantic_slicer.py\""
                        )
                        # We use Popen instead of run to avoid blocking and errors if processes are not found
                        cleanup_process = subprocess.Popen(cleanup_command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
                        cleanup_process.communicate(timeout=5) # Wait no more than 5 seconds
                        progress_callback("✅ Sanitary cleanup completed.")
                    except Exception as cleanup_err:
                        progress_callback(f"⚠️ Error during sanitary cleanup: {cleanup_err}")
                    # <<< END OF NEW BLOCK >>>

                # After completion - start conversion (logic remains)
                if not self.stop_event.is_set() and task.get('interview_files'):
                    progress_callback("Starting interview conversion...")
                    conversion_task = task.copy()
                    self.conversion_queue.put(conversion_task)
                    self.ensure_worker_running('conversion')

                self.semantic_slicer_queue.task_done()

            except queue.Empty:
                progress_callback("[Semantic Slicer]: Queue is empty. Worker is finishing.")
                break
            except Exception as e:
                progress_callback(f"[Semantic Slicer]: ❌ Critical error in control worker: {e}\n{traceback.format_exc()}")
                break
        self.semantic_slicer_worker_thread = None

    def _conversion_worker_loop(self):
        progress_callback = lambda msg: self.log_to_panel(self.slicer_log, 'slicer', msg) # Write to slice log
        while True:
            if self.stop_event.is_set(): progress_callback("Worker (conversion) stopped."); break
            try:
                task = self.conversion_queue.get(timeout=1)
                interview_files = task.get('interview_files', [])
                progress_callback(f"▶️ Starting background P-Pro conversion of {len(interview_files)} interviews...")
                for video_file in interview_files:
                    if self.stop_event.is_set(): break
                    video_slicer_v2.convert_master_file_only(video_file, progress_callback)
                self.conversion_queue.task_done()
            except queue.Empty:
                progress_callback("Queue (conversion) is empty. Worker is finishing."); break
            except Exception as e:
                progress_callback(f"❌ Critical error in worker (conversion): {e}"); break
        self.conversion_worker_thread = None

    def _composite_worker_loop(self):
        """Worker for processing composite video creation tasks (Sandwich Worker)."""
        progress_callback = lambda msg: self.log_to_panel(self.video_dl_log, 'video_dl', msg)
        while True:
            if self.stop_event.is_set():
                progress_callback("Worker (composite video) stopped.")
                break
            try:
                task = self.composite_queue.get(timeout=1)
                overwrite_decision = task.get('overwrite_decision', 'append')

                # Import composite video handler
                import video_song_sandwich_worker

                # Call handler
                result_path = video_song_sandwich_worker.process_sandwich_task(
                    task=task,
                    base_archive_path=Path(self.app.get_setting('media_archive_path', "")),
                    progress_callback=progress_callback,
                    stop_event=self.stop_event
                )

                # Update statistics in tree after successful completion
                if result_path and not self.stop_event.is_set():
                    eid = task.get('eid')
                    ssid = task.get('ssid')
                    name = task.get('name')
                    role = task.get('role')
                    song_name = task.get('song_name', '')

                    entity_path, _ = entity_manager_v2.get_or_create_entity_path(
                        Path(self.app.get_setting('media_archive_path', "")),
                        eid, name, role, ssid=ssid, song_name=song_name
                    )

                    if entity_path:
                        folder_path = entity_path / "raw_videos"
                        actual_count = len([f for f in folder_path.iterdir()
                                           if f.is_file() and not f.name.startswith('.')]) if folder_path.exists() else 0
                        self.app.root.after(0, self.update_treeview_row, ssid, {'clip': actual_count})

                        # --- SLICING INTO SLICES (if enabled for composite) ---
                        if task.get('with_slicing', False):
                            slice_duration = task.get('slice_duration', 4.8)
                            is_sandwich = task.get('is_sandwich', False)

                            # Find created composite file
                            if result_path and result_path.exists():
                                slices_task = task.copy()
                                slices_task['slice_duration'] = slice_duration
                                slices_task['entity_path'] = entity_path
                                slices_task['downloaded_files'] = [result_path]
                                slices_task['is_sandwich'] = is_sandwich
                                self.slices_queue.put(slices_task)
                                self.log_to_panel(self.video_dl_log, 'video_dl',
                                                  f"➔ Task added to slice parts from composite (duration: {slice_duration} sec).")
                                self.ensure_worker_running('slices')

                self.composite_queue.task_done()
            except queue.Empty:
                progress_callback("Queue (composite video) is empty. Worker is finishing.")
                break
            except Exception as e:
                progress_callback(f"❌ Critical error in worker (composite video): {e}")
                break
        self.composite_worker_thread = None

    def _generic_download_worker(self, task_queue, log_panel, log_key, processor_func, target_folder):
        progress_callback = lambda msg: self.log_to_panel(log_panel, log_key, msg)
        while True:
            if self.stop_event.is_set(): progress_callback(f"Worker stopped."); break
            try:
                task = task_queue.get(timeout=1)
                overwrite_decision = task.get('overwrite_decision', 'append')
                was_successful = processor_func(
                    task=task, base_archive_path=Path(self.app.get_setting('media_archive_path', "")),
                    progress_callback=progress_callback, stop_event=self.stop_event,
                    overwrite_decision=overwrite_decision
                )
                if was_successful and not self.stop_event.is_set():
                    eid = task.get('eid');
                    name = task.get('name');
                    role = task.get('role')
                    entity_path, _ = entity_manager_v2.get_or_create_entity_path(Path(self.app.get_setting('media_archive_path', "")), eid, name,
                                                                                 role)
                    if entity_path:
                        actual_folder_path = entity_path / target_folder
                        actual_count = len([f for f in actual_folder_path.iterdir() if
                                            f.is_file()]) if actual_folder_path.exists() else 0
                        new_stats = {'unchecked_photos_total': actual_count}
                        entity_manager_v2.update_readiness_stats(entity_path, new_stats, progress_callback)
                        self.app.root.after(0, self.update_treeview_row, eid, {'unchecked_photos': actual_count})
                task_queue.task_done()
            except queue.Empty:
                progress_callback("Queue is empty. Worker is finishing.");
                break
            except Exception as e:
                progress_callback(f"Critical error in worker: {e}");
                break

    def on_entity_select(self, event=None):
        selection = self.tree.selection()
        if not selection: return
        item_id = selection[0]
        item_values = self.tree.item(item_id, 'values')
        tags = self.tree.item(item_id, 'tags')
        
        # tags: (eid, role, ssid, song_name)
        self.active_entity_data = {
            'name': item_values[0], # Artist Name
            'eid': tags[0],
            'role': tags[1],
            'ssid': tags[2],
            'song_name': tags[3]
        }
        self.set_action_panel_state('normal')
        self.regular_video_links_text.delete('1.0', tk.END)
        self.interview_video_links_text.delete('1.0', tk.END)

    def _normalize_single_id(self, part: str) -> str:
        part = part.strip()
        if not part: return ""
        
        # If starts with a digit - normalize to VIDxxxx + suffix
        if part[0].isdigit():
            match = re.match(r'(\d+)(.*)', part)
            if match:
                num_str, suffix = match.groups()
                return f"VID{num_str.zfill(4)}{suffix}"
        
        # If starts with VID (any case) - normalize only digits
        if part.upper().startswith('VID'):
            match = re.search(r'VID(\d+)(.*)', part, re.IGNORECASE)
            if match:
                num_str, suffix = match.groups()
                return f"VID{num_str.zfill(4)}{suffix}"
        
        # In other cases (PIZDA112, etc.) return as is
        return part

    def _parse_input_to_sheets(self, raw_input: str) -> List[str]:
        final_sheets = []
        parts = [p.strip() for p in raw_input.split(',') if p.strip()]
        
        for part in parts:
            if '-' in part:
                range_match = re.match(r'(\d+)([a-zA-Z0-9]*)-(\d+)([a-zA-Z0-9]*)', part)
                if range_match:
                    start_num, start_suffix, end_num, end_suffix = range_match.groups()
                    if start_suffix == end_suffix:
                        s, e = int(start_num), int(end_num)
                        step = 1 if s <= e else -1
                        for n in range(s, e + step, step):
                            final_sheets.append(self._normalize_single_id(f"{n}{start_suffix}"))
                        continue
            final_sheets.append(self._normalize_single_id(part))
        return final_sheets

    def _find_sheet_case_insensitive(self, video_id: str):
        try:
            workbook = load_workbook(self.DB_PATH, read_only=True)
            found_name = None
            for sheet_name in workbook.sheetnames:
                if sheet_name.lower() == video_id.lower():
                    found_name = sheet_name
                    break
            workbook.close()
            return found_name
        except Exception:
            return None

    def _on_load_click(self):
        raw_input = self.video_id_entry.get().strip()
        if not raw_input: return

        requested_sheets = self._parse_input_to_sheets(raw_input)
        if not requested_sheets: return
        
        self.video_id_entry.delete(0, tk.END)
        self.video_id_entry.insert(0, ", ".join(requested_sheets))

        found_sheets = []
        missing_sheets = []
        
        for vid in requested_sheets:
            actual = self._find_sheet_case_insensitive(vid)
            if actual:
                found_sheets.append(actual)
            else:
                missing_sheets.append(vid)

        if missing_sheets:
            msg = "The following sheets were not found in the database:\n\n" + "\n".join(missing_sheets)
            if found_sheets:
                msg += "\n\nContinue loading only the found sheets?"
                if not messagebox.askyesno("Sheets not found", msg, parent=self):
                    self.scan_status_label.config(text="Loading canceled.", foreground="red")
                    return
            else:
                messagebox.showerror("Error", "None of the specified sheets were found.", parent=self)
                self.scan_status_label.config(text="Sheets not found.", foreground="red")
                return

        status_text = f"Project {found_sheets[0]} loaded" if len(found_sheets) == 1 else f"Projects {', '.join(found_sheets)} loaded"
        self.scan_status_label.config(text=status_text, foreground="blue")
        
        threading.Thread(target=self.populate_tasks_from_project, args=(found_sheets,), daemon=True).start()

    def _on_load_from_json_click(self):
        raw_input = self.video_id_entry.get().strip()
        if not raw_input: return

        requested_pids = self._parse_input_to_sheets(raw_input)
        if not requested_pids: return

        self.video_id_entry.delete(0, tk.END)
        self.video_id_entry.insert(0, ", ".join(requested_pids))

        seed_dir = Path(self.app.WORK_ROOT_PATH) / "database" / "seed"
        found_files = []
        missing_files = []

        for pid in requested_pids:
            # Form filename
            if not pid.startswith("PID"):
                pid_name = f"PID{pid}"
            else:
                pid_name = pid

            seed_file = seed_dir / f"{pid_name}_seed.json"

            if seed_file.exists():
                found_files.append(seed_file)
            else:
                missing_files.append(pid)

        if missing_files:
            msg = "The following seed files were not found:\n\n" + "\n".join(missing_files)
            if found_files:
                msg += "\n\nContinue loading only the found files?"
                if not messagebox.askyesno("Files not found", msg, parent=self):
                    self.scan_status_label.config(text="Loading canceled.", foreground="red")
                    return
            else:
                messagebox.showerror("Error", "None of the specified seed files were found.", parent=self)
                self.scan_status_label.config(text="Files not found.", foreground="red")
                return

        status_text = f"Project from JSON loaded" if len(found_files) == 1 else f"{len(found_files)} projects from JSON loaded"
        self.scan_status_label.config(text=status_text, foreground="blue")

        threading.Thread(target=self.populate_tasks_from_json, args=(found_files,), daemon=True).start()

    def populate_tasks_from_project(self, sheet_names):
        try:
            # Force reread database on every "Load" click to see new entities
            self.songs_db = pd.read_excel(self.DB_PATH, sheet_name='Songs_Database')
            self.master_db = pd.read_excel(self.DB_PATH, sheet_name='MasterDatabase_v1')

            # Dictionary for storing visuals: {ssid: visuals_value}
            visuals_dict = {}

            # Read visuals from project sheet (ONLY for blocks from column C)
            if len(sheet_names) == 1:
                project_df = pd.read_excel(self.DB_PATH, sheet_name=sheet_names[0])
                # Column C (index 2) - SSID, column J (index 9) - visuals
                if len(project_df.columns) > 9:
                    for _, row in project_df.iterrows():
                        ssid = row.iloc[2] if len(row) > 2 else None
                        visuals = row.iloc[9] if len(row) > 9 else ""
                        if ssid and pd.notna(ssid):
                            visuals_dict[ssid] = visuals if pd.notna(visuals) else ""

            all_unique_ssids = []
            source_order_counter = 0
            for sheet_name in sheet_names:
                df = pd.read_excel(self.DB_PATH, sheet_name=sheet_name)

                # SSID from Intro (column R, index 17) - DO NOT read visuals for intro!
                if len(df.columns) > 17:
                    intro_ssids = df.iloc[:, 17].dropna().unique().tolist()
                    for s in intro_ssids:
                        if s not in all_unique_ssids:
                            all_unique_ssids.append(s)
                            source_order_counter += 1

                # SSID from blocks (column C, index 2) - Read visuals for these
                if len(df.columns) > 2:
                    block_ssids = df.iloc[:, 2].dropna().unique().tolist()
                    for s in block_ssids:
                        if s not in all_unique_ssids:
                            all_unique_ssids.append(s)
                            source_order_counter += 1

            archive_path = Path(self.app.get_setting('media_archive_path', ""))
            tasks_data = []

            for idx, ssid in enumerate(all_unique_ssids):
                song_row = self.songs_db[self.songs_db['SSID'] == ssid]
                if song_row.empty: continue
                
                song_name = song_row.iloc[0]['Song_Name']
                eid = song_row.iloc[0]['EID']
                artist_name = song_row.iloc[0]['Name']
                
                artist_row = self.master_db[self.master_db['EID'] == eid]
                role = artist_row.iloc[0]['Role'] if not artist_row.empty else "Artist"

                def sanitize(t): return t.lower().replace(' ', '_')
                def clean(t): return re.sub(r'[^a-z0-9_-]', '', sanitize(str(t)))

                artist_folder_name = f"{clean(artist_name)}_{eid}"
                song_folder_name = f"{ssid}_{clean(song_name)}_by_{clean(artist_name)}_{eid}"
                
                raw_videos_path = archive_path / artist_folder_name / song_folder_name / "raw_videos"
                clip_count = 0
                if raw_videos_path.exists():
                    clip_count = len([f for f in raw_videos_path.iterdir() if f.is_file() and not f.name.startswith('.')])

                tasks_data.append({
                    'artist': artist_name, 'song': song_name, 'ssid': ssid,
                    'clip': clip_count, 'slices': 0, 'eid': eid, 'role': role,
                    'source_order': idx, 'visuals': visuals_dict.get(ssid, '')
                })

            # Conditional sorting: if 1 project - by Excel order, if more - alphabetical
            if len(sheet_names) == 1:
                tasks_data.sort(key=lambda x: x.get('source_order', 0))
            else:
                # Sorting: First Clip == 0, then Clip > 0. Inside - by artist alphabet.
                tasks_data.sort(key=lambda x: (x['clip'] > 0, x['artist'].lower()))

            def update_ui():
                for i in self.tree.get_children(): self.tree.delete(i)
                for item in tasks_data:
                    self.tree.insert("", "end", values=(
                        item['artist'], item['song'], item['ssid'], item['clip'], item['slices'], item.get('visuals', '')
                    ), tags=(item['eid'], item['role'], item['ssid'], item['song']))

                if tasks_data:
                    first = self.tree.get_children()[0]
                    self.tree.selection_set(first)
                    self.tree.focus(first)

                # Save source info
                self.loaded_source = 'excel'
                self.loaded_files = sheet_names
                self._update_order_buttons_state()

            self.app.root.after(0, update_ui)

        except Exception as e:
            self.app.root.after(0, lambda: self.scan_status_label.config(text=f"Error: {e}", foreground="red"))

        except Exception as e:
            self.app.root.after(0, lambda: self.scan_status_label.config(text=f"Error: {e}"))

    def populate_tasks_from_json(self, json_files):
        try:
            # Force reread database on every "Load" click to see new entities
            self.songs_db = pd.read_excel(self.DB_PATH, sheet_name='Songs_Database')
            self.master_db = pd.read_excel(self.DB_PATH, sheet_name='MasterDatabase_v1')

            all_ssids_with_order = []

            for json_file in json_files:
                with open(json_file, 'r', encoding='utf-8') as f:
                    seed_data = json.load(f)

                # Extract songs from JSON
                songs = seed_data.get('songs', [])
                for song_entry in songs:
                    ssid = song_entry.get('ssid')
                    index = song_entry.get('index', 0)
                    visuals = song_entry.get('visuals', '')
                    if ssid:
                        all_ssids_with_order.append({'ssid': ssid, 'source_order': index, 'visuals': visuals})

            archive_path = Path(self.app.get_setting('media_archive_path', ""))
            tasks_data = []

            for item in all_ssids_with_order:
                ssid = item['ssid']
                source_order = item['source_order']

                song_row = self.songs_db[self.songs_db['SSID'] == ssid]
                if song_row.empty: continue

                song_name = song_row.iloc[0]['Song_Name']
                eid = song_row.iloc[0]['EID']
                artist_name = song_row.iloc[0]['Name']

                artist_row = self.master_db[self.master_db['EID'] == eid]
                role = artist_row.iloc[0]['Role'] if not artist_row.empty else "Artist"

                def sanitize(t): return t.lower().replace(' ', '_')
                def clean(t): return re.sub(r'[^a-z0-9_-]', '', sanitize(str(t)))

                artist_folder_name = f"{clean(artist_name)}_{eid}"
                song_folder_name = f"{ssid}_{clean(song_name)}_by_{clean(artist_name)}_{eid}"

                raw_videos_path = archive_path / artist_folder_name / song_folder_name / "raw_videos"
                clip_count = 0
                if raw_videos_path.exists():
                    clip_count = len([f for f in raw_videos_path.iterdir() if f.is_file() and not f.name.startswith('.')])

                tasks_data.append({
                    'artist': artist_name, 'song': song_name, 'ssid': ssid,
                    'clip': clip_count, 'slices': 0, 'eid': eid, 'role': role,
                    'source_order': source_order, 'visuals': item.get('visuals', '')
                })

            # Conditional sorting: if 1 project - by JSON order, if more - alphabetical
            if len(json_files) == 1:
                tasks_data.sort(key=lambda x: x.get('source_order', 0))
            else:
                tasks_data.sort(key=lambda x: (x['clip'] > 0, x['artist'].lower()))

            def update_ui():
                for i in self.tree.get_children(): self.tree.delete(i)
                for item in tasks_data:
                    self.tree.insert("", "end", values=(
                        item['artist'], item['song'], item['ssid'], item['clip'], item['slices'], item.get('visuals', '')
                    ), tags=(item['eid'], item['role'], item['ssid'], item['song']))

                if tasks_data:
                    first = self.tree.get_children()[0]
                    self.tree.selection_set(first)
                    self.tree.focus(first)

                # Save source info
                self.loaded_source = 'json'
                self.loaded_files = [f.name for f in json_files]
                self._update_order_buttons_state()

            self.app.root.after(0, update_ui)

        except Exception as e:
            self.app.root.after(0, lambda: self.scan_status_label.config(text=f"Error: {e}", foreground="red"))

    def update_treeview_row(self, ssid, new_stats):
        item_id_to_update = None
        for item_id in self.tree.get_children():
            tags = self.tree.item(item_id, 'tags')
            # tags: (eid, role, ssid, song_name)
            if tags and len(tags) > 2 and tags[2] == ssid:
                item_id_to_update = item_id
                break
        if not item_id_to_update: return
        current_values = list(self.tree.item(item_id_to_update, 'values'))

        # In the new interface update only Clip (index 3) and Slices (index 4)
        if 'clip' in new_stats:
            current_values[3] = new_stats['clip']
        if 'slices' in new_stats:
            current_values[4] = new_stats['slices']

        self.tree.item(item_id_to_update, values=tuple(current_values))

    def open_youtube(self):
        if not self.active_entity_data: return
        artist_name = self.active_entity_data.get('name')
        song_name = self.active_entity_data.get('song_name', "")
        
        # Format: "song_name artist_name video"
        query = f"{song_name} {artist_name} video".strip()
        
        if query: webbrowser.open(f"https://www.youtube.com/results?search_query={urllib.parse.quote_plus(query)}")

    def _get_active_entity_path(self):
        if not self.active_entity_data or not self.active_entity_data.get('eid'): return None
        return entity_manager_v2.find_path_by_eid(Path(self.app.get_setting('media_archive_path', "")), self.active_entity_data['eid'],
                                                  self.app.log_message_to_selection_tab)

    def open_clips_folder(self):
        entity_path = self._get_active_entity_path()
        if not entity_path: self.app.log_message_to_selection_tab(
            f"Folder for {self.active_entity_data.get('eid', 'N/A')} not found."); return
        self._open_path_in_finder(entity_path / "selected_clips")

    def open_slices_folder(self):
        """Opens slices folder (if files exist) or raw_videos (if slices is empty/missing)."""
        if not self.active_entity_data or not self.active_entity_data.get('eid'):
            self.app.log_message_to_selection_tab("First select a song from the list.")
            return

        name = self.active_entity_data.get('name')
        eid = self.active_entity_data.get('eid')
        role = self.active_entity_data.get('role')
        ssid = self.active_entity_data.get('ssid')
        song_name = self.active_entity_data.get('song_name')

        entity_path, msg = entity_manager_v2.get_or_create_entity_path(
            Path(self.app.get_setting('media_archive_path', "")),
            eid, name, role, ssid=ssid, song_name=song_name
        )

        if not entity_path:
            self.app.log_message_to_selection_tab(
            f"Folder for {ssid} not found: {msg}")
            return

        slices_path = entity_path / "slices"
        raw_videos_path = entity_path / "raw_videos"

        # Logic: open slices if exists and not empty, otherwise raw_videos
        if slices_path.exists() and any(slices_path.iterdir()):
            target_path = slices_path
            folder_name = "Slices"
        elif raw_videos_path.exists():
            target_path = raw_videos_path
            folder_name = "Raw Videos"
        else:
            # Neither folder exists — show message
            self.app.log_message_to_selection_tab(f"There is no slices/ or raw_videos/ folder for {ssid}")
            return

        from tkinter import filedialog
        filedialog.askopenfilename(title=f"Folder {folder_name}: {ssid}", initialdir=str(target_path))

    def open_in_bridge(self):
        entity_path = self._get_active_entity_path()
        if not entity_path: self.app.log_message_to_selection_tab(
            f"Folder for {self.active_entity_data.get('eid', 'N/A')} not found."); return
        target_path = entity_path / "unchecked_photos"
        bridge_app_path = "/Applications/Adobe Bridge 2025/Adobe Bridge 2025.app"
        try:
            if not Path(bridge_app_path).exists():
                self.app.log_message_to_selection_tab(f"Error: Adobe Bridge not found at path {bridge_app_path}");
                self._open_path_in_finder(target_path);
                return
            if not target_path.exists(): target_path.mkdir(parents=True, exist_ok=True)
            subprocess.run(['open', '-a', bridge_app_path, str(target_path)], check=True)
        except Exception as e:
            self.app.log_message_to_selection_tab(f"Error opening Adobe Bridge: {e}. Opening in Finder.");
            self._open_path_in_finder(target_path)

    def _open_path_in_finder(self, path_to_open: Path):
        try:
            if not path_to_open.exists():
                path_to_open.mkdir(parents=True, exist_ok=True)
            if sys.platform == "darwin":
                subprocess.run(['open', '-R', str(path_to_open)], check=True)
            elif sys.platform == "win32":
                subprocess.run(['explorer', str(path_to_open)], check=True)
            else:
                subprocess.run(['xdg-open', str(path_to_open)], check=True)
        except Exception as e:
            self.app.log_message_to_selection_tab(f"Could not open folder {path_to_open}: {e}")

    def _move_entity_up(self):
        # Check: only JSON and only ONE project
        if self.loaded_source != 'json' or len(self.loaded_files) != 1:
            messagebox.showwarning("Unavailable",
                "Order adjustment is only available for a SINGLE JSON project.",
                parent=self)
            return

        sel = self.tree.selection()
        if not sel: return

        for s in sel:
            idx = self.tree.index(s)
            if idx > 0:
                self.tree.move(s, self.tree.parent(s), idx-1)

        self._mark_order_as_modified()

    def _move_entity_down(self):
        # Check: only JSON and only ONE project
        if self.loaded_source != 'json' or len(self.loaded_files) != 1:
            messagebox.showwarning("Unavailable",
                "Order adjustment is only available for a SINGLE JSON project.",
                parent=self)
            return

        sel = self.tree.selection()
        if not sel: return

        for s in reversed(sel):
            idx = self.tree.index(s)
            self.tree.move(s, self.tree.parent(s), idx+1)

        self._mark_order_as_modified()

    def _mark_order_as_modified(self):
        # Mark order as modified
        self.order_modified = True
        self.order_save_btn.config(text="Save order (*)")

    def _save_order_to_json(self):
        # Check: only JSON and only ONE project
        if self.loaded_source != 'json' or len(self.loaded_files) != 1:
            messagebox.showwarning("Unavailable",
                "Order saving is only available for a SINGLE JSON project.",
                parent=self)
            return

        if not self.loaded_files:
            messagebox.showerror("Error", "No loaded files.", parent=self)
            return

        # Get current order from tree
        current_order = []
        for idx, item_id in enumerate(self.tree.get_children()):
            tags = self.tree.item(item_id, 'tags')
            # tags: (eid, role, ssid, song_name)
            if tags and len(tags) > 2:
                ssid = tags[2]
                current_order.append({'ssid': ssid, 'new_index': idx + 1})

        # Save to JSON file
        try:
            seed_dir = Path(self.app.WORK_ROOT_PATH) / "database" / "seed"
            json_filename = self.loaded_files[0]
            json_path = seed_dir / json_filename

            if not json_path.exists():
                messagebox.showerror("Error", f"File {json_filename} not found.", parent=self)
                return

            with open(json_path, 'r', encoding='utf-8') as f:
                seed_data = json.load(f)

            # Update order in songs[]
            songs = seed_data.get('songs', [])
            for song_entry in songs:
                ssid = song_entry.get('ssid')
                for order_item in current_order:
                    if order_item['ssid'] == ssid:
                        song_entry['index'] = order_item['new_index']
                        break

            # Sort songs[] by new index
            seed_data['songs'] = sorted(songs, key=lambda x: x.get('index', 0))

            # Save back to file
            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump(seed_data, f, indent=4, ensure_ascii=False)

            messagebox.showinfo("Success", f"Order saved in {json_filename}.", parent=self)
            self.order_modified = False
            self.order_save_btn.config(text="Save order")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to save: {e}", parent=self)

    def _update_order_buttons_state(self):
        # Update source status and buttons state
        if self.loaded_source == 'json' and len(self.loaded_files) == 1:
            source_text = f"JSON ({self.loaded_files[0]})"
            self.order_source_label.config(text=source_text, foreground="green")
            state = 'normal'
        elif self.loaded_source == 'json' and len(self.loaded_files) > 1:
            source_text = f"JSON ({len(self.loaded_files)} files - order adjustment unavailable)"
            self.order_source_label.config(text=source_text, foreground="orange")
            state = 'disabled'
        elif self.loaded_source == 'excel':
            source_text = f"Excel (order adjustment unavailable)"
            self.order_source_label.config(text=source_text, foreground="orange")
            state = 'disabled'
        else:
            source_text = "Not loaded"
            self.order_source_label.config(text=source_text, foreground="gray")
            state = 'disabled'

        # Set button states
        self.order_up_btn.config(state=state)
        self.order_down_btn.config(state=state)
        self.order_save_btn.config(state=state)

    def _set_visuals_plus(self):
        """Sets visual tag +1 for selected row"""
        # Check: single project loaded
        if len(self.loaded_files) != 1:
            messagebox.showwarning("Unavailable",
                "Visual tag setting is only available for a SINGLE loaded project.",
                parent=self)
            return

        # Check: selected row exists
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("Attention", "Select a row in the table.", parent=self)
            return

        item_id = sel[0]
        tags = self.tree.item(item_id, 'tags')
        # tags: (eid, role, ssid, song_name)
        ssid = tags[2] if len(tags) > 2 else None

        if not ssid:
            messagebox.showerror("Error", "Could not determine SSID.", parent=self)
            return

        # Update value in treeview
        current_values = list(self.tree.item(item_id, 'values'))
        current_values[5] = "+1"  # Visuals column (index 5)
        self.tree.item(item_id, values=tuple(current_values))

        # Save to source (Excel or JSON)
        self._save_visuals_to_source(ssid, "+1")

    def _set_visuals_minus(self):
        """Sets visual tag -1 for selected row"""
        # Check: single project loaded
        if len(self.loaded_files) != 1:
            messagebox.showwarning("Unavailable",
                "Visual tag setting is only available for a SINGLE loaded project.",
                parent=self)
            return

        # Check: selected row exists
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("Attention", "Select a row in the table.", parent=self)
            return

        item_id = sel[0]
        tags = self.tree.item(item_id, 'tags')
        ssid = tags[2] if len(tags) > 2 else None

        if not ssid:
            messagebox.showerror("Error", "Could not determine SSID.", parent=self)
            return

        # Update value in treeview
        current_values = list(self.tree.item(item_id, 'values'))
        current_values[5] = "-1"
        self.tree.item(item_id, values=tuple(current_values))

        # Save to source
        self._save_visuals_to_source(ssid, "-1")

    def _clear_visuals(self):
        """Clears visual tag for selected row"""
        # Check: single project loaded
        if len(self.loaded_files) != 1:
            messagebox.showwarning("Unavailable",
                "Visual tag clearing is only available for a SINGLE loaded project.",
                parent=self)
            return

        # Check: selected row exists
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("Attention", "Select a row in the table.", parent=self)
            return

        item_id = sel[0]
        tags = self.tree.item(item_id, 'tags')
        ssid = tags[2] if len(tags) > 2 else None

        if not ssid:
            messagebox.showerror("Error", "Could not determine SSID.", parent=self)
            return

        # Update value in treeview
        current_values = list(self.tree.item(item_id, 'values'))
        current_values[5] = ""
        self.tree.item(item_id, values=tuple(current_values))

        # Save to source
        self._save_visuals_to_source(ssid, "")

    def _save_visuals_to_source(self, ssid, visuals_value):
        """Saves visuals value to source (Excel or JSON)"""
        try:
            if self.loaded_source == 'excel':
                self._save_visuals_to_excel(ssid, visuals_value)
            elif self.loaded_source == 'json':
                self._save_visuals_to_json(ssid, visuals_value)
        except Exception as e:
            messagebox.showerror("Save error", f"Failed to save: {e}", parent=self)

    def _save_visuals_to_excel(self, ssid, visuals_value):
        """SAFELY saves visuals to Excel column J"""
        from openpyxl import load_workbook

        excel_path = self.DB_PATH
        project_sheet_name = self.loaded_files[0]

        # SAFE Excel loading
        workbook = load_workbook(excel_path)

        if project_sheet_name not in workbook.sheetnames:
            workbook.close()
            raise ValueError(f"Sheet {project_sheet_name} not found in Excel")

        sheet = workbook[project_sheet_name]

        # Find row where column C (column=3) == ssid
        row_number = None
        for row_idx in range(1, sheet.max_row + 1):
            cell_value = sheet.cell(row=row_idx, column=3).value
            if cell_value == ssid:
                row_number = row_idx
                break

        if row_number is None:
            workbook.close()
            raise ValueError(f"SSID {ssid} not found in sheet {project_sheet_name}")

        # Write to column J (column=10)
        sheet.cell(row=row_number, column=10, value=visuals_value if visuals_value else None)

        # SAFE saving
        workbook.save(excel_path)
        workbook.close()

    def _save_visuals_to_json(self, ssid, visuals_value):
        """Saves visuals to JSON seed file"""
        seed_dir = Path(self.app.WORK_ROOT_PATH) / "database" / "seed"
        json_filename = self.loaded_files[0]
        json_path = seed_dir / json_filename

        if not json_path.exists():
            raise ValueError(f"File {json_filename} not found")

        with open(json_path, 'r', encoding='utf-8') as f:
            seed_data = json.load(f)

        # Find song by SSID
        found = False
        for song in seed_data.get('songs', []):
            if song.get('ssid') == ssid:
                if visuals_value == "":
                    # Remove key if value is empty
                    song.pop('visuals', None)
                else:
                    song['visuals'] = visuals_value
                found = True
                break

        if not found:
            raise ValueError(f"SSID {ssid} not found in JSON")

        # Save back
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(seed_data, f, indent=4, ensure_ascii=False)

    def set_action_panel_state(self, state):
        selected_tab_widget = self.nametowidget(self.action_panel.select())
        for widget in selected_tab_widget.winfo_children():
            self._set_widget_state_recursively(widget, state)

    def _set_widget_state_recursively(self, widget, state):
        try:
            if widget.winfo_class() == 'TFrame' and any(
                    isinstance(child, ttk.Button) for child in widget.winfo_children()):
                if "youtube_btn" in str(widget.winfo_children()[0]): return
            widget.configure(state=state)
        except tk.TclError:
            pass
        for child in widget.winfo_children(): self._set_widget_state_recursively(child, state)