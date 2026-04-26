# --- text_processing_interface.py ---

import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox, Listbox
import threading
import time
import re
import difflib
import sys
from pathlib import Path
from typing import List, Dict

import text_splitter_worker


class TextProcessingTab(ttk.Frame):
    def __init__(self, parent, controller):
        super().__init__(parent)
        self.controller = controller
        self.all_blocks_data = []
        self.current_pid = None
        self.timer_id = None
        self.start_time = 0
        self.font_size = self.controller.get_setting('splitter_font_size', 14)
        self.stop_processing = False
        self.processed_pids = []
        self.queued_pids = []

        self._create_widgets()
        self._configure_tags()
        self._update_font_size()

        # macOS-specif fix removed - causes button malfunctions
        # if sys.platform.startswith('darwin'):
        #     self._setup_macos_focus_fix()

    def _create_widgets(self):
        # --- Top Panel ---
        top_frame = ttk.Frame(self)
        top_frame.grid(row=0, column=0, columnspan=3, padx=10, pady=(10, 0), sticky="ew")

        controls_left_frame = ttk.Frame(top_frame)
        controls_left_frame.pack(side="left")

        pid_frame = ttk.Frame(controls_left_frame)
        pid_frame.pack(fill="x", anchor="w")
        ttk.Label(pid_frame, text="PID(s):").pack(side="left", padx=(0, 5))
        self.pid_entry = ttk.Entry(pid_frame, width=30)
        self.pid_entry.pack(side="left")

        buttons_frame = ttk.Frame(controls_left_frame)
        buttons_frame.pack(fill="x", anchor="w", pady=(5, 0))
        self.process_button = ttk.Button(buttons_frame, text="Process", command=self._on_process_click)
        self.process_button.pack(side="left")
        self.stop_button = ttk.Button(buttons_frame, text="Stop", command=self._on_stop_click, state="disabled")
        self.stop_button.pack(side="left", padx=5)
        self.save_button = ttk.Button(buttons_frame, text="Save to Excel", command=self._on_save_click,
                                      state="disabled")
        self.save_button.pack(side="left", padx=5)
        self.database_button = ttk.Button(buttons_frame, text="Find in Database", command=self._on_database_click,
                                          state="disabled")
        self.database_button.pack(side="left", padx=5)

        auto_accept_frame = ttk.Frame(top_frame)
        auto_accept_frame.pack(side="left", padx=20)

        self.auto_accept_var = tk.BooleanVar(value=False)
        self.auto_accept_check = ttk.Checkbutton(auto_accept_frame, text="Auto-accept if deviation <=",
                                                 variable=self.auto_accept_var)
        self.auto_accept_check.pack(side="left")

        self.threshold_var = tk.DoubleVar(value=2.0)
        self.threshold_spinbox = ttk.Spinbox(auto_accept_frame, from_=0.0, to=100.0, increment=0.1,
                                             textvariable=self.threshold_var, width=5)
        self.threshold_spinbox.pack(side="left", padx=(5, 0))
        ttk.Label(auto_accept_frame, text="%").pack(side="left")

        controls_right_frame = ttk.Frame(top_frame)
        controls_right_frame.pack(side="right")

        status_frame = ttk.Frame(controls_right_frame)
        status_frame.pack(fill="x", anchor="e")
        self.status_label = ttk.Label(status_frame, text="", foreground="blue", width=25, anchor="w")
        self.timer_label = ttk.Label(status_frame, text="", foreground="gray", width=15, anchor="w")
        self.status_label.pack(side="left")
        self.timer_label.pack(side="left", padx=10)

        font_frame = ttk.Frame(controls_right_frame)
        font_frame.pack(fill="x", anchor="e", pady=(5, 0))
        ttk.Button(font_frame, text="A-", command=self._decrease_font, width=3).pack(side="left")
        self.font_size_label = ttk.Label(font_frame, text=str(self.font_size), width=3, anchor="center")
        self.font_size_label.pack(side="left")
        ttk.Button(font_frame, text="A+", command=self._increase_font, width=3).pack(side="left")

        counts_frame = ttk.Frame(self)
        counts_frame.grid(row=1, column=0, columnspan=3, padx=10, pady=5, sticky="ew")
        self.block_stats_label = ttk.Label(counts_frame, text="Block Statistics", anchor="w")
        self.block_stats_label.pack(side="left")

        self.grid_rowconfigure(2, weight=10)
        self.grid_rowconfigure(3, weight=0, minsize=30)
        self.grid_rowconfigure(4, weight=1, minsize=100)

        self.grid_columnconfigure(0, weight=1, minsize=200)
        self.grid_columnconfigure(1, weight=4)
        self.grid_columnconfigure(2, weight=4)

        nav_frame = ttk.Frame(self)
        nav_frame.grid(row=2, column=0, padx=(10, 5), pady=5, sticky="nsew")
        nav_frame.rowconfigure(0, weight=1); 
        nav_frame.columnconfigure(0, weight=1)
        self.blocks_listbox = Listbox(nav_frame, exportselection=False, borderwidth=0, highlightthickness=0)
        self.blocks_listbox.grid(row=0, column=0, sticky="nsew")
        nav_scrollbar = ttk.Scrollbar(nav_frame, orient="vertical", command=self.blocks_listbox.yview)
        nav_scrollbar.grid(row=0, column=1, sticky="ns")
        self.blocks_listbox.config(yscrollcommand=nav_scrollbar.set)
        self.blocks_listbox.bind("<<ListboxSelect>>", self._on_block_select)

        source_frame = ttk.Frame(self)
        source_frame.grid(row=2, column=1, padx=5, pady=5, sticky="nsew")
        source_frame.rowconfigure(0, weight=1); 
        source_frame.columnconfigure(0, weight=1)
        self.source_text_widget = scrolledtext.ScrolledText(source_frame, wrap=tk.WORD, state="disabled",
                                                            relief="solid", borderwidth=1)
        self.source_text_widget.grid(row=0, column=0, sticky="nsew")

        processed_frame = ttk.Frame(self)
        processed_frame.grid(row=2, column=2, padx=(5, 10), pady=5, sticky="nsew")
        processed_frame.rowconfigure(0, weight=1); 
        processed_frame.columnconfigure(0, weight=1)
        self.processed_text_widget = scrolledtext.ScrolledText(processed_frame, wrap=tk.WORD, relief="solid",
                                                               borderwidth=1, undo=True)
        self.processed_text_widget.grid(row=0, column=0, sticky="nsew")

        queue_frame = ttk.Frame(self)
        queue_frame.grid(row=3, column=0, columnspan=3, padx=10, pady=5, sticky="ew")
        self.queue_label = ttk.Label(queue_frame, text="", foreground="blue", anchor="w")
        self.queue_label.pack(fill="x")

        log_frame = ttk.LabelFrame(self, text="Operations Log", padding=5)
        log_frame.grid(row=4, column=0, columnspan=3, padx=10, pady=(5, 10), sticky="nsew")
        log_frame.rowconfigure(0, weight=1); 
        log_frame.columnconfigure(0, weight=1)
        self.log_text_widget = scrolledtext.ScrolledText(log_frame, wrap=tk.WORD, height=6, state="normal")
        self.log_text_widget.grid(row=0, column=0, sticky="nsew")
        # self.log_text_widget.bind("<KeyPress>", lambda e: "break")  # Unblocked for copy/paste
        self.log_text_widget.bind("<Button-3>", lambda e: None)  # Allow right-click menu

    def _configure_tags(self):
        self.source_text_widget.tag_configure("replaced", background="#A0C4FF")
        self.processed_text_widget.tag_configure("inserted", background="#A6F6A0")

    def _update_font_size(self):
        font_tuple = ("Helvetica", self.font_size)
        self.source_text_widget.config(font=font_tuple)
        self.processed_text_widget.config(font=font_tuple)
        self.blocks_listbox.config(font=("Helvetica", self.font_size - 2))
        self.log_text_widget.config(font=("Helvetica", self.font_size - 4))
        self.font_size_label.config(text=str(self.font_size))
        self.controller.set_setting('splitter_font_size', self.font_size)

    def _increase_font(self):
        self.font_size += 1; self._update_font_size()

    def _decrease_font(self):
        self.font_size = max(8, self.font_size - 1); self._update_font_size()

    def _start_timer(self):
        self.start_time = time.time()

        def update():
            self.timer_label.config(text=f"{int(time.time() - self.start_time)} sec.")
            self.timer_id = self.after(1000, update)

        update()

    def _stop_timer(self):
        if self.timer_id: self.after_cancel(self.timer_id); self.timer_id = None
        self.timer_label.config(text=f"Completed in {int(time.time() - self.start_time)} sec.")

    def _parse_pid_input(self, pid_string: str) -> List[str]:
        pids = set()
        # First remove all spaces, then the prefix 'pid' case-insensitively
        cleaned_string = pid_string.replace(" ", "")
        cleaned_string = re.sub(r'pid', '', cleaned_string, flags=re.IGNORECASE)
        parts = cleaned_string.split(',')

        for part in parts:
            if not part: continue
            if '-' in part:
                try:
                    start_str, end_str = part.split('-')
                    if not start_str or not end_str: continue
                    start, end = map(int, (start_str, end_str))
                    for i in range(start, end + 1):
                        pids.add(f"PID{i:04d}")
                except (ValueError, IndexError):
                    self._log(f"Invalid PID range: {part}", is_error=True); 
                    return []
            else:
                try:
                    pids.add(f"PID{int(part):04d}")
                except ValueError:
                    self._log(f"Invalid PID format: {part}", is_error=True); return []
        return sorted(list(pids))

    def _on_process_click(self):
        print(f"[DEBUG] TEXT_PROCESSING: _on_process_click called!")
        pid_list = self._parse_pid_input(self.pid_entry.get())
        print(f"[DEBUG] TEXT_PROCESSING: PID parsed: {pid_list}")
        if not pid_list:
            messagebox.showerror("Error", "Could not recognize PID(s). Check format (e.g., 2, 4-6, 10).",
                                 parent=self)
            print(f"[DEBUG] TEXT_PROCESSING: Invalid PID format, exiting")
            return
        print(f"[DEBUG] TEXT_PROCESSING: Processing will continue with {len(pid_list)} PIDs")

        # --- Move check here ---
        if self.auto_accept_var.get() and len(pid_list) > 1:
             pids_to_process = []
             pids_to_skip = []
             from pathlib import Path
             import openpyxl
             db_path = Path(self.controller.get_setting('work_root_path')) / "database" / "main_database.xlsx"

             if db_path.exists():
                 try:
                     workbook = openpyxl.load_workbook(db_path)
                     existing_sheets = workbook.sheetnames
                     workbook.close()
                 except Exception as e:
                     self._log(f"Could not read Excel file for verification: {e}", is_error=True)
                     existing_sheets = []
             else:
                 existing_sheets = []

             for pid in pid_list:
                 vid_number_match = re.search(r'\d+', pid)
                 sheet_name = f"VID{vid_number_match.group()}s" if vid_number_match else None
                 if sheet_name and sheet_name in existing_sheets:
                     pids_to_skip.append(pid)
                 else:
                     pids_to_process.append(pid)

             if pids_to_skip:
                 self._log(f"Batch mode: The following PIDs will be skipped because sheets for them already exist: {', '.join(pids_to_skip)}")
                 if not pids_to_process:
                     self._log("No new PIDs to process. Operation completed.")
                     return

             pid_list = pids_to_process
             self._log(f"Batch mode: {len(pid_list)} PIDs accepted for processing: {', '.join(pid_list)}")
        # --- End of migration ---

        self._clear_ui_widgets()
        self._set_ui_state(is_busy=True)
        self._start_timer()
        self.stop_processing = False
        self.processed_pids = []
        self.queued_pids = pid_list[1:] if len(pid_list) > 1 else []
        self._update_queue_display()

        if self.auto_accept_var.get():
            threading.Thread(target=self._run_batch_processing, args=(pid_list,), daemon=True).start()
        else:
            self.current_pid = pid_list[0]
            if len(pid_list) > 1:
                self._log(
                    f"REVIEW MODE: Only the first PID is being processed: {self.current_pid}. Check the box for batch processing.")
            self._run_single_processing(self.current_pid)

    def _run_single_processing(self, pid: str):
        print(f"[DEBUG] TEXT_PROCESSING: _run_single_processing started for PID {pid}")
        settings = {'pid': pid, 'projects_root_path': self.controller.get_setting('work_root_path'),
                    'mapper_prompt_path': self.controller.PROMPTS_PATH / "SplitterPrompt_Stage1_Mapper.txt",
                    'corrector_prompt_path': self.controller.PROMPTS_PATH / "SplitterPrompt_Stage2_Corrector.txt"}
        print(f"[DEBUG] TEXT_PROCESSING: Settings prepared, starting thread...")
        threading.Thread(target=self._run_worker_in_thread, args=(settings, self._process_and_display_result, self._log),
                         daemon=True).start()

    def _run_batch_processing(self, pid_list: List[str]):
        is_stopped_for_review = False
        try:
            for i, pid in enumerate(pid_list):
                if self.stop_processing:
                    self._log("Processing stopped by user.")
                    break
                self.current_pid = pid
                # DO NOT remove from queue here! Remove ONLY after successful completion
                self._update_queue_display()
                self.after(0, lambda p=pid, c=i + 1, t=len(pid_list): self._update_status(
                    f"AUTO MODE: Processing {p} ({c}/{t})..."))

                settings = {'pid': pid, 'projects_root_path': self.controller.get_setting('work_root_path'),
                            'mapper_prompt_path': self.controller.PROMPTS_PATH / "SplitterPrompt_Stage1_Mapper.txt",
                            'corrector_prompt_path': self.controller.PROMPTS_PATH / "SplitterPrompt_Stage2_Corrector.txt"}

                result_holder = {};
                event = threading.Event()

                def callback(result):
                    result_holder['result'] = result; event.set()

                worker_thread = threading.Thread(target=self._run_worker_in_thread, args=(settings, callback), daemon=True)
                worker_thread.start()
                event.wait(timeout=300)  # Timeout 5 minutes to avoid hanging forever

                result = result_holder.get('result')
                if not result or result['status'] == 'error':
                    self.after(0, lambda p=pid, msg=result.get('message', ''): self._update_status(
                        f"Error processing {p}. Batch processing stopped."))
                    is_stopped_for_review = True;
                    break

                self.all_blocks_data = result.get('final_blocks', [])

                is_problematic = False
                threshold = self.threshold_var.get()
                for block in self.all_blocks_data:
                    s_words, p_words = len(block.get('source_text', '').split()), len(
                        block.get('processed_text', '').split())

                    # Fixed mathematical error when dividing by zero
                    if s_words > 0:
                        delta = abs((p_words - s_words) / s_words * 100)
                    elif p_words > 0:
                        # If source text is empty but processed is not, consider it 100% change
                        delta = 100.0
                    else:
                        # If both texts are empty, consider deviation 0%
                        delta = 0.0

                    if delta > threshold:
                        self._log(
                            f"{pid}: Block '{block.get('caption')}' failed verification (deviation {delta:.1f}% > {threshold}%)")
                        is_problematic = True
                        break

                if not is_problematic:
                    self.after(0, lambda p=pid: self._log(f"{p}: All blocks passed verification. Auto-saving..."))
                    # Run database matching in auto mode
                    self._log(f"{p}: Starting automatic database search...")
                    settings = {
                        'pid': pid,
                        'projects_root_path': self.controller.get_setting('work_root_path')
                    }
                    result = text_splitter_worker.run_database_matching(
                        settings,
                        self.all_blocks_data,
                        lambda msg: None,  # No status callback in auto mode
                        lambda msg, is_error=False: self._log(f"{p}: {msg}", is_error),
                        self.controller.ai_manager
                    )
                    if result['status'] == 'success':
                        self.all_blocks_data = result.get('final_blocks', self.all_blocks_data)
                        matched = result.get('blocks_matched', 0)
                        total = result.get('blocks_total', 0)
                        self._log(f"{p}: Found {matched}/{total} matches in database")

                        # Save directly, bypassing _check_if_database_search_needed check
                        self._log(f"{p}: Saving to Excel...")
                        settings = {
                            'pid': pid,  # Use current PID from loop, not self.current_pid
                            'projects_root_path': self.controller.get_setting('work_root_path')
                        }
                        final_blocks_for_excel = list(self.all_blocks_data)
                        # Perform save synchronously in this thread
                        success = text_splitter_worker.save_results_to_excel(
                            settings,
                            final_blocks_for_excel,
                            lambda msg: self._log(f"{p}: {msg}"),
                            lambda msg, is_error=False: self._log(f"{p}: {msg}", is_error)
                        )

                        # Check save result
                        if not success:
                            self._log(f"{p}: Error saving to Excel", is_error=True)
                            is_stopped_for_review = True
                            break

                        # SUCCESS! Now remove PID from queue
                        if self.queued_pids:
                            removed = self.queued_pids.pop(0)
                            self._log(f"{p}: Successfully processed and saved. Removed from queue: {removed}")
                        self._update_queue_display()
                    else:
                        self._log(f"{p}: Error searching in database: {result.get('message', '')}", is_error=True)
                        is_stopped_for_review = True
                        break
                    time.sleep(1)
                else:
                    is_stopped_for_review = True
                    self.after(0, self._process_and_display_result, result)
                    break
        finally:
            self.after(0, self._set_ui_state, False)
            self.after(0, self._stop_timer)
            if not is_stopped_for_review:
                self.after(0, lambda: self._update_status("All PIDs processed successfully."))

    def _run_worker_in_thread(self, settings, callback, log_callback=None):
        if log_callback is None:
            log_callback = self._log

        # Check stop flag before starting
        if hasattr(self, 'stop_processing') and self.stop_processing:
            print("[DEBUG] TEXT_PROCESSING: Processing already stopped, returning early")
            self.after(0, callback, {'status': 'stopped'})
            return

        print(f"[DEBUG] TEXT_PROCESSING: Starting worker thread with settings keys: {list(settings.keys())}")

        # Create a wrapper status callback that checks stop flag
        def status_check_wrapper(msg):
            if hasattr(self, 'stop_processing') and self.stop_processing:
                # Don't update status if stopped, just return
                return
            self._update_status(msg)

        result = text_splitter_worker.run_splitting_process(settings, status_check_wrapper, log_callback, self.controller.ai_manager)

        # Check if we were stopped during processing
        if hasattr(self, 'stop_processing') and self.stop_processing:
            result = {'status': 'stopped'}

        print(f"[DEBUG] TEXT_PROCESSING: Worker finished, result status: {result.get('status', 'error')}")
        self.after(0, callback, result)

    def _process_and_display_result(self, result):
        if self.auto_accept_var.get() == False:
            self._stop_timer()
            self.after(0, self._set_ui_state, False)  # Ensure UI state reset even in review mode

        if result.get('status') == 'stopped':
            self._update_status("Processing stopped by user")
            self._log("Processing interrupted by 'Stop' button")
            return

        if result['status'] == 'error':
            self._update_status(f"Error: {result['message']}");
            return

        self.all_blocks_data = result.get('final_blocks', [])

        self._clear_ui_widgets()
        if self.current_pid:
            pid_num = re.search(r'\d+', self.current_pid)
            if pid_num: self.pid_entry.insert(0, pid_num.group())

        for i, block_data in enumerate(self.all_blocks_data):
            caption = block_data.get('caption', '')
            display_caption = "[Intro]" if i == 0 and not caption else caption
            self.blocks_listbox.insert(tk.END, display_caption)
        if self.all_blocks_data:
            self.blocks_listbox.selection_set(0)
            self._on_block_select(None)
        self.save_button.config(state="normal")
        self.database_button.config(state="normal")
        self._update_status("Ready. Awaiting review.")

    def _on_block_select(self, event=None):
        self.update_idletasks()
        selection_indices = self.blocks_listbox.curselection()
        if not selection_indices: return

        block_data = self.all_blocks_data[selection_indices[0]]

        source_text, processed_text = block_data.get('source_text', ''), block_data.get('processed_text', '')

        for widget in [self.source_text_widget, self.processed_text_widget]:
            widget.config(state="normal");
            widget.delete("1.0", tk.END)

        self.source_text_widget.insert("1.0", source_text)
        self.processed_text_widget.insert("1.0", processed_text)

        s_words, p_words = source_text.split(), processed_text.split()
        matcher = difflib.SequenceMatcher(None, s_words, p_words, autojunk=False)

        for tag, i1, i2, j1, j2 in matcher.get_opcodes():
            if tag == 'replace':
                self._highlight_text(self.source_text_widget, s_words, i1, i2, "replaced")
                self._highlight_text(self.processed_text_widget, p_words, j1, j2, "inserted")

        self.source_text_widget.config(state="disabled")
        self._update_block_counts(source_text, processed_text)

    def _highlight_text(self, widget, words, start_word, end_word, tag):
        text_before = " ".join(words[:start_word])
        start_index = f"1.0 + {len(text_before) + (1 if text_before else 0)} chars"
        text_to_end = " ".join(words[:end_word])
        end_index = f"1.0 + {len(text_to_end)} chars"
        widget.tag_add(tag, start_index, end_index)

    def _update_block_counts(self, source, processed):
        s_words, p_words = len(source.split()), len(processed.split())
        delta = (p_words - s_words) / s_words * 100 if s_words > 0 else float('inf') if p_words > 0 else 0.0

        # Get EID/SSID info if available
        selection_indices = self.blocks_listbox.curselection()
        eid_info = ""
        ssid_info = ""
        if selection_indices and self.all_blocks_data:
            block_data = self.all_blocks_data[selection_indices[0]]
            eid = block_data.get('eid', '')
            ssid = block_data.get('ssid', '')
            if eid:
                eid_info = f" | EID: {eid}"
            if ssid:
                ssid_info = f" | SSID: {ssid}"

        self.block_stats_label.config(
            text=f"Block Stats: Source ({s_words} w.) | Processed ({p_words} w.) | Δ: {delta:+.1f}%{eid_info}{ssid_info}")

    def _set_ui_state(self, is_busy: bool):
        self.process_button.config(state="disabled" if is_busy else "normal")
        self.stop_button.config(state="normal" if is_busy else "disabled")
        # NEVER disable save and database buttons - user should always be able to save or search
        self.save_button.config(state="normal")
        self.database_button.config(state="normal")
        print(f"[DEBUG] TEXT_PROCESSING: _set_ui_state called with is_busy={is_busy}. Process: {self.process_button.cget('state')}, Stop: {self.stop_button.cget('state')}, Save: {self.save_button.cget('state')}, Database: {self.database_button.cget('state')})")

    def _on_stop_click(self):
        self.stop_processing = True
        self._log("Stopping processing at user request.")
        self._update_status("Stopping...")
        # Force UI to unlock after a short delay in case worker doesn't respond
        self.after(1000, self._force_unlock_ui)

    def _force_unlock_ui(self):
        """Force unlock UI in case worker thread is stuck"""
        if self.stop_processing:
            self._set_ui_state(is_busy=False)
            self._stop_timer()
            self._update_status("Stopped")

    def _check_sheet_exists_and_confirm(self, db_path, sheet_name, is_auto=False):
        """Check if sheet exists and ask for confirmation to overwrite if not in auto mode."""
        if not db_path.exists() or not sheet_name:
            return True

        if is_auto:
            return True  # In auto-mode don't ask, just overwrite

        try:
            import openpyxl
            workbook = openpyxl.load_workbook(db_path)
            if sheet_name in workbook.sheetnames:
                response = messagebox.askyesno(
                    "Sheet already exists",
                    f"Sheet '{sheet_name}' already exists in database.\n\nOverwrite it?",
                    parent=self
                )
                workbook.close()
                return response
            workbook.close()
        except Exception as e:
            self._log(f"Could not check sheet in Excel: {e}", is_error=True)
            pass  # If error, allow to continue

        return True

    def _on_save_click(self, is_auto=False):
        if not self.all_blocks_data:
            if not is_auto:
                messagebox.showwarning("No data", "No data to save.", parent=self)
            return

        # Update text of current block from widget
        if not is_auto:
            current_selection = self.blocks_listbox.curselection()
            if current_selection:
                selected_idx = current_selection[0]
                self.all_blocks_data[selected_idx]['processed_text'] = self.processed_text_widget.get("1.0", "end-1c").strip()

        # Check if sheet exists BEFORE starting thread
        db_path = Path(self.controller.get_setting('work_root_path')) / "database" / "main_database.xlsx"
        vid_number_match = re.search(r'\d+', self.current_pid)
        sheet_name = f"VID{vid_number_match.group()}s" if vid_number_match else None

        if not self._check_sheet_exists_and_confirm(db_path, sheet_name, is_auto):
            self._update_status("Saving cancelled by user.")
            return

        # Check if search is needed or EID/SSID already present
        needs_search = self._check_if_database_search_needed()

        if needs_search:
            # Launch database search and subsequent save
            self._set_ui_state(is_busy=True)
            self._update_status("Searching in database...")
            settings = {
                'pid': self.current_pid,
                'projects_root_path': self.controller.get_setting('work_root_path')
            }
            threading.Thread(
                target=self._run_database_matching_and_save,
                args=(settings,),
                daemon=True
            ).start()
        else:
            # Skip search and save immediately
            self._set_ui_state(is_busy=True)
            self._update_status("Saving to Excel...")
            settings = {
                'pid': self.current_pid,
                'projects_root_path': self.controller.get_setting('work_root_path')
            }
            final_blocks_for_excel = list(self.all_blocks_data)
            threading.Thread(
                target=text_splitter_worker.save_results_to_excel,
                args=(settings, final_blocks_for_excel, self._update_status, self._log),
                daemon=True
            ).start()

    def _check_if_database_search_needed(self) -> bool:
        """
        Checks if database search needs to be launched.
        Returns False if all blocks (except intro) already have EID/SSID.
        """
        if not self.all_blocks_data:
            return True  # If no data, need to run search on save

        for block in self.all_blocks_data:
            original_caption = block.get('caption', '').strip()

            # Skip intro blocks
            if not original_caption or original_caption.lower() == 'intro':
                continue

            # If block has no EID and SSID, search is needed
            if not block.get('eid') or not block.get('ssid'):
                self._log("Blocks without EID/SSID found, starting database search...")
                return True

        self._log("All blocks already have EID/SSID, database search not required.")
        return False

    def _on_database_click(self):
        """Handle database matching button click."""
        if not self.all_blocks_data:
            messagebox.showwarning("No data", "No processed blocks for database search.", parent=self)
            return

        # Update any edited block data
        current_selection = self.blocks_listbox.curselection()
        if current_selection:
            selected_idx = current_selection[0]
            self.all_blocks_data[selected_idx]['processed_text'] = self.processed_text_widget.get("1.0", "end-1c").strip()

        self._set_ui_state(is_busy=True)
        self._update_status("Searching in database...")

        # Prepare settings for database matching
        settings = {
            'pid': self.current_pid,
            'projects_root_path': self.controller.get_setting('work_root_path')
        }

        # Run database matching in a separate thread
        threading.Thread(
            target=self._run_database_matching,
            args=(settings,),
            daemon=True
        ).start()

    def _run_database_matching(self, settings):
        """Run database matching in a separate thread."""
        try:
            # Call the database matching function
            result = text_splitter_worker.run_database_matching(
                settings,
                self.all_blocks_data,
                self._update_status,
                self._log,
                self.controller.ai_manager
            )

            # Handle result in main thread
            self.after(0, self._handle_database_result, result)

        except Exception as e:
            error_msg = f"Database search error: {str(e)}"
            self._log(error_msg, is_error=True)
        finally:
            self.after(0, self._set_ui_state, False)
            self.after(0, self._update_status, "Search error!")

    def _format_stats_message(self, stats, blocks_matched, blocks_total):
        """Formats the statistics for the messagebox."""
        if not stats:
            return f"Found matches: {blocks_matched} out of {blocks_total} blocks."

        artist_stats = stats.get('artist', {})
        song_stats = stats.get('song', {})

        message = f"""Database search complete.

Found matches: {blocks_matched} out of {blocks_total} blocks.

--- Artist Statistics ---
  • Exact matches found: {artist_stats.get('exact', 0)}
  • Found via AI: {artist_stats.get('ai', 0)}
  • New ones created: {artist_stats.get('created', 0)}

--- Song Statistics ---
  • Exact matches found: {song_stats.get('exact', 0)}
  • Found via AI: {song_stats.get('ai', 0)}
  • New ones created: {song_stats.get('created', 0)}"""
        return message

    def _handle_database_result(self, result):
        """Handle the result from database matching."""
        self._set_ui_state(is_busy=False)

        if result['status'] == 'error':
            messagebox.showerror("Error", f"Database search error:\n{result.get('message', 'Unknown error')}", parent=self)
            self._update_status("Database search error")
            return

        # Update the blocks with EID/SSID data
        self.all_blocks_data = result.get('final_blocks', self.all_blocks_data)

        blocks_matched = result.get('blocks_matched', 0)
        blocks_total = result.get('blocks_total', 0)
        stats = result.get('stats')

        self._update_status(f"Found {blocks_matched}/{blocks_total} matches")

        # Update current block display
        current_selection = self.blocks_listbox.curselection()
        if current_selection:
            self._on_block_select(None)

        # Show completion message with detailed stats
        messagebox.showinfo(
            "Search complete",
            self._format_stats_message(stats, blocks_matched, blocks_total),
            parent=self
        )

    def _run_database_matching_and_save(self, settings):
        """Run database matching then save to Excel in a separate thread."""
        try:
            # Call the database matching function first
            result = text_splitter_worker.run_database_matching(
                settings,
                self.all_blocks_data,
                self._update_status,
                self._log,
                self.controller.ai_manager
            )

            # Handle result in main thread
            self.after(0, self._handle_database_result_and_save, result)

        except Exception as e:
            error_msg = f"Database search error: {str(e)}"
            self._log(error_msg, is_error=True)
        finally:
            self.after(0, self._set_ui_state, False)
            self.after(0, self._update_status, "Search error!")

    def _handle_database_result_and_save(self, result):
        """Handle the result from database matching then save to Excel."""
        if result['status'] == 'error':
            messagebox.showerror("Error", f"Database search error:\n{result.get('message', 'Unknown error')}", parent=self)
            self._update_status("Database search error")
            self._set_ui_state(is_busy=False)
            return

        # Update the blocks with EID/SSID data
        self.all_blocks_data = result.get('final_blocks', self.all_blocks_data)
        blocks_matched = result.get('blocks_matched', 0)
        blocks_total = result.get('blocks_total', 0)
        stats = result.get('stats')

        self._update_status(f"Found {blocks_matched}/{blocks_total}. Saving...")

        # Now save to Excel with EID/SSID data
        final_blocks_for_excel = list(self.all_blocks_data)
        settings = {
            'pid': self.current_pid,
            'projects_root_path': self.controller.get_setting('work_root_path')
        }

        # Save to Excel in background thread
        threading.Thread(
            target=text_splitter_worker.save_results_to_excel,
            args=(settings, final_blocks_for_excel, self._update_status, self._log),
            daemon=True
        ).start()

        # Refresh the display to show EID/SSID info
        current_selection = self.blocks_listbox.curselection()
        if current_selection:
            self._on_block_select(None)


        # Show completion message
        messagebox.showinfo(
            "Search and save completed",
             self._format_stats_message(stats, blocks_matched, blocks_total),
            parent=self
        )


    def _clear_ui_widgets(self):
        self.pid_entry.delete(0, tk.END)
        self.blocks_listbox.delete(0, tk.END)
        for widget in [self.source_text_widget, self.processed_text_widget, self.log_text_widget]:
            widget.config(state="normal");
            widget.delete("1.0", tk.END)
        self.source_text_widget.config(state="disabled")
        self._update_status("")
        self.save_button.config(state="disabled")
        self.database_button.config(state="disabled")
        self.timer_label.config(text="")
        self.block_stats_label.config(text="Block Statistics")
        self.queue_label.config(text="")
        self.processed_pids = []
        self.queued_pids = []

    def _update_queue_display(self):
        processed = ", ".join(self.processed_pids) if self.processed_pids else "None"
        queued = ", ".join(self.queued_pids) if self.queued_pids else "None"
        text = f"Processed: {processed} | Queued: {queued}"
        self.after(0, lambda: self.queue_label.config(text=text))

    def _update_status(self, message: str):
        self.after(0, lambda: self.status_label.config(text=message))

    def _log(self, message: str, is_error: bool = False):
        timestamp = time.strftime('%H:%M:%S')
        pid_info = f"[{self.current_pid}]" if self.current_pid else ""
        log_message = f"[{timestamp}] [{'ERROR' if is_error else 'INFO'}] {pid_info} {message}\n"
        # Debug: print(f"[TextProcessingTab] _log called: {log_message.strip()}")

        def update_widget():
            # Debug: print(f"[TextProcessingTab] About to update widget with: {log_message.strip()}")
            self._update_log_widget(log_message)

        self.after(0, update_widget)

    def _update_log_widget(self, message: str):
        try:
            if hasattr(self, 'log_text_widget') and self.log_text_widget and self.log_text_widget.winfo_exists():
                # Save initial widget state (should be "normal")
                initial_state = self.log_text_widget.cget("state")
                # Debug: print(f"Widget initial state: {initial_state}")

                # Widget should be in "normal" state (per session)
                self.log_text_widget.config(state="normal")
                # Debug: print(f"Widget state after setting to normal: {self.log_text_widget.cget('state')}")

                # Add message
                self.log_text_widget.insert("end", message)
                self.log_text_widget.see("end")

                # Return to "normal" (widget is read-only)
                self.log_text_widget.config(state="normal")
                # Debug: print(f"Widget final state: {self.log_text_widget.cget('state')}")
                # Debug: print(f"SUCCESS: Added log message: {message.strip()}")
        except Exception as e:
            print(f"ERROR: Failed to update log widget: {e}")
            import traceback
            print(f"Traceback: {traceback.format_exc()}")

    # REMOVED: _setup_macos_focus_fix() - was causing button malfunctions