import json
import re
import unicodedata
from pathlib import Path
from datetime import datetime
import threading
import concurrent.futures
import time
import traceback
from typing import Dict, List, Optional, Tuple, Any, Callable
from openpyxl import load_workbook, Workbook
from fuzzywuzzy import fuzz
from pydub import AudioSegment
from pydub.playback import play

class EntityRecognizer:
    """Autonomous recognition of artists and songs based on Excel database with AI pre-parsing."""

    def __init__(self, app):
        self.app = app
        self.work_root = Path(app.WORK_ROOT_PATH)
        self.db_path = self.work_root / "database" / "main_database.xlsx"
        self.prompts_path = self.work_root / "common_assets" / "prompts"
        
        # Data cache
        self.artists_by_name = {} # norm_name -> {eid, name}
        self.songs_by_eid = {}    # eid -> list of {ssid, song_title, normalized_title}
        
        # Prompts
        self.artist_matching_prompt = self._load_prompt("script_artist_matching.txt")
        self.song_matching_prompt = self._load_prompt("script_song_matching.txt")
        self.artist_canonical_prompt = self._load_prompt("script_artist_canonical.txt")
        self.extract_list_prompt = self._load_prompt("script_extract_list.txt")

        self._load_databases()

    def _load_prompt(self, filename: str) -> str:
        p = self.prompts_path / filename
        return p.read_text(encoding='utf-8') if p.exists() else ""

    def _load_databases(self):
        if not self.db_path.exists(): return
        try:
            wb = load_workbook(self.db_path, data_only=True)
            if 'MasterDatabase_v1' in wb.sheetnames:
                ws = wb['MasterDatabase_v1']
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] and row[1]:
                        eid, name = str(row[0]), str(row[1])
                        self.artists_by_name[self._normalize_text(name).lower()] = {'eid': eid, 'name': name}
            if 'Songs_Database' in wb.sheetnames:
                ws = wb['Songs_Database']
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] and row[1] and row[3]:
                        ssid, title, eid = str(row[0]), str(row[1]), str(row[3])
                        if eid not in self.songs_by_eid: self.songs_by_eid[eid] = []
                        self.songs_by_eid[eid].append({'ssid': ssid, 'song_title': title, 'normalized_title': self._normalize_text(title)})
            wb.close()
        except Exception: pass

    def _normalize_text(self, text: str) -> str:
        if not text: return ""
        text = unicodedata.normalize('NFKD', str(text))
        text = ''.join([c for c in text if not unicodedata.combining(c)])
        text = text.lower()
        text = re.sub(r'[^a-z0-9\s]', '', text)
        return ' '.join(text.split())

    def _clean_caption_part(self, text: str) -> str:
        if not text: return ""
        text = re.sub(r'^(Number\s*)?\d+\.\s*', '', str(text)).strip()
        text = re.sub(r'^"|"$', '', text).strip()
        text = re.sub(r'\.\.\.$', '', text).strip()
        return text.strip().rstrip('.').strip()

    def find_artist_exact(self, artist_name: str) -> Optional[str]:
        norm = self._normalize_text(artist_name)
        if norm in self.artists_by_name: return self.artists_by_name[norm]['eid']
        v1, v2 = norm.replace(' and ', ' & '), norm.replace(' & ', ' and ')
        if v1 in self.artists_by_name: return self.artists_by_name[v1]['eid']
        if v2 in self.artists_by_name: return self.artists_by_name[v2]['eid']
        return None

    def find_artist_ai(self, artist_name: str, candidates: List[Dict]) -> Optional[str]:
        scored = []
        for c in candidates:
            score = fuzz.token_set_ratio(artist_name.lower(), c['name'].lower())
            if score >= 45: scored.append((score, c))
        scored.sort(key=lambda x: x[0], reverse=True)
        filtered = [item[1] for item in scored[:100]]
        if not filtered: return None
        cand_text = "\n".join([f"{i+1}. {c['name']}" for i, c in enumerate(filtered)])
        prompt = self.artist_matching_prompt.format(artist_name=artist_name, candidate_text=cand_text, context="Music")
        try:
            res = self.app.ai_manager.execute_ai_task("text_processing", {"prompt": prompt}).get("text", "").strip()
            match = re.search(r'^(\d+)', res)
            if match:
                idx = int(match.group(1)) - 1
                if 0 <= idx < len(filtered): return filtered[idx]['eid']
        except Exception: pass
        return None

    def find_song_exact(self, eid: str, song_title: str) -> Optional[str]:
        if eid not in self.songs_by_eid: return None
        norm = self._normalize_text(song_title)
        for s in self.songs_by_eid[eid]:
            if s['normalized_title'] == norm: return s['ssid']
        return None

    def find_song_fuzzy(self, eid: str, song_title: str, threshold: int = 85) -> Optional[str]:
        if eid not in self.songs_by_eid: return None
        target_norm = self._normalize_text(song_title)
        best_match, max_s = None, 0
        for s in self.songs_by_eid[eid]:
            score = fuzz.ratio(target_norm, s['normalized_title'])
            if score > max_s: max_s, best_match = score, s['ssid']
        return best_match if max_s >= threshold else None

    def find_song_ai(self, eid: str, song_title: str, songs: List[Dict]) -> Optional[str]:
        if not songs: return None
        artist = self.get_artist_name_by_eid(eid)
        cand_text = "\n".join([f"{i+1}. {s['song_title']}" for i, s in enumerate(songs)])
        prompt = self.song_matching_prompt.format(song_title=song_title, artist_name=artist, candidate_text=cand_text, context="Search")
        try:
            res = self.app.ai_manager.execute_ai_task("text_processing", {"prompt": prompt}).get("text", "").strip()
            match = re.search(r'^(\d+)', res)
            if match:
                idx = int(match.group(1)) - 1
                if 0 <= idx < len(songs): return songs[idx]['ssid']
        except Exception: pass
        return None

    def get_artist_name_by_eid(self, eid: str) -> str:
        for data in self.artists_by_name.values():
            if data['eid'] == eid: return data['name']
        return "Unknown"

    def get_song_title_by_ssid(self, ssid: str) -> str:
        for songs in self.songs_by_eid.values():
            for s in songs:
                if s['ssid'] == ssid: return s['song_title']
        return ""

    def _get_canonical_data(self, artist_name: str) -> Tuple[str, str]:
        """AI request for canonical name and role."""
        try:
            prompt = self.artist_canonical_prompt.format(artist_name=artist_name, context=artist_name)
            res = self.app.ai_manager.execute_ai_task("text_processing", {"prompt": prompt}).get("text", "")
            m_name = re.search(r'CANONICAL_NAME:\s*(.+)', res, re.IGNORECASE)
            m_role = re.search(r'ROLE:\s*(\w+)', res, re.IGNORECASE)
            clean_name = self._clean_caption_part(m_name.group(1)) if m_name else artist_name
            role = m_role.group(1).lower() if m_role else "artist"
            return clean_name, role
        except Exception: 
            return artist_name, "artist"

    def check_list(self, raw_text: str, log_callback: Callable[[str], None]) -> list:
        results = []
        try:
            log_callback("[1/2] Intelligent list splitting via AI...")
            final_prompt = f"{self.extract_list_prompt}\n\nINPUT:\n{raw_text}"
            
            print(f"\n>>> OUTGOING PROMPT [text_processing]:\n{final_prompt}\n")
            ai_res = self.app.ai_manager.execute_ai_task("text_processing", {"prompt": final_prompt})
            
            if "error" in ai_res and ai_res["error"]:
                print(f"!!! AI ERROR: {ai_res['error']}")
                log_callback(f"AI ERROR: {ai_res['error']}")
                return []

            ai_text = ai_res.get("text", "")
            print(f"\n<<< INCOMING RESPONSE [text_processing]:\n{ai_text}\n")

            clean_pairs = []
            if ai_text:
                json_match = re.search(r'\[.*\]', re.sub(r'^```json\s*|```$', '', ai_text, flags=re.MULTILINE), re.DOTALL)
                if json_match: 
                    try:
                        clean_pairs = json.loads(json_match.group(0))
                    except Exception as je:
                        print(f"!!! JSON PARSE ERROR: {je}")

            if not clean_pairs: 
                log_callback("ERROR: AI could not split the list."); return []

            log_callback(f"[2/2] Matching with database...")
            for idx, pair in enumerate(clean_pairs):
                orig_artist = self._clean_caption_part(pair.get("artist", ""))
                orig_song = self._clean_caption_part(pair.get("song", ""))
                
                # --- ARTIST SEARCH ---
                eid = self.find_artist_exact(orig_artist)
                a_status = "found"
                if not eid:
                    eid = self.find_artist_ai(orig_artist, list(self.artists_by_name.values()))
                    if not eid:
                        a_status = "new"
                else: artist_name = self.get_artist_name_by_eid(eid)
                else: artist_name = self.get_artist_name_by_eid(eid)

                # --- SONG SEARCH ---
                ssid, s_status = None, "new"
                song_title = orig_song # Dirty by default
                if eid and orig_song:
                    ssid = self.find_song_exact(eid, orig_song)
                    if not ssid:
                        ssid = self.find_song_fuzzy(eid, orig_song)
                        if not ssid:
                            ssid = self.find_song_ai(eid, orig_song, self.songs_by_eid.get(eid, []))
                    
                    if ssid:
                        s_status = "found"
                        song_title = self.get_song_title_by_ssid(ssid)
                
                results.append({"artist": artist_name, "eid": eid, "artist_status": a_status, "song": song_title, "ssid": ssid, "song_status": s_status})
                
                # Generate log showing replacement
                a_log = f"{orig_artist} -> {artist_name} ({eid or 'NEW'})" if orig_artist != artist_name else f"{artist_name} ({eid or 'NEW'})"
                s_log = f"{orig_song} -> {song_title} ({ssid or 'NEW'})" if orig_song != song_title else f"{song_title} ({ssid or 'NEW'})"
                log_callback(f"Row {idx+1}: {a_log} | {s_log}")
            
            return results
        except Exception as e: 
            log_callback(f"CRITICAL ERROR: {e}"); return []

class SeedManager:
    """Seed data management and Excel synchronization."""

    @staticmethod
    def _play_alert(app, sound_type):
        """Plays a sound alert (success/error)."""
        def play_sound():
            try:
                alerts_path = Path(app.SOUND_ALERTS_PATH)
                filename = "generic_success.mp3" if sound_type == "success" else "generic_error.mp3"
                sound_path = alerts_path / filename
                
                if sound_path.exists():
                    audio = AudioSegment.from_mp3(str(sound_path))
                    play(audio)
            except Exception as e:
                print(f"Sound play error: {e}")

        threading.Thread(target=play_sound, daemon=True).start()

    @staticmethod
    def check_excel_lock(db_path: Path) -> bool:
        """Checks if Excel file is locked (opened in application)."""
        # 1. Check for temporary lock file created by Excel
        lock_file = db_path.parent / f"~${db_path.name}"
        if lock_file.exists():
            return True
        
        # 2. Attempt exclusive open for read/write
        try:
            if db_path.exists():
                with open(db_path, "r+b") as f:
                    pass
            return False
        except (PermissionError, IOError):
            return True

    @staticmethod
    def load_archetype_data(app, archetype_id: str, version: str = "v1") -> dict:
        """Reads archetype data from version-specific file (archetypes_master_data_{version}.txt)."""
        try:
            # Search ONLY current version file
            filename = f"archetypes_master_data_{version}.txt"
            master_file = Path(app.WORK_ROOT_PATH) / "common_assets" / "prompts" / filename

            if not master_file.exists():
                # If file not found - return empty dict so caller can error out
                return {}

            content = master_file.read_text(encoding='utf-8')
            # Regex to find block: from [ARCHETYPE_ID: N] to next block or end of file
            block_pattern = rf'(?i)\[ARCHETYPE_ID:\s*{re.escape(str(archetype_id))}\](.*?)(?=\[ARCHETYPE_ID:|\Z)'
            match_block = re.search(block_pattern, content, re.DOTALL)

            if not match_block:
                return {}

            data_block = match_block.group(1)
            result = {"id": archetype_id}

            # Parameters: find key at start of line, case-insensitive, take everything until EOL
            patterns = {
                "target_length": r'(?i)^TARGET_LENGTH:\s*(.*)',
                "beat_01": r'(?i)^\[BEAT_01\]:\s*(.*)',
                "beat_02": r'(?i)^\[BEAT_02\]:\s*(.*)',
                "beat_03": r'(?i)^\[BEAT_03\]:\s*(.*)',
                "beat_04": r'(?i)^\[BEAT_04\]:\s*(.*)',
                "beat_05": r'(?i)^\[BEAT_05\]:\s*(.*)',
                "writer_instruction": r'(?i)^WRITER_INSTRUCTION:\s*(.*)',
                "benchmark_story": r'(?i)^BENCHMARK_STORY:\s*(.*)'
            }

            for key, pattern in patterns.items():
                m = re.search(pattern, data_block, re.MULTILINE)
                if m:
                    result[key] = m.group(1).strip().replace('\r', '')
            return result
        except Exception as e:
            print(f"Error loading archetype data: {e}")
            return {}

    @staticmethod
    def load_seed(work_root: str, pid_input: str) -> dict:
        if not work_root or not pid_input: return None
        pid_input = pid_input.strip()
        filename = pid_input
        if not filename.upper().startswith("PID"): filename = f"PID{filename}"
        if not filename.endswith("_seed.json"): filename = f"{filename}_seed.json"
        p = Path(work_root) / "database" / "seed" / filename
        if p.exists():
            try:
                with open(p, 'r', encoding='utf-8') as f: return json.load(f)
            except Exception: return None
        return None

    @staticmethod
    def save_seed(app, data_dict: dict, user_pid_input: str, version_selected: str, tree_data: list, archetype_number: str = "") -> dict:
        work_root = Path(app.WORK_ROOT_PATH)
        seed_dir = work_root / "database" / "seed"; seed_dir.mkdir(parents=True, exist_ok=True)
        recognizer = EntityRecognizer(app)
        db_path = work_root / "database" / "main_database.xlsx"
        
        # --- CHECK EXCEL FILE LOCK ---
        if SeedManager.check_excel_lock(db_path):
            SeedManager._play_alert(app, "error")
            return {
                "status": "error", 
                "message": "DATABASE BUSY. Close main_database.xlsx in Excel and try again."
            }

        total_songs = len(tree_data)
        
        # Phase 1 (Excel): Registering new entities and saving workbook
        try:
            wb = load_workbook(db_path)
            for row in tree_data:
                artist_name, eid, a_status = row['artist'], row['eid'], row['artist_status']
                song_title, ssid, s_status = row['song'], row['ssid'], row['song_status']
                
                # Write new to Excel (names already canonicalized in check_list)
                if a_status == "new":
                    eid = SeedManager._create_new_artist_in_wb(wb, app, recognizer, artist_name)
                    row['eid'] = eid  # Update in tree_data for JSON
                
                if s_status == "new" and eid:
                    ssid = SeedManager._create_new_song_in_wb(wb, app, recognizer, eid, song_title, artist_name)
                    row['ssid'] = ssid  # Update in tree_data for JSON
            
            wb.save(db_path)
            wb.close()
        except Exception as e:
            SeedManager._play_alert(app, "error")
            return {"status": "error", "message": f"Excel save error: {e}"}

        # Phase 2 (JSON): Forming final list and saving seed
        final_songs = []
        for idx, row in enumerate(tree_data):
            title_num = total_songs - idx
            final_songs.append({
                "index": idx + 1, 
                "title_number": f"{title_num}. ", 
                "artist": row['artist'], 
                "eid": row['eid'], 
                "song": row['song'], 
                "ssid": row['ssid']
            })

        # PID Logic
        user_pid_input = user_pid_input.strip()
        final_pid = ""
        if re.search(r'\d+v\d+', user_pid_input): final_pid = user_pid_input
        elif re.fullmatch(r'\d{4}', user_pid_input): final_pid = f"{user_pid_input}{version_selected}"
        elif re.fullmatch(r'\d', user_pid_input):
            pids = [int(re.search(r'PID(\d+)', f.name).group(1)) for f in seed_dir.glob(f"PID{user_pid_input}*_seed.json") if re.search(r'PID(\d+)', f.name)]
            next_n = max(pids) + 1 if pids else int(f"{user_pid_input}001")
            final_pid = f"{next_n}{version_selected}"
        else:
            pids = [int(re.search(r'PID(\d+)', f.name).group(1)) for f in seed_dir.glob("PID*_seed.json") if re.search(r'PID(\d+)', f.name)]
            next_n = max(pids) + 1 if pids else 1
            final_pid = f"{next_n:04d}{version_selected}"

        filename = f"PID{final_pid}_seed.json" if not final_pid.upper().startswith("PID") else f"{final_pid}_seed.json"
        seed_data = {"metadata": {"pid": final_pid, "version": version_selected, "archetype_number": archetype_number, "date": datetime.now().isoformat(), "total_songs": total_songs}, "songs": final_songs, "script_settings": {"min_words": app.get_setting("script_creation", {}).get("min_words", "14")}} 
        for key in ["inputs", "perspective", "titles_desc"]:
            if key in data_dict: seed_data[key] = data_dict[key]

        # Extract Archetype for metadata
        if "perspective" in seed_data and "selected" in seed_data["perspective"]:
            pers_text = seed_data["perspective"]["selected"]
            arch_match = re.search(r'ARCHETYPE:\s*(.+)', pers_text, re.IGNORECASE)
            if arch_match:
                seed_data["metadata"]["archetype"] = arch_match.group(1).strip()

        try:
            with open(seed_dir / filename, 'w', encoding='utf-8') as f:
                json.dump(seed_data, f, indent=4, ensure_ascii=False)
        except Exception as e:
            SeedManager._play_alert(app, "error")
            return {"status": "error", "message": f"JSON save error: {e}"}

        SeedManager._play_alert(app, "success")
        return {"status": "success", "filename": filename, "pid": final_pid}
    
    @staticmethod
    def save_seed_json(app, seed_data: dict) -> bool:
        """Saves seed data to JSON without touching Excel (Atomic-like for Research Phase)."""
        try:
            pid = seed_data.get("metadata", {}).get("pid", "")
            if not pid: return False
            filename = f"PID{pid}_seed.json" if not pid.upper().startswith("PID") else f"{pid}_seed.json"
            seed_dir = Path(app.WORK_ROOT_PATH) / "database" / "seed"
            seed_dir.mkdir(parents=True, exist_ok=True)
            file_path = seed_dir / filename
            
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(seed_data, f, indent=4, ensure_ascii=False)
            return True
        except Exception as e: 
            print(f"Error saving seed JSON: {e}")
            return False

    @staticmethod
    def _create_new_artist_in_wb(wb, app, recognizer, artist_name) -> str:
        # Name is already clean, get only role for Excel
        _, role = recognizer._get_canonical_data(artist_name)
        ws = wb['MasterDatabase_v1']; max_eid = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                num = int(str(row[0])[3:])
                if num > max_eid: max_eid = num
            except: pass
        new_eid = f"EID{max_eid + 1:04d}"; ws.append([new_eid, artist_name, role, 1]); return new_eid

    @staticmethod
    def _create_new_song_in_wb(wb, app, recognizer, eid, song_title, artist_name) -> str:
        ws = wb['Songs_Database']; max_ssid = 0
        for r in ws.iter_rows(min_row=2, values_only=True):
            try:
                num = int(str(r[0])[4:])
                if num > max_ssid: max_ssid = num
            except: pass
        new_ssid = f"SSID{max_ssid + 1:04d}"
        ws.append([new_ssid, song_title, None, eid, artist_name]); return new_ssid

class Brainstormer:
    """AI interaction logic"""

    @staticmethod
    def parse_chat_into_messages(full_text: str) -> list:
        messages = []
        parts = re.split(r'(>>> (?:ASSISTANT|USER):)', full_text)
        current_role = None
        for part in parts:
            part = part.strip()
            if not part: continue
            if part == ">>> ASSISTANT:": current_role = "assistant"
            elif part == ">>> USER:": current_role = "user"
            else:
                if current_role: messages.append({"role": current_role, "content": part})
        return messages

    @staticmethod
    def chat_with_ai(app, full_chat_text: str, songs_list: str, chronology: str, version: str = "v1") -> str:
        work_root = app.WORK_ROOT_PATH
        # Version-aware prompt loading
        prompt_file = Path(work_root) / "common_assets" / "prompts" / f"script_brainstorm_theme_{version}.txt"
        if not prompt_file.exists():
            # Fallback to original
            prompt_file = Path(work_root) / "common_assets" / "prompts" / "script_brainstorm_theme.txt"
            
        if not prompt_file.exists(): return "Error: Prompt file not found."
        template = prompt_file.read_text(encoding='utf-8')
        context = f"CONTEXT: Chronology: {chronology}. Songs: {songs_list}."
        messages = Brainstormer.parse_chat_into_messages(full_chat_text)
        history_str = ""
        if messages:
            for msg in messages:
                role_label = "USER" if msg['role'] == 'user' else "ASSISTANT"
                history_str += f"{role_label}: {msg['content']}\n\n"
        final_prompt = f"{template}\n\n{context}" if not history_str else f"{template}\n\n{context}\n\nCURRENT DIALOG:\n{history_str}\n\nASSISTANT, reply to the last line."
        result = app.ai_manager.execute_ai_task("idea_generation", {"prompt": final_prompt})
        return result.get("text", "AI returned an empty response.")

    @staticmethod
    def generate_titles(app, songs_list: str, chronology: str, selected_perspective: str, version: str = "v1") -> dict:
        work_root = app.WORK_ROOT_PATH
        # Version-aware prompt loading
        prompt_file = Path(work_root) / "common_assets" / "prompts" / f"script_brainstorm_title_{version}.txt"
        if not prompt_file.exists():
             prompt_file = Path(work_root) / "common_assets" / "prompts" / "script_brainstorm_title.txt"

        if not prompt_file.exists(): return {"error": "Prompt file not found."}
        template = prompt_file.read_text(encoding='utf-8')
        final_prompt = template.replace("[CHRONOLOGY]", chronology).replace("[SELECTED_PERSPECTIVE]", selected_perspective).replace("[LIST_OF_SONGS]", songs_list)
        result = app.ai_manager.execute_ai_task("idea_generation", {"prompt": final_prompt})
        raw_text = result.get("text", "").strip()
        clean_json_text = re.sub(r'^```json\s*|```$', '', raw_text, flags=re.MULTILINE)
        try:
            json_match = re.search(r'\{.*\}', clean_json_text, re.DOTALL)
            if json_match:
                data = json.loads(json_match.group(0))
                data["raw_debug"] = raw_text
                return data
            return {"error": "AI did not return JSON.", "raw_debug": raw_text}
        except Exception as e: return {"error": f"Parse error: {e}", "raw_debug": raw_text}

class ResearchAgent:
    """Agent for mining facts about songs using AI."""
    
    @staticmethod
    def mine_song_facts(app, artist, song, period, perspective, prompt_path, assembled_prompt=None) -> dict:
        try:
            if assembled_prompt:
                final_prompt = assembled_prompt.replace("[SONG]", song).replace("[ARTIST]", artist).replace("[CHRONOLOGY]", period).replace("[PERSPECTIVE]", perspective)
            else:
                p_path = Path(prompt_path)
                if p_path.exists():
                    prompt_template = p_path.read_text(encoding='utf-8')
                else:
                    prompt_template = ""
                # Construct the final prompt
                final_prompt = (
                    f"{prompt_template}\n\n"
                    f"TARGET: \"{song}\" by {artist}\n"
                    f"CHRONOLOGY: {period}\n"
                    f"UNIFYING PRINCIPLE: {perspective}\n\n"
                    "Provide the output strictly in the specified format (FIELD_NAME: Value)."
                )

        except Exception:
            final_prompt = ""

        try:
            # Logging requirement 3.3
            print(f"\n>>> OUTGOING PROMPT [fact_research]:\n{final_prompt}\n")
            
            # Execute AI task
            response = app.ai_manager.execute_ai_task("fact_research", {"prompt": final_prompt})
            text = response.get("text", "")
            
            print(f"\n<<< INCOMING RESPONSE [fact_research]:\n{text}\n")
            
            # Debug output to console as requested
            print(f"--- AI RESPONSE FOR {song} ---\n{text}\n-----------------------------")

            facts = {}
            # Universal parser: accepts any UPPERCASE_KEY: Value
            for line in text.split('\n'):
                if ':' in line:
                    parts = line.split(':', 1)
                    key = parts[0].strip()
                    val = parts[1].strip()
                    # Check if key is uppercase (allowing spaces/underscores)
                    if key.isupper() and len(key) > 1:
                        facts[key] = val
            
            if not facts: # Fallback parsing
                 for line in text.splitlines():
                    if ':' in line:
                        parts = line.split(':', 1)
                        key = parts[0].strip()
                        val = parts[1].strip()
                        if key.isupper() and len(key) > 1:
                            facts[key] = val
                            
            return facts
        except Exception as e: 
            print(f"Error mining facts for {song}: {e}")
            return {}

class CompositionAgent:
    @staticmethod
    def synthesize_script(app, seed_data, arch_data=None):
        try:
            songs = seed_data.get("songs", [])
            total_songs = len(songs) # Determine base for countdown
            data_sets_str = ""
            for idx, s in enumerate(songs):
                # Using index as 1-based index from list
                idx_num = idx + 1
                facts_str = json.dumps(s.get("research_data", {}), ensure_ascii=False)
                # CALCULATION: first song in list (SONG [1]) should get number 13
                # second (SONG [2]) - 12, and so on.
                display_num = total_songs - idx
                data_sets_str += f"SONG [{idx_num}]: {s.get('artist')} - {s.get('song')}. [POSITION_NUMBER]: {display_num}. FACTS: {facts_str}\n"

            perspective = seed_data.get("perspective", {}).get("selected", "")
            chronology = seed_data.get("inputs", {}).get("chronology", "")
            version = seed_data.get("metadata", {}).get("version", "v1")
            
            # Version-aware prompt loading (Template-first standard)
            prompt_path = Path(app.WORK_ROOT_PATH) / "common_assets" / "prompts" / f"script_writer_industrial_{version}_template.txt"
            if not prompt_path.exists():
                 # Fallback to legacy path
                 prompt_path = Path(app.WORK_ROOT_PATH) / "common_assets" / "prompts" / f"script_writer_industrial_{version}.txt"
            if not prompt_path.exists():
                 # Final fallback
                 prompt_path = Path(app.WORK_ROOT_PATH) / "common_assets" / "prompts" / "script_writer_industrial.txt"

            if prompt_path.exists():
                prompt_content = prompt_path.read_text(encoding='utf-8')
            else:
                return None

            # Template assembly (Monolith logic)
            if arch_data:
                prompt_content = prompt_content.replace("{{TARGET_LENGTH}}", arch_data.get("target_length", "100-120 words"))
                prompt_content = prompt_content.replace("{{WRITER_INSTRUCTION}}", arch_data.get("writer_instruction", ""))
                prompt_content = prompt_content.replace("{{BENCHMARK_STORY}}", arch_data.get("benchmark_story", ""))

            full_prompt = prompt_content.replace("[SELECTED_PERSPECTIVE]", perspective).replace("[CHRONOLOGY]", chronology).replace("[DATA_SETS]", data_sets_str)
            
            # Logging requirement 3.3
            print(f"\n>>> OUTGOING PROMPT [script_writing]:\n{full_prompt}\n")

            response = app.ai_manager.execute_ai_task("script_writing", {"prompt": full_prompt})
            response_text = response.get("text", "")

            print(f"\n<<< INCOMING RESPONSE [script_writing]:\n{response_text}\n")

            # Parsing
            clean_text = response_text.replace(">>>", "").replace(">>", "").strip()
            
            # Extract Blocks - Flexible regex to catch any ID format (like B01_FINAL)
            blocks = []
            block_matches = re.findall(r'\[BLOCK_START:\s*([^\]]+)\](.*?)(?=\[BLOCK_START:|\[BLOCK_END\]|\[/BLOCK_START\]|\[/BLOCK_END\]|\[INTRO_TEASERS_DATA\]|\Z)', clean_text, re.DOTALL)
            for b_id, b_text in block_matches:
                 # Remove only hallucinated tags and Markdown trash
                 # WE DO NOT TOUCH regular square brackets inside the text!
                 clean_b_text = b_text.strip()
                 clean_b_text = re.sub(r'\[/BLOCK_START\]|\[BLOCK_END\]|\[/BLOCK_END\]', '', clean_b_text)
                 clean_b_text = clean_b_text.replace('#', '').replace('---', '').strip()

                 blocks.append({"block_id": b_id.strip(), "text": clean_b_text})
            
            if not blocks:
                return {"error": "non_informative", "raw_text": response_text}

            # Extract Teasers
            teasers_indices = []
            teaser_match = re.search(r'\[INTRO_TEASERS_DATA\]\s*\[(.*?)\]', clean_text)
            if teaser_match:
                try:
                    indices_str = teaser_match.group(1)
                    if indices_str.strip():
                        teasers_indices = [int(x.strip()) for x in indices_str.split(',') if x.strip().isdigit()]
                except Exception: pass

            seed_data["raw_script_output"] = response_text
            seed_data["parsed_blocks"] = blocks
            seed_data["teaser_indices"] = teasers_indices
            
            with ScriptOrchestrator._file_lock:
                 SeedManager.save_seed_json(app, seed_data)
                 
            return seed_data
        except Exception as e: 
            print(f"Error in synthesize_script: {e}")
            traceback.print_exc()
            return None

class ExportManager:
    @staticmethod
    def _split_text_programmatically(text: str, min_words: int = 14) -> List[str]:
        """
        Copied from version_creation_worker.py
        """
        abbreviations = ["Mr.", "Mrs.", "Ms.", "Dr.", "St.", "Jr.", "Sr.", "vs.", "approx.", "incl.", "etc.", 
                        "Jan.", "Feb.", "Mar.", "Apr.", "Jun.", "Jul.", "Aug.", "Sep.", "Oct.", "Nov.", "Dec."]
        
        # 1. Protect abbreviations by temporarily replacing dot with unique marker
        protected_text = text
        for abbr in abbreviations:
            # Protect only isolated words (using \b), preserving original case
            prefix = abbr.rstrip('.')
            pattern = rf"\b({re.escape(prefix)})\."
            protected_text = re.sub(pattern, r"\1|DOT|", protected_text, flags=re.IGNORECASE)
        
        # 2. Now safely split by dots/questions/exclamations followed by space and capital letter
        # No variable look-behind here, logic is simple
        pattern = r'(?<=[.!?])\s+(?=[A-Z"“])|(?<=[.!?]["”])\s+(?=[A-Z"“])'
        sentences = re.split(pattern, protected_text.strip())
        
        final_sentences = []
        for s in sentences:
            if s.strip():
                final_sentences.append(s.replace("|DOT|", ".").strip())

        fragments = []
        buffer = []
        current_word_count = 0

        for sentence in final_sentences:
            words = sentence.split()
            buffer.append(sentence)
            current_word_count += len(words)

            if current_word_count >= min_words:
                fragments.append(" ".join(buffer))
                buffer = []
                current_word_count = 0

        if buffer:
            if fragments:
                fragments[-1] = fragments[-1] + " " + " ".join(buffer)
            else:
                fragments.append(" ".join(buffer))

        return fragments

    @staticmethod
    def process_and_export(app, pid, seed_data):
        db_path = Path(app.WORK_ROOT_PATH) / "database" / "main_database.xlsx"
        if SeedManager.check_excel_lock(db_path):
            return False

        parsed_blocks = seed_data.get("parsed_blocks", [])
        songs = seed_data.get("songs", [])
        
        # Validation
        if not parsed_blocks:
            return False
            
        # Check for B01 presence
        if not any(b['block_id'] == 'B01' for b in parsed_blocks):
            return False
            
        # Check block count: songs count + 1 (intro)
        if len(parsed_blocks) < len(songs) + 1:
            return False

        try:
            min_words_val = int(seed_data.get("script_settings", {}).get("min_words", "14"))
            parsed_blocks = seed_data.get("parsed_blocks", [])
            
            # Sort blocks to ensure order (B01, B02...)
            parsed_blocks.sort(key=lambda x: x['block_id'])
            
            songs = seed_data.get("songs", [])
            teaser_indices = seed_data.get("teaser_indices", [])

            processed_fragments = []
            
            for block in parsed_blocks:
                b_id = block["block_id"]
                b_text = block["text"]
                
                fragments = ExportManager._split_text_programmatically(b_text, min_words=min_words_val)
                
                ssid = ""
                eid = ""
                caption = ""
                
                if b_id.startswith("B") and b_id != "B01":
                    try:
                        # Robust number extraction: find the first sequence of digits
                        num_match = re.search(r'(\d+)', b_id)
                        if num_match:
                            val = int(num_match.group(1))
                            # Standard logic: B02 -> index 0 (Song 1)
                            song_idx = val - 2
                            if 0 <= song_idx < len(songs):
                                s_data = songs[song_idx]
                                ssid = s_data.get("ssid", "")
                                eid = s_data.get("eid", "")
                                # Caption Format: 13. "Song" by Artist
                                title_num = s_data.get('title_number', '') 
                                song_name = s_data.get('song', '')
                                artist_name = s_data.get('artist', '')
                                caption = f'{title_num}"{song_name}" by {artist_name}'
                    except: pass
                
                for i, frag_text in enumerate(fragments):
                    frag_obj = {
                        "block_id": f"{b_id}-{i+1:02d}",
                        "text": frag_text,
                        "ssid": ssid if i == 0 else "",
                        "eid": eid if i == 0 else "",
                        "caption": caption if i == 0 else "",
                    }
                    processed_fragments.append(frag_obj)

            intro_data_rows = []
            for t_idx in teaser_indices:
                list_idx = t_idx - 1
                if 0 <= list_idx < len(songs):
                    s_item = songs[list_idx]
                    orig_block_id = f"B{list_idx + 2:02d}"
                    
                    intro_row = {
                        "intro_song_number": t_idx,
                        "song_name": s_item.get("song"),
                        "ssid": s_item.get("ssid"),
                        "artist": s_item.get("artist"),
                        "eid": s_item.get("eid"),
                        "orig_block_id": orig_block_id,
                        "id_source": "Own Assignment"
                    }
                    intro_data_rows.append(intro_row)

            if not db_path.exists(): return False
            
            wb = load_workbook(db_path)
            sheet_name = f"VID{pid}"
            
            if sheet_name in wb.sheetnames:
                 del wb[sheet_name]
            
            ws = wb.create_sheet(sheet_name)
            
            ws['A1'] = "Block ID"
            ws['B1'] = "Text"
            ws['C1'] = "SSID"
            ws['D1'] = "EID"
            ws['F1'] = "url"
            ws['K1'] = "Caption"
            
            ws['P1'] = "Intro Song Number"
            ws['Q1'] = "Intro Song Name"
            ws['R1'] = "SSID"
            ws['S1'] = "Intro Artist"
            ws['T1'] = "EID"
            ws['U1'] = "Original Block ID"
            ws['V1'] = "Identification Source"
            
            ws['F2'] = "own text"
            
            for i, frag in enumerate(processed_fragments):
                r = i + 3
                ws[f'A{r}'] = frag['block_id']
                ws[f'B{r}'] = frag['text']
                ws[f'C{r}'] = frag['ssid']
                ws[f'D{r}'] = frag['eid']
                ws[f'K{r}'] = frag['caption']

            for i, item in enumerate(intro_data_rows):
                r = i + 3
                ws[f'P{r}'] = i + 1
                ws[f'Q{r}'] = item['song_name']
                ws[f'R{r}'] = item['ssid']
                ws[f'S{r}'] = item['artist']
                ws[f'T{r}'] = item['eid']
                ws[f'U{r}'] = item['orig_block_id']
                ws[f'V{r}'] = item['id_source']

            try:
                wb.save(db_path)
                return len(processed_fragments)
            except:
                return False

        except Exception as e: 
            print(f"Export Error: {e}")
            traceback.print_exc()
            return False

class ScriptOrchestrator:
    """Orchestrates the script creation process."""
    
    _file_lock = threading.Lock()

    @staticmethod
    def run_full_chain(app, pid, log_callback):
        """Main entry point for the script creation chain."""
        try:
            log_callback(f"=== SCRIPT CREATION CHAIN START (PID: {pid}) ===")
            
            seed_data = SeedManager.load_seed(app.WORK_ROOT_PATH, pid)
            if not seed_data:
                log_callback("Failed to load seed data.")
                return

            # 1. IMMEDIATELY extract version from project. This is the Single Point of Truth.
            version = seed_data.get("metadata", {}).get("version", "v1")

            # 2. Extract Archetype ID
            archetype_id = seed_data.get("metadata", {}).get("archetype_number")

            arch_data = {}
            if archetype_id:
                # 3. Now load archetype data using GUARANTEED correct version
                arch_data = SeedManager.load_archetype_data(app, archetype_id, version)
                if not arch_data:
                    log_callback(f"ERROR: Handbook archetypes_master_data_{version}.txt not found or Archetype {archetype_id} is missing in it. Execution stopped.")
                    return
                log_callback(f"Loaded archetype data ID: {archetype_id} ({version})")

            # --- ASSEMBLY LOGIC (MONOLITH) ---
            scheme_info = "STANDARD (5 beats)"
            fallback_reason = None

            if not archetype_id:
                fallback_reason = "Archetype ID not specified in project"
            elif not arch_data:
                fallback_reason = f"Archetype {archetype_id} not found in handbook or parsing error"

            detective_prompt = None
            if not fallback_reason:
                template_path = Path(app.WORK_ROOT_PATH) / "common_assets" / "prompts" / f"script_research_detective_{version}_template.txt"
                if template_path.exists():
                    template = template_path.read_text(encoding='utf-8')
                    # Check for beats presence (min 3, optionally 4 and 5)
                    b1, b2, b3 = arch_data.get("beat_01"), arch_data.get("beat_02"), arch_data.get("beat_03")
                    b4, b5 = arch_data.get("beat_04"), arch_data.get("beat_05")
                    if b1 and b2 and b3:
                        detective_prompt = template.replace("{{BEAT_01_TASK}}", b1).replace("{{BEAT_02_TASK}}", b2).replace("{{BEAT_03_TASK}}", b3)
                        beat_count = 3
                        if b4:
                            detective_prompt = detective_prompt.replace("{{BEAT_04_TASK}}", b4)
                            beat_count = 4
                        if b5:
                            detective_prompt = detective_prompt.replace("{{BEAT_05_TASK}}", b5)
                            beat_count = 5
                        scheme_info = f"ASSEMBLY {version.upper()} ({beat_count} beats)"
                    else:
                        fallback_reason = "Handbook is missing [BEAT_01-03] tasks"
                else:
                    fallback_reason = f"Template file {version}_template not found"

            if fallback_reason:
                log_callback(f"FALLBACK: {fallback_reason}. Using {scheme_info}")
            else:
                log_callback(f"MODE: {scheme_info}")

            # Step 1: Research
            success = ScriptOrchestrator.run_research_phase(app, pid, log_callback, assembled_prompt=detective_prompt)
            if not success:
                log_callback("Chain interrupted at fact gathering phase.")
                return

            # Step 2: Synthesis
            writer_model = app.ai_manager.get_model_string_for_task("script_writing")
            seed_data = SeedManager.load_seed(app.WORK_ROOT_PATH, pid) # Reload to get research facts
            
            if not seed_data:
                log_callback("Failed to load seed data for synthesis.")
                return
                
            songs = seed_data.get("songs", [])
            log_callback(f"[WRITER] Sending {len(songs)} datasets to {writer_model}...")

            log_callback("[WRITER] Request sent. Waiting for model response...")
            # Template assembly is now handled inside synthesize_script
            script_data = CompositionAgent.synthesize_script(app, seed_data, arch_data=arch_data)
            
            if not script_data or not script_data.get("parsed_blocks") or script_data.get("error") == "non_informative":
                 log_callback("[WRITER] Script was not synthesized. Excel sheet creation cancelled.")
                 return

            # Step 3: Export
            word_count = len(script_data.get("raw_script_output", "").split())
            log_callback(f"[WRITER] Text received ({word_count} words). Parsing blocks...")
            log_callback("Programmatic cutting and Excel sheet formation...")
            result_export = ExportManager.process_and_export(app, pid, script_data)

            if result_export is not False:
                log_callback(f"SUCCESS: Sheet VID{pid} created in database. Total rows written: {result_export}")
            else:
                log_callback("RED: Error during cutting or Excel writing. Check terminal for details.")

        except Exception as e:
            log_callback(f"Critical orchestrator error: {e}")
            traceback.print_exc()

    @staticmethod
    def run_research_phase(app, pid, log_callback, assembled_prompt=None):
        detective_model = app.ai_manager.get_model_string_for_task("fact_research")
        log_callback(f"[DETECTIVE] Starting fact gathering via {detective_model}...")
        
        seed_data = SeedManager.load_seed(app.WORK_ROOT_PATH, pid)
        if not seed_data:
            log_callback("Failed to load project data. Check PID.")
            return False

        version = seed_data.get("metadata", {}).get("version", "v1")
        songs = seed_data.get("songs", [])
        if not songs:
            log_callback("Song list is empty.")
            return False

        # Template-based standard (Default for v1 and others)
        prompt_path = Path(app.WORK_ROOT_PATH) / "common_assets" / "prompts" / f"script_research_detective_{version}_template.txt"
        if not prompt_path.exists():
             log_callback(f"CRITICAL ERROR: Research template {prompt_path.name} not found. Fallback forbidden.")
             return False

        chronology = seed_data.get("inputs", {}).get("chronology", "Unknown Period")
        perspective = seed_data.get("perspective", {}).get("selected", "General Analysis")

        total_songs = len(songs)
        completed_count = 0
        
        with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
            future_to_index = {
                executor.submit(
                    ResearchAgent.mine_song_facts, 
                    app, 
                    s['artist'], 
                    s['song'], 
                    chronology, 
                    perspective, 
                    str(prompt_path),
                    assembled_prompt=assembled_prompt
                ): i for i, s in enumerate(songs)
            }

            for future in concurrent.futures.as_completed(future_to_index):
                idx = future_to_index[future]
                song_item = songs[idx] 
                
                try:
                    facts = future.result()
                    
                    if not facts:
                        log_callback(f"Song {idx+1}: Empty response. Retrying...")
                        facts = ResearchAgent.mine_song_facts(
                            app, 
                            song_item['artist'], 
                            song_item['song'], 
                            chronology, 
                            perspective, 
                            str(prompt_path)
                        )

                    if facts:
                        with ScriptOrchestrator._file_lock:
                            current_seed = SeedManager.load_seed(app.WORK_ROOT_PATH, pid)
                            if current_seed and "songs" in current_seed:
                                if 0 <= idx < len(current_seed["songs"]):
                                    current_seed["songs"][idx]["research_data"] = facts
                                    if SeedManager.save_seed_json(app, current_seed):
                                        log_callback(f"[OK] Song {idx+1}: Found plot hooks for ({song_item['artist']} - {song_item['song']}).")
                                    else:
                                        log_callback(f"Song {idx+1}: Save error.")
                                else:
                                     log_callback(f"Song {idx+1}: Index out of range.")
                            else:
                                log_callback(f"Error re-reading file for song {idx+1}.")
                    else:
                        log_callback(f"Song {idx+1}: Failed to gather facts.")

                except Exception as e:
                    log_callback(f"Thread error for song {idx+1}: {e}")
                
                completed_count += 1

        log_callback(f"[DETECTIVE] Research completed.")
        return True
