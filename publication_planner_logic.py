# publication_planner_logic.py (v6.1 - Added pytz import)

import pandas as pd
from pathlib import Path
import re
from datetime import datetime, timedelta
import json
from typing import List, Dict, Any, Tuple
import openpyxl
import pytz  # <<< CHANGE: Added missing import


# --- Utility functions ---
def parse_base_id(filename: str) -> str:
    """Extracts the base ID (e.g., VID123) from a filename."""
    if not isinstance(filename, str): return ""
    match = re.match(r"(VID\d+)", filename, re.IGNORECASE)
    return match.group(1).upper() if match else ""


# --- Logic classes ---

class ChannelConfigManager:
    """Manages channel configuration from JSON file."""

    def __init__(self, config_path: Path):
        self.config_path = config_path
        self.channels = []
        self.load_config()

    def load_config(self):
        """Loads configuration or creates default."""
        try:
            with open(self.config_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                if not isinstance(data, list):
                    raise json.JSONDecodeError("Config is not a list", "", 0)
                self.channels = data
            if len(self.channels) < 10:
                existing_ids = {ch['id'] for ch in self.channels}
                for i in range(1, 11):
                    if i not in existing_ids:
                        self.channels.append(
                            {"id": i, "name": f"{i} Channel", "url": "", "proxy": "", "channel_id": ""})
                self.channels.sort(key=lambda x: x['id'])
                self.save_config()
        except (FileNotFoundError, json.JSONDecodeError):
            self.channels = [{"id": i, "name": f"{i} Channel", "url": "", "proxy": "", "channel_id": ""} for i in
                             range(1, 11)]
            self.save_config()

    def save_config(self):
        """Saves configuration to JSON."""
        try:
            self.channels.sort(key=lambda x: x['id'])
            with open(self.config_path, 'w', encoding='utf-8') as f:
                json.dump(self.channels, f, ensure_ascii=False, indent=4)
        except Exception as e:
            print(f"Error saving channel configuration file: {e}")

    def get_channel_names(self) -> List[str]:
        return [ch['name'] for ch in self.channels]

    def get_channels_data(self) -> List[Dict]:
        return self.channels

    def update_channels_data(self, new_data: List[Dict]):
        self.channels = new_data
        self.save_config()


class ScannedDataManager:
    """Works with cumulative database 'Posting_Scanned'."""

    def __init__(self, db_path: Path):
        self.db_path = db_path
        self.sheet_name = 'Posting_Scanned'
        self.columns = ['ChannelID', 'ChannelName', 'PublicationTimestamp', 'VideoTitle', 'VideoURL',
                        'OriginalFilename']
        self.local_tz = pytz.timezone('Europe/Warsaw')

    def ensure_sheet_exists(self):
        """Checks/creates Posting_Scanned sheet at first position."""
        try:
            workbook = openpyxl.load_workbook(self.db_path)
            if self.sheet_name not in workbook.sheetnames:
                new_sheet = workbook.create_sheet(self.sheet_name, 0)
                new_sheet.append(self.columns)
                workbook.save(self.db_path)
        except FileNotFoundError:
            df = pd.DataFrame(columns=self.columns)
            df.to_excel(self.db_path, sheet_name=self.sheet_name, index=False)
            workbook = openpyxl.load_workbook(self.db_path)
            if self.sheet_name in workbook.sheetnames:
                workbook.move_sheet(workbook[self.sheet_name], offset=-len(workbook.sheetnames) + 1)
                workbook.save(self.db_path)
        except Exception as e:
            print(f"Error in ensure_sheet_exists: {e}")

    def get_scanned_data(self) -> pd.DataFrame:
        """Reads all data from Posting_Scanned."""
        try:
            df = pd.read_excel(self.db_path, sheet_name=self.sheet_name)
            if 'PublicationTimestamp' in df.columns:
                df['PublicationTimestamp'] = pd.to_datetime(df['PublicationTimestamp'], errors='coerce')
            return df
        except Exception:
            return pd.DataFrame(columns=self.columns)

    def add_scanned_videos(self, videos_data: List[Dict]):
        """Adds videos cumulatively and safely overwrites the sheet."""
        if not videos_data: return
        df = self.get_scanned_data()
        existing_urls = set(df['VideoURL']) if 'VideoURL' in df.columns else set()
        new_records = []
        for v in videos_data:
            if v['url'] not in existing_urls:
                aware_time = v['published_time']
                local_time = aware_time.astimezone(self.local_tz)
                naive_local_time = local_time.replace(tzinfo=None)
                new_records.append({
                    'ChannelID': v['channel_id'],
                    'ChannelName': v['channel_name'],
                    'PublicationTimestamp': naive_local_time,
                    'VideoTitle': v['title'],
                    'VideoURL': v['url'],
                    'OriginalFilename': v.get('filename', '')
                })
        if new_records:
            new_df = pd.DataFrame(new_records)
            combined_df = pd.concat([df, new_df], ignore_index=True)
            self._safe_write_df_to_excel(combined_df)
            print(f"Added {len(new_records)} new records.")
        else:
            print("No new videos found to be added.")

    def update_filename(self, video_url: str, new_filename: str):
        """Updates filename and safely overwrites the sheet."""
        df = self.get_scanned_data()
        if 'VideoURL' not in df.columns: return
        mask = df['VideoURL'] == video_url
        if not mask.any(): return
        df.loc[mask, 'OriginalFilename'] = new_filename
        self._safe_write_df_to_excel(df)
        print(f"Filename for {video_url} updated.")

    def _safe_write_df_to_excel(self, df: pd.DataFrame):
        """
        Safely overwrites 'Posting_Scanned' sheet using _new and _old schema.
        """
        df.sort_values(by='PublicationTimestamp', ascending=False, inplace=True)
        try:
            workbook = openpyxl.load_workbook(self.db_path)
            temp_new_sheet_name = f"{self.sheet_name}_new"
            temp_old_sheet_name = f"{self.sheet_name}_old"

            if temp_new_sheet_name in workbook.sheetnames:
                workbook.remove(workbook[temp_new_sheet_name])
            sheet_new = workbook.create_sheet(temp_new_sheet_name)

            # Convert DataFrame to format suitable for openpyxl
            rows = [list(df.columns)] + df.values.tolist()
            for r in rows:
                sheet_new.append(r)

            sheet_new.freeze_panes = "A2"
            workbook.save(self.db_path)

            if self.sheet_name in workbook.sheetnames:
                if temp_old_sheet_name in workbook.sheetnames:
                    workbook.remove(workbook[temp_old_sheet_name])
                sheet_old = workbook[self.sheet_name]
                sheet_old.title = temp_old_sheet_name
            workbook.save(self.db_path)

            sheet_to_rename = workbook[temp_new_sheet_name]
            sheet_to_rename.title = self.sheet_name
            workbook.move_sheet(sheet_to_rename, offset=-len(workbook.sheetnames))
            workbook.save(self.db_path)

            if temp_old_sheet_name in workbook.sheetnames:
                workbook.remove(workbook[temp_old_sheet_name])
            workbook.save(self.db_path)

            print(f"Sheet '{self.sheet_name}' safely updated and sorted.")

        except Exception as e:
            print(f"Critical error during safe write to {self.sheet_name}: {e}")


class PublicationHistory:
    """Generates 'wide' view on the fly."""

    def __init__(self, scanned_data_manager: ScannedDataManager, channel_config_manager: ChannelConfigManager):
        self.scanned_manager = scanned_data_manager
        self.channel_manager = channel_config_manager
        self.history = []

    def generate_history_view(self):
        scanned_df = self.scanned_manager.get_scanned_data()
        if scanned_df.empty:
            self.history = [];
            return

        ready_df = scanned_df.copy()
        ready_df['DateOnly'] = ready_df['PublicationTimestamp'].dt.date

        grouped_by_date = ready_df.groupby('DateOnly')
        processed_history = []
        for date_obj, group in grouped_by_date:
            pubs_by_channel = group.groupby('ChannelID')
            max_pubs_on_date = max((len(channel_group) for _, channel_group in pubs_by_channel), default=0)

            for i in range(max_pubs_on_date):
                row_data = {'date': date_obj}
                for channel_info in self.channel_manager.get_channels_data():
                    channel_name = channel_info['name']
                    channel_id = channel_info['channel_id']

                    row_data[channel_name] = ""

                    if channel_id and channel_id in pubs_by_channel.groups:
                        channel_pubs = pubs_by_channel.get_group(channel_id).sort_values(
                            'PublicationTimestamp').to_dict('records')
                        if i < len(channel_pubs):
                            pub = channel_pubs[i]
                            # --- FIXED LOGIC ---
                            filename = pub.get('OriginalFilename')
                            if isinstance(filename, str) and filename.strip():
                                row_data[channel_name] = filename
                            else:
                                row_data[channel_name] = True
                            # --- END OF FIXED LOGIC ---

                processed_history.append(row_data)

        self.history = sorted(processed_history, key=lambda x: x['date'], reverse=True)

    def get_channel_names(self) -> List[str]:
        return self.channel_manager.get_channel_names()

    def find_publications_by_base_id(self, base_id: str) -> List[Dict[str, Any]]:
        scanned_df = self.scanned_manager.get_scanned_data()
        if scanned_df.empty or 'OriginalFilename' not in scanned_df.columns: return []

        scanned_df['base_id'] = scanned_df['OriginalFilename'].astype(str).apply(parse_base_id)
        results = scanned_df[scanned_df['base_id'] == base_id].to_dict('records')
        return [{'date': r['PublicationTimestamp'].date(), 'channel_name': r['ChannelName'],
                 'filename': r['OriginalFilename']} for r in results]


class PublicationPlanner:
    def __init__(self, history: PublicationHistory):
        self.history = history

    def analyze_files(self, new_filenames: List[str], min_gap: int, ideal_gap: int) -> List[Dict[str, Any]]:
        results = []
        today = datetime.now().date()
        for filename in new_filenames:
            base_id = parse_base_id(filename)
            if not base_id:
                results.append({'filename': filename, 'status': 'ID Error', 'details': 'Failed to determine ID.'})
                continue

            self.history.generate_history_view()
            related_pubs = self.history.find_publications_by_base_id(base_id)
            if not related_pubs:
                results.append({'filename': filename, 'status': 'Ready to publish',
                                'details': 'Versions of this video have not been published yet.', 'min_date': today,
                                'ideal_date': today})
            else:
                last_pub = max(related_pubs, key=lambda x: x['date'])
                last_pub_date = last_pub['date']
                min_date = last_pub_date + timedelta(days=min_gap)
                ideal_date = last_pub_date + timedelta(days=ideal_gap)
                status = 'Ready to publish' if today >= min_date else 'Postpone'
                results.append({'filename': filename, 'status': status,
                                'details': f"Last version was {last_pub_date.strftime('%d.%m.%Y')} on '{last_pub['channel_name']}'",
                                'min_date': min_date, 'ideal_date': ideal_date})
        return results