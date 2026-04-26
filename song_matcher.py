# song_matcher.py
"""
SongMatcher class for matching artists and songs in video captions.
Uses AI assistance for fuzzy matching and creates new entries when needed.
"""

import logging
import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any
from openpyxl import load_workbook
import unicodedata
from fuzzywuzzy import fuzz

class SongMatcher:
    """
    Main class for matching artists and songs from video captions.
    Handles loading, caching, exact matching, AI-assisted matching and canonical creation.
    """

    def __init__(self, database_path: str, ai_manager, prompts_path: Path):
        self.database_path = Path(database_path)
        self.ai_manager = ai_manager
        self.logger = logging.getLogger(__name__)
        self.prompts_path = Path(prompts_path)

        # Load prompt templates
        self.artist_matching_prompt = self._load_prompt_file("ArtistMatchingPrompt.txt", 
            "Find artist {artist_name} in candidates:\n{candidate_text}\nContext: {context}")
        self.song_matching_prompt = self._load_prompt_file("SongMatchingPrompt.txt", 
            "Find song {song_title} by {artist_name} in candidates:\n{candidate_text}\nContext: {context}")
        self.artist_canonical_prompt = self._load_prompt_file("ArtistCanonicalPrompt.txt", 
            "Determine canonical name and role for: {artist_name}\nContext: {context}")

        # Cache
        self.artists_by_name = {}  # normalized_name -> {eid, name}
        self.songs_by_eid = {}     # eid -> list of {ssid, song_title, normalized_title}
        self.video_map = {}        # (ssid, eid) -> video_id

        self._load_databases()

    def _load_prompt_file(self, filename: str, fallback: str) -> str:
        try:
            p = self.prompts_path / filename
            if p.exists():
                return p.read_text(encoding='utf-8')
            return fallback
        except Exception:
            return fallback

    def _load_databases(self):
        """Load and cache database sheets."""
        try:
            if not self.database_path.exists():
                self.logger.error(f"Database not found: {self.database_path}")
                return

            wb = load_workbook(self.database_path, data_only=True)

            # MasterDatabase_v1
            if 'MasterDatabase_v1' in wb.sheetnames:
                ws = wb['MasterDatabase_v1']
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] and row[1]:
                        eid, name = str(row[0]), str(row[1])
                        norm_name = self._normalize_text(name).lower()
                        self.artists_by_name[norm_name] = {'eid': eid, 'name': name}

            # Songs_Database
            if 'Songs_Database' in wb.sheetnames:
                ws = wb['Songs_Database']
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] and row[1] and row[3]:
                        ssid, title, eid = str(row[0]), str(row[1]), str(row[3])
                        if eid not in self.songs_by_eid:
                            self.songs_by_eid[eid] = []
                        self.songs_by_eid[eid].append({
                            'ssid': ssid,
                            'song_title': title,
                            'normalized_title': self._normalize_text(title)
                        })

            # EID_SSID_VIDEO-ID
            if 'EID_SSID_VIDEO-ID' in wb.sheetnames:
                ws = wb['EID_SSID_VIDEO-ID']
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] and row[1] and row[2]:
                        self.video_map[(str(row[0]), str(row[1]))] = str(row[2])

            wb.close()
        except Exception as e:
            self.logger.error(f"Error loading databases: {e}")

    def _normalize_text(self, text: str) -> str:
        """Standard normalization for lookup keys."""
        if not text: return ""
        text = unicodedata.normalize('NFKD', str(text))
        text = ''.join([c for c in text if not unicodedata.combining(c)])
        text = text.lower()
        text = re.sub(r'[^a-z0-9\s]', '', text)
        return ' '.join(text.split())

    def _clean_caption_part(self, text: str) -> str:
        """Aggressive cleaning of names/titles."""
        if not text: return ""
        text = str(text)
        # Remove numbering, quotes, ellipses
        text = re.sub(r'^(Number\s*)?\d+\.\s*', '', text).strip()
        text = re.sub(r'^"|"$', '', text).strip()
        text = re.sub(r'\.\.\.$', '', text).strip()
        # ABSOLUTE CLEAN: remove dots at the end
        return text.strip().rstrip('.').strip()

    def _get_significant_words(self, text: str) -> frozenset:
        """Extract words excluding common noise."""
        words = re.findall(r'\b[a-zA-Z0-9]+\b', str(text).lower())
        stop = {'the', 'a', 'an', 'and', '&', 'of', 'by', 'for', 'with', 'to'}
        return frozenset({w for w in words if w not in stop and len(w) > 1})

    def _prepare_for_matching(self, text: str) -> Tuple[Optional[str], Optional[str]]:
        """Splits incoming block title into Artist and Song."""
        if not text: return None, None
        
        # Cut at separator or newline
        cut_pos = min([text.find('>>') if '>>' in text else len(text), 
                       text.find('\n') if '\n' in text else len(text)])
        line = text[:cut_pos].strip()

        # Remove numbering, quotes, years
        line = re.sub(r'^\d+\.\s*', '', line).strip()
        line = re.sub(r'^"|"$', '', line).strip()
        line = re.sub(r'\s*\(\s*\d{4}\s*\)$', '', line).strip()
        line = re.sub(r'\s+\d{4}$', '', line).strip()

        # Split by the LAST occurrence of " by " to handle titles like "If You Don't Know Me by Now"
        # We find all occurrences and take the last one
        by_matches = list(re.finditer(r'\s+by\s+', line, flags=re.IGNORECASE))
        if by_matches:
            last_match = by_matches[-1]
            song = self._clean_caption_part(line[:last_match.start()])
            artist = self._clean_caption_part(line[last_match.end():])
            
            # Sentence cut heuristic for artist bio leak
            if ". " in artist:
                sub = artist.split(". ", 1)
                if sub[1] and sub[1][0].isupper():
                    artist = sub[0].strip()
            
            return artist, song
        
        # Fallback to simple Artist
        cleaned = self._clean_caption_part(line)
        return (cleaned, None) if cleaned and len(cleaned) < 50 else (None, None)

    def _get_canonical_data(self, artist_name: str) -> Tuple[str, str]:
        """Asks AI for official name and role."""
        try:
            prompt = self.artist_canonical_prompt.format(artist_name=artist_name, context=artist_name)
            res = self.ai_manager.execute_ai_task("text_processing", {"prompt": prompt}).get("text", "")
            
            name_match = re.search(r'CANONICAL_NAME:\s*(.+)', res, re.IGNORECASE)
            role_match = re.search(r'ROLE:\s*(\w+)', res, re.IGNORECASE)
            
            clean_name = self._clean_caption_part(name_match.group(1)) if name_match else self._clean_caption_part(artist_name)
            role = role_match.group(1).lower() if role_match else "artist"
            if role not in ['artist', 'group', 'vocalist', 'instrumentalist']:
                role = "artist"
            return clean_name, role
        except Exception:
            return self._clean_caption_part(artist_name), "artist"

    def match_artist_and_song(self, caption: str, full_text: str = "", video_id: str = None) -> Tuple[Optional[str], Optional[str], str, str]:
        """Main entry point for matching."""
        try:
            artist_name, song_title = self._prepare_for_matching(caption)
            if not artist_name:
                # Last resort AI match on raw caption
                eid, m_type = self._match_or_create_artist(caption[:100], full_text)
                return eid, None, m_type, 'parse_failed'

            eid, a_type = self._match_or_create_artist(artist_name, full_text)
            ssid, s_type = (None, 'none')
            if eid and song_title:
                ssid, s_type = self._match_or_create_song(eid, song_title, full_text)

            if video_id and eid and ssid and (a_type == 'created' or s_type == 'created'):
                self._write_video_mapping_to_excel(ssid, eid, video_id)
            
            return eid, ssid, a_type, s_type
        except Exception as e:
            self.logger.error(f"Matching error: {e}")
            return None, None, 'none', 'none'

    def _match_or_create_artist(self, artist_name: str, full_text: str = "") -> Tuple[Optional[str], str]:
        # 1. Exact / Normalized
        eid = self.find_artist_exact(artist_name)
        if eid: return eid, 'exact'

        # 2. AI Match from candidates
        candidates = list(self.artists_by_name.values())
        eid = self.find_artist_ai(artist_name, full_text, candidates)
        if eid: return eid, 'ai'

        # 3. Canonical Create (AI-Norm -> Exact Check -> Create)
        eid = self.create_new_artist(artist_name)
        return (eid, 'created') if eid else (None, 'none')

    def _match_or_create_song(self, eid: str, song_title: str, full_text: str = "") -> Tuple[Optional[str], str]:
        ssid = self.find_song_exact(eid, song_title)
        if ssid: return ssid, 'exact'

        # 2. Fuzzy match within this artist's songs
        ssid = self.find_song_fuzzy(eid, song_title)
        if ssid: return ssid, 'exact'

        # 3. AI Match from candidates
        if eid in self.songs_by_eid:
            ssid = self.find_song_ai(eid, song_title, full_text, self.songs_by_eid[eid])
            if ssid: return ssid, 'ai'

        ssid = self.create_new_song(eid, song_title)
        return (ssid, 'created') if ssid else (None, 'none')

    def find_song_fuzzy(self, eid: str, song_title: str, threshold: int = 85) -> Optional[str]:
        """Fuzzy comparison of normalized titles."""
        if eid not in self.songs_by_eid: return None
        target_norm = self._normalize_text(song_title)
        
        best_match = None
        highest_score = 0
        
        for s in self.songs_by_eid[eid]:
            score = fuzz.ratio(target_norm, s['normalized_title'])
            if score > highest_score:
                highest_score = score
                best_match = s['ssid']
        
        if highest_score >= threshold:
            self.logger.info(f"Fuzzy match: '{song_title}' score {highest_score}")
            return best_match
        return None

    def find_artist_exact(self, artist_name: str) -> Optional[str]:
        """Sequential search: Strict -> Variations -> Significant Words."""
        norm = self._normalize_text(artist_name)
        if norm in self.artists_by_name:
            return self.artists_by_name[norm]['eid']

        # Variations
        v1 = norm.replace(' and ', ' & ')
        v2 = norm.replace(' & ', ' and ')
        if v1 in self.artists_by_name: return self.artists_by_name[v1]['eid']
        if v2 in self.artists_by_name: return self.artists_by_name[v2]['eid']

        # Significant Words
        sig_target = self._get_significant_words(artist_name)
        if sig_target:
            for cached_norm, data in self.artists_by_name.items():
                if self._get_significant_words(data['name']) == sig_target:
                    return data['eid']

        self.logger.warning(f"DIAGNOSTIC: FAILED - No match for: '{artist_name}'")
        return None

    def find_artist_ai(self, artist_name: str, full_text: str, candidates: List[Dict]) -> Optional[str]:
        if not candidates: return None
        
        # 1. Filter candidates by a low fuzzy threshold to find "potential suspects"
        scored_candidates = []
        target_name = artist_name.lower()
        for c in candidates:
            # token_set_ratio is excellent for handling typos and small variations
            score = fuzz.token_set_ratio(target_name, c['name'].lower())
            if score >= 45: # Low, safe threshold to catch minimally relevant artists
                scored_candidates.append((score, c))
        
        # 2. Sort by score descending (most relevant first)
        scored_candidates.sort(key=lambda x: x[0], reverse=True)
        
        # 3. Take top 100 for AI review
        filtered_candidates = [item[1] for item in scored_candidates[:100]]
        
        if not filtered_candidates:
            self.logger.warning(f"AI Match: No candidates passed the low threshold for '{artist_name}'")
            return None

        cand_text = "\n".join([f"{i+1}. {c['name']}" for i, c in enumerate(filtered_candidates)])
        prompt = self.artist_matching_prompt.format(artist_name=artist_name, candidate_text=cand_text, context=full_text[:500])
        
        try:
            res = self.ai_manager.execute_ai_task("text_processing", {"prompt": prompt}).get("text", "").strip()
            match = re.search(r'^(\d+)', res)
            if match:
                idx = int(match.group(1)) - 1
                if 0 <= idx < len(filtered_candidates):
                    return filtered_candidates[idx]['eid']
        except Exception as e:
            self.logger.error(f"AI Artist Match error: {e}")
        return None

    def find_song_exact(self, eid: str, song_title: str) -> Optional[str]:
        if eid not in self.songs_by_eid: return None
        norm = self._normalize_text(song_title)
        for s in self.songs_by_eid[eid]:
            if s['normalized_title'] == norm: return s['ssid']
        return None

    def find_song_ai(self, eid: str, song_title: str, full_text: str, songs: List[Dict]) -> Optional[str]:
        if not songs: return None
        artist = self.get_artist_name_by_eid(eid)
        cand_text = "\n".join([f"{i+1}. {s['song_title']}" for i, s in enumerate(songs)])
        prompt = self.song_matching_prompt.format(song_title=song_title, artist_name=artist, candidate_text=cand_text, context=full_text[:500])
        try:
            res = self.ai_manager.execute_ai_task("text_processing", {"prompt": prompt}).get("text", "").strip()
            match = re.search(r'^(\d+)', res)
            if match:
                idx = int(match.group(1)) - 1
                if 0 <= idx < len(songs):
                    return songs[idx]['ssid']
        except Exception: pass
        return None

    def create_new_artist(self, artist_name: str) -> Optional[str]:
        """Normalizes name via AI, checks for duplicates again, then creates."""
        canonical_name, role = self._get_canonical_data(artist_name)
        
        # FINAL GUARD against duplicates
        existing_eid = self.find_artist_exact(canonical_name)
        if existing_eid:
            self.logger.info(f"AI normalized '{artist_name}' -> '{canonical_name}' (MATCH FOUND: {existing_eid})")
            return existing_eid

        if len(canonical_name) > 50:
            self.logger.error(f"Artist name too long ({len(canonical_name)}): {canonical_name}")
            return None

        try:
            wb = load_workbook(self.database_path)
            ws = wb['MasterDatabase_v1']
            max_eid = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                try:
                    num = int(str(row[0])[3:])
                    if num > max_eid: max_eid = num
                except:
                    pass
            new_eid = f"EID{max_eid + 1:04d}"
            ws.append([new_eid, canonical_name, role, 1])
            wb.save(self.database_path)
            self.artists_by_name[self._normalize_text(canonical_name)] = {'eid': new_eid, 'name': canonical_name}
            return new_eid
        except Exception as e:
            self.logger.error(f"Error creating artist: {e}")
            return None

    def create_new_song(self, eid: str, song_title: str) -> Optional[str]:
        title = self._clean_caption_part(song_title)
        if len(title) > 60: return None
        try:
            wb = load_workbook(self.database_path)
            ws = wb['Songs_Database']
            max_ssid = 0
            for r in ws.iter_rows(min_row=2, values_only=True):
                try:
                    num = int(str(r[0])[4:])
                    if num > max_ssid: max_ssid = num
                except:
                    pass
            new_ssid = f"SSID{max_ssid + 1:04d}"
            artist_name = self.get_artist_name_by_eid(eid)
            ws.append([new_ssid, title, None, eid, artist_name])
            wb.save(self.database_path)
            if eid not in self.songs_by_eid: self.songs_by_eid[eid] = []
            self.songs_by_eid[eid].append({'ssid': new_ssid, 'song_title': title, 'normalized_title': self._normalize_text(title)})
            return new_ssid
        except Exception as e:
            self.logger.error(f"Error creating song: {e}")
            return None

    def _write_video_mapping_to_excel(self, ssid: str, eid: str, video_id: str):
        try:
            wb = load_workbook(self.database_path)
            ws = wb['EID_SSID_VIDEO-ID'] if 'EID_SSID_VIDEO-ID' in wb.sheetnames else wb.create_sheet('EID_SSID_VIDEO-ID')
            if ws.max_row == 1 and not ws['A1'].value: ws.append(['SSID', 'EID', 'Publication_Video_ID'])
            ws.append([ssid, eid, video_id])
            wb.save(self.database_path)
        except Exception: pass

    def get_artist_name_by_eid(self, eid: str) -> str:
        for data in self.artists_by_name.values():
            if data['eid'] == eid: return data['name']
        return "Unknown"

    def get_song_title_by_ssid(self, ssid: str) -> str:
        for songs in self.songs_by_eid.values():
            for s in songs:
                if s['ssid'] == ssid: return s['song_title']
        return ""

    def get_statistics(self) -> Dict[str, Any]:
        return {'total_artists': len(self.artists_by_name), 
                'total_songs': sum(len(s) for s in self.songs_by_eid.values()),
                'artists_with_songs': len(self.songs_by_eid)}
