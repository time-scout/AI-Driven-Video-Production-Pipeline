# --- text_splitter_worker.py ---

import re
import json
from pathlib import Path
from typing import Callable, Optional, List, Dict
import traceback
import openpyxl  # <--- FIX: Added import for Excel operations

try:
    import tomli as toml
except ImportError:
    import toml

# Google AI is now used via AI_Manager
from song_matcher import SongMatcher

try:
    import tiktoken
except ImportError:
    tiktoken = None


# --- Helper functions ---

def _split_into_sentences(text: str) -> List[str]:
    # List of abbreviations after which a dot does not mean the end of a sentence
    abbreviations = r"(?:Mr|Mrs|Ms|Dr|St|Jr|Sr|vs|approx|incl|etc|Jan|Feb|Mar|Apr|Jun|Jul|Aug|Sep|Oct|Nov|Dec)"
    # Regular expression: look for dot/question/exclamation followed by space and uppercase letter,
    # provided the preceding part is NOT one of our abbreviations.
    pattern = rf'(?<!{abbreviations})(?<=[.!?])\s+(?=[A-Z"“])'
    sentences = re.split(pattern, text.replace('\n', ' ').strip())
    return [s.strip() for s in sentences if s.strip()]


def _count_tokens(text: str, log_callback: Callable) -> int:
# ... (code below unchanged until run_splitting_process)
    if not tiktoken:
        log_callback("WARNING: 'tiktoken' not installed. Using approximate count.")
        return len(text) // 4
    try:
        encoding = tiktoken.get_encoding("cl100k_base")
        return len(encoding.encode(text))
    except Exception as e:
        log_callback(f"ERROR counting tokens: {e}", is_error=True)
        return len(text) // 4


def _read_text_file_safe(file_path: Path) -> str:
    """
    Safe reading of text file with various encoding support.
    """
    encodings_to_try = ['utf-8', 'cp1251', 'windows-1251', 'iso-8859-1', 'cp1252', 'latin1']

    for encoding in encodings_to_try:
        try:
            return file_path.read_text(encoding=encoding)
        except UnicodeDecodeError:
            continue

    # If nothing helped, try reading as binary and decode with errors ignored
    try:
        return file_path.read_bytes().decode('utf-8', errors='ignore')
    except Exception:
        return ""


def _extract_title_for_matching(processed_text: str, fallback_caption: str) -> str:
    """
    Smart extraction of a song title line from the processed text.

    Looks for title patterns in text:
    1. Quoted lines: "Song Title" by Artist
    2. "Song by Artist" patterns
    3. "Song - Artist" or "Song – Artist" patterns
    4. Lines with a year in parentheses (1974)
    5. First line if it looks like a title (short and contains "by")

    Args:
        processed_text: Processed text from AI
        fallback_caption: Original caption as a backup

    Returns:
        String with title for recognition
    """
    # PRIORITY 1: If we have a high-quality Caption from AI (format: "Song" by Artist),
    # we trust it 100% and don't try to guess the title from the text.
    # This prevents errors where the first sentence of a story is mistaken for a title.
    if fallback_caption and '"' in fallback_caption and " by " in fallback_caption:
        return fallback_caption

    if not processed_text:
        return fallback_caption or ""

    lines = processed_text.split('\n')

    # Pattern 1: Look for lines with quotes (most reliable indicator)
    for line in lines:
        line = line.strip()
        # Match: "Song Title" by Artist (1974)
        # Use greedy match for the title inside quotes to handle nested quotes if any,
        # but the key is how we handle "by".
        quote_match = re.match(r'^"(.+)"\s+by\s+([^>>.]+)', line, re.IGNORECASE)
        if quote_match and len(line) < 100:
            return line

    # Pattern 2: Look for "Song by Artist" pattern without quotes
    for line in lines:
        line = line.strip()
        # Find all " by " occurrences and try to treat the last one as the separator
        by_matches = list(re.finditer(r'\s+by\s+', line, flags=re.IGNORECASE))
        if by_matches:
            last_match = by_matches[-1]
            potential_title = line[:last_match.start()].strip()
            # If the line isn't too long, it's likely a title line
            if len(line) < 100:
                return line

    # Pattern 3: Look for lines with year in parentheses
    for line in lines:
        line = line.strip()
        # Match: Song Title (1974)
        year_match = re.search(r'\(\s*\d{4}\s*\)', line)
        cut_line = line.split('>>')[0] if '>>' in line else line
        if year_match and len(cut_line) < 100:
            return cut_line

    # Pattern 4: Look for "Song - Artist" pattern
    for line in lines:
        line = line.strip()
        if re.search(r'\s*[-–—]\s+', line):
            parts = re.split(r'\s*[-–—]\s+', line, maxsplit=1)
            cut_line = line.split('>>')[0] if '>>' in line else line
            if len(parts) == 2 and len(cut_line) < 100:
                return cut_line

    # Pattern 5: First line if it's short and contains "by"
    if lines:
        first_line = lines[0].strip()
        cut_first = first_line.split('>>')[0] if '>>' in first_line else first_line
        if (len(cut_first) < 100 and
            re.search(r'\s+by\s+', cut_first, re.IGNORECASE)):
            return cut_first

    # Pattern 6: Look for any line that looks like a title (short + artist info)
    for line in lines:
        line = line.strip()
        cut_line = line.split('>>')[0] if '>>' in line else line
        # Skip very long lines, lines with URLs, or lines that look like descriptions
        if (20 < len(cut_line) < 100 and
            not cut_line.lower().startswith(('http', 'www', 'this', 'the song', 'track')) and
            (re.search(r'\s+by\s+|\([^)]*\)|"', cut_line, re.IGNORECASE) or
             any(word in cut_line.lower() for word in ['song', 'track', 'music', 'record']))):
            return cut_line

    # Fallback: return the original caption
    return fallback_caption or ""


def _call_ai_manager(prompt: str, log_callback: Callable, ai_manager) -> Optional[str]:
    """AI call via AI_Manager using configured model for 'text_processing'"""
    try:
        response = ai_manager.execute_ai_task(
            task_category="text_processing",
            input_data={"prompt": prompt}
        )

        if response.get("error"):
            log_callback(f"CRITICAL ERROR during AI call: {response['error']}", is_error=True)
            return None

        # AI_Manager doesn't return token stats, so this block is commented out
        # log_callback(f"Token stats: Input={usage.prompt_token_count}, Output={usage.candidates_token_count}")

        return response.get("text")
    except Exception as e:
        error_details = f"CRITICAL ERROR during AI call: {e}\n{traceback.format_exc()}"
        log_callback(error_details, is_error=True)
        return None


# --- Pipeline Stages ---

def run_stage1_mapper(raw_text: str, prompt_path: Path, status_callback: Callable,
                      log_callback: Callable, ai_manager) -> Optional[List[Dict]]:
    status_callback("Stage 1: Text correction and block creation...")
    log_callback("Starting Stage 1 (Mapper)...")

    prompt_template = _read_text_file_safe(prompt_path)
    if not prompt_template:
        log_callback(f"ERROR: Failed to read prompt file {prompt_path}", is_error=True)
        return None
    full_prompt = f"{prompt_template}\n\n{raw_text}"

    response_text = _call_ai_manager(full_prompt, log_callback, ai_manager)
    if not response_text:
        return None

    try:
        # First look for JSON in markdown blocks
        json_match = re.search(r'```json\s*(\{.*?\})\s*```', response_text, re.DOTALL)
        if json_match:
            response_text = json_match.group(1)
        else:
            # If not found, look for JSON object in entire response
            # Find start { and end } of JSON
            start_idx = response_text.find('{')
            if start_idx != -1:
                # Find matching closing brace
                brace_count = 0
                for i in range(start_idx, len(response_text)):
                    if response_text[i] == '{':
                        brace_count += 1
                    elif response_text[i] == '}':
                        brace_count -= 1
                        if brace_count == 0:
                            response_text = response_text[start_idx:i+1]
                            break

        # Clean JSON from common AI errors
        response_text = re.sub(r'\]\s*,\s*\}', ']}', response_text)  # Remove trailing comma before }
        response_text = re.sub(r'" ,', '",', response_text)           # Remove space before comma
        response_text = re.sub(r'\.\s*"', '. "', response_text)        # Fix dots inside strings
        response_text = re.sub(r'\.\s*([A-Z])', r'. \1', response_text) # Dots before uppercase

        data = json.loads(response_text)
        blocks = data.get("blocks", [])

        if not blocks:
            log_callback("ERROR: AI did not return blocks.", is_error=True)
            return None

        log_callback(f"Received {len(blocks)} blocks from AI.")

        # Clean text fields from HTML encoding
        cleaned_blocks = []
        for block in blocks:
            cleaned_block = {
                "caption": block.get("caption", "").replace("&amp;", "&"),
                "text": block.get("text", "").replace("&amp;", "&")
            }
            cleaned_blocks.append(cleaned_block)

        # Simple validation: check that blocks are not empty
        for i, block in enumerate(cleaned_blocks):
            if not block.get("text", "").strip():
                log_callback(f"VALIDATION ERROR: Block {i + 1} has empty text.", is_error=True)
                return None
            log_callback(f"- Block {i + 1} ('{block.get('caption', 'N/A')}'): {len(block['text'])} characters")

        log_callback("Block validation successful.")
        return cleaned_blocks

    except json.JSONDecodeError:
        log_callback(f"ERROR: AI returned invalid JSON at Stage 1.\nResponse:\n{response_text}", is_error=True)
        return None


def run_stage2_corrector(blocks_map: List[Dict], sentences: List[str], prompt_path: Path,
                         status_callback: Callable, log_callback: Callable, ai_manager) -> List[Dict]:
    status_callback("Stage 2: Block correction...")
    log_callback("Starting Stage 2 (Corrector)...")

    corrected_blocks = []
    prompt_template = _read_text_file_safe(prompt_path)
    if not prompt_template:
        log_callback(f"ERROR: Failed to read corrector prompt file {prompt_path}", is_error=True)
        return []
    total_blocks = len(blocks_map)

    for i, block_info in enumerate(blocks_map):
        status_callback(f"Stage 2: Processing block {i + 1}/{total_blocks}...")
        block_text = block_info.get('text', '')

        if not block_text:
            log_callback(f"ERROR: Block {i + 1} has no text.", is_error=True)
            continue

        token_count = _count_tokens(block_text, log_callback)
        if token_count > 7000:
            log_callback(
                f"WARNING: Block {i + 1} ({token_count} tokens) is too large. Response limit errors possible.",
                is_error=True)

        full_prompt = f"{prompt_template}\n{block_text}"
        corrected_text = _call_ai_manager(full_prompt, log_callback, ai_manager)

        if corrected_text is None:
            log_callback(f"ERROR: Failed to process block {i + 1}. Using source text.", is_error=True)
            corrected_text = block_text
        else:
            # Clean text from HTML encoding
            corrected_text = corrected_text.replace("&amp;", "&")

        block_info['source_text'] = block_text
        block_info['processed_text'] = corrected_text
        corrected_blocks.append(block_info)

    return corrected_blocks


# --- Main Orchestrator ---
def run_splitting_process(settings: dict, status_callback: Callable, log_callback: Callable, ai_manager) -> dict:
    if not tiktoken:
        msg = "'tiktoken' library not installed."
        log_callback(msg, is_error=True)
        return {'status': 'error', 'message': msg}

    pid = settings['pid']
    root_path = Path(settings['projects_root_path'])
    mapper_prompt_path = settings['mapper_prompt_path']
    corrector_prompt_path = settings['corrector_prompt_path']
    transcript_path = root_path / "parsed_data" / f"{pid}_transcript.txt"

    # API key is now managed via AI_Manager

    if not transcript_path.exists():
        msg = f'Transcript file not found: {transcript_path}'
        log_callback(msg, is_error=True)
        return {'status': 'error', 'message': msg}

    source_text = _read_text_file_safe(transcript_path)
    if not source_text:
        msg = f'Failed to read transcript file (encoding issue): {transcript_path}'
        log_callback(msg, is_error=True)
        return {'status': 'error', 'message': msg}

    # Data type detection
    if '[BLOCK_START:' in source_text:
        log_callback("Data type: Author script (with tags). Using programmatic block splitting.")
        # Step 1: Programmatic block extraction (ensures B01 accuracy)
        block_pattern = r'\[BLOCK_START:\s*(B(\d+))\]\s*(.*?)\s*\[BLOCK_END\]'
        tags = re.findall(block_pattern, source_text, re.DOTALL)
        
        # Sort by numerical ID
        sorted_tags = sorted(tags, key=lambda x: int(x[1]))
        
        # 1. Count total songs (excluding intro B01)
        total_songs = len([t for t in sorted_tags if t[0] != 'B01'])
        song_counter = 0

        blocks_map = []
        total_author_blocks = len(sorted_tags)
        
        for i, (full_id, num_id, content) in enumerate(sorted_tags):
            status_callback(f"Generating titles: {i+1}/{total_author_blocks}...")
            clean_content = content.strip()
            
            if full_id == 'B01':
                blocks_map.append({'caption': 'Intro', 'text': clean_content})
            else:
                # 2. Calculate number for countdown
                display_num = total_songs - song_counter
                song_counter += 1

                # In order for the AI Mapper not to call the block "Intro", we create context from two blocks.
                # 3. Substitute display_num instead of num_id in context for AI
                synthetic_context = f"Intro\nThis is a placeholder.\n\nNumber {display_num}, \"PLACEHOLDER\" by SYSTEM.\n{clean_content}"
                
                # Call original mapper on a small piece
                ai_results = run_stage1_mapper(synthetic_context, mapper_prompt_path, 
                                              lambda m: None, lambda m, e=False: None, ai_manager)
                
                # We need the header of the SECOND block from the AI response
                caption = ai_results[1]['caption'] if ai_results and len(ai_results) > 1 else ai_results[0]['caption'] if ai_results else f"{display_num}."
                blocks_map.append({
                    'caption': caption,
                    'text': clean_content  # Your ORIGINAL text
                })
    else:
        log_callback("Data type: Standard transcript. Using original AI Mapper.")
        blocks_map = run_stage1_mapper(source_text, mapper_prompt_path, status_callback, log_callback, ai_manager)

    if not blocks_map:
        return {'status': 'error', 'message': 'Stage 1 (Mapper) did not return result or failed validation.'}

    final_blocks_data = run_stage2_corrector(blocks_map, [], corrector_prompt_path, status_callback,
                                             log_callback, ai_manager)

    log_callback("Processing completed successfully.")
    return {'status': 'success', 'final_blocks': final_blocks_data}


# --- Excel Save Function ---
def save_results_to_excel(settings: dict, final_blocks: List[Dict], status_callback: Callable,
                          log_callback: Callable) -> bool:
    pid = settings['pid']
    root_path = Path(settings['projects_root_path'])
    db_path = root_path / "database" / "main_database.xlsx"
    json_path = root_path / "parsed_data" / f"{pid}_transcript.json"

    try:
        status_callback("Preparing to save to Excel...")
        log_callback(f"Starting save for {pid} in {db_path}")
        log_callback(f"DEBUG: final_blocks contains {len(final_blocks)} blocks")

        for idx, blk in enumerate(final_blocks):
            proc_text = blk.get('processed_text', '')
            cap = blk.get('caption', '')
            log_callback(f"DEBUG: Block {idx+1}: caption='{cap[:50]}...', processed_text length={len(proc_text)}")
            if len(proc_text) > 0:
                log_callback(f"DEBUG: processed_text preview: {proc_text[:100].replace(chr(10), ' ')}...")

        url = ""
        if json_path.exists():
            with open(json_path, 'r', encoding='utf-8') as f:
                url = json.load(f).get('url', "")

        vid_number_match = re.search(r'\d+', pid)
        if not vid_number_match:
            log_callback(f"ERROR: Failed to extract number from PID: {pid}", is_error=True);
            return False
        sheet_name = f"VID{vid_number_match.group()}s"

        workbook = openpyxl.load_workbook(db_path) if db_path.exists() else openpyxl.Workbook()

        if sheet_name in workbook.sheetnames:
            log_callback(f"Deleting existing sheet '{sheet_name}'...")
            del workbook[sheet_name]

        sheet = workbook.create_sheet(title=sheet_name, index=0)
        log_callback(f"Created new sheet '{sheet_name}'.")

        # Delete Sheet1 if there are other sheets
        if 'Sheet1' in workbook.sheetnames and len(workbook.sheetnames) > 1:
            workbook.remove(workbook['Sheet1'])
            log_callback("Deleted sheet 'Sheet1'.")

        sheet.append(["Block ID", "Text", "SSID", "EID", "Clean Caption", "", "", "", "", "Split Info", "Caption"])
        sheet.append(["", "", "", "", "", "url", url])

        bold_font = openpyxl.styles.Font(bold=True)
        wrap_alignment = openpyxl.styles.Alignment(wrap_text=False, horizontal='left')

        row_idx = 3
        log_callback(f"DEBUG: Starting writing {len(final_blocks)} blocks to sheet")
        for i, block_data in enumerate(final_blocks):
            sheet[f'A{row_idx}'] = f"B{i + 1:02d}"
            sheet[f'C{row_idx}'] = block_data.get('ssid', '')
            sheet[f'D{row_idx}'] = block_data.get('eid', '')
            sheet[f'E{row_idx}'] = block_data.get('clean_caption', '')
            sheet[f'K{row_idx}'] = block_data.get('caption', '')
            sheet[f'K{row_idx}'].font = bold_font

            text_cell = sheet[f'B{row_idx}']
            # Use the 'processed_text' key that we standardized
            text_cell.value = block_data.get('processed_text', '')
            text_cell.alignment = wrap_alignment
            log_callback(f"DEBUG: Wrote block {i+1} to row {row_idx} (SSID={block_data.get('ssid', '')}, EID={block_data.get('eid', '')})")
            row_idx += 1

        log_callback(f"DEBUG: Sheet '{sheet_name}' contains {sheet.max_row} rows and {sheet.max_column} columns")

        # Sort sheets by number
        def sort_key(sheet):
            match = re.search(r'\d+', sheet.title)
            return int(match.group()) if match else 0
        workbook._sheets.sort(key=sort_key)
        log_callback("Sheets sorted by number.")

        status_callback("Saving Excel file...")
        workbook.save(db_path)
        log_callback(f"SUCCESS: Data for {pid} successfully saved to Excel.")
        # Check file size
        import os
        if os.path.exists(db_path):
            file_size = os.path.getsize(db_path)
            log_callback(f"DEBUG: File size after saving: {file_size} bytes")
        status_callback("Saved!")
        return True

    except Exception as e:
        log_callback(f"CRITICAL ERROR while saving to Excel: {traceback.format_exc()}", is_error=True)
        status_callback("Save error!")
        return False


def create_clean_caption(original_caption: str, eid: str, ssid: str, matcher: SongMatcher) -> str:
    """Creates a clean caption based on database data"""
    # Keep original caption for Intro blocks
    if original_caption.lower() == "intro":
        return original_caption
    
    # Extract number from original caption
    match = re.match(r'^(\d+)\.', original_caption)
    number = match.group(1) + "." if match else ""

    # If no EID or SSID, return original
    if not eid or not ssid or eid.lower() == 'nan' or ssid.lower() == 'nan':
        return original_caption

    # Get clean names
    artist_name = matcher.get_artist_name_by_eid(eid)
    song_title = matcher.get_song_title_by_ssid(ssid)

    # If data not found, return original
    if not artist_name or not song_title:
        return original_caption

    # Form clean caption
    return f'{number} "{song_title}" by {artist_name}'


# --- Database Search Function ---
def run_database_matching(settings: dict, final_blocks: List[Dict], status_callback: Callable,
                         log_callback: Callable, ai_manager) -> dict:
    """
    Run database matching for processed blocks.

    Args:
        settings: Dictionary containing 'pid' and 'projects_root_path'
        final_blocks: List of processed blocks with caption and text
        status_callback: Function to update status
        log_callback: Function to log messages
        ai_manager: AI manager instance for AI-assisted matching

    Returns:
        Dictionary with status, updated blocks, and statistics
    """
    # Import SongMatcher locally to avoid circular dependencies
    from song_matcher import SongMatcher

    pid = settings['pid']
    root_path = Path(settings['projects_root_path'])
    db_path = root_path / "database" / "main_database.xlsx"
    json_path = root_path / "parsed_data" / f"{pid}_transcript.json"

    try:
        status_callback("Initializing database search...")
        log_callback(f"Starting database search for {pid}")

        # Load YouTube ID from JSON
        video_id = None
        if json_path.exists():
            with open(json_path, 'r', encoding='utf-8') as f:
                video_id = json.load(f).get('video_id', None)
                if video_id:
                    log_callback(f"Loaded video_id: {video_id}")

        # Construct prompts path (root_path is already WORK_ROOT)
        prompts_path = root_path / "common_assets" / "prompts"

        # Initialize SongMatcher with prompts path
        matcher = SongMatcher(str(db_path), ai_manager, prompts_path)

        # Statistics
        blocks_matched = 0
        blocks_total = 0
        stats = {'artist': {'exact': 0, 'ai': 0, 'created': 0},
                 'song': {'exact': 0, 'ai': 0, 'created': 0}}

        # Process each block
        for block in final_blocks:
            original_caption = block.get('caption', '').strip()
            processed_text = block.get('processed_text', '')

            # Skip intro blocks
            if not original_caption or original_caption.lower() == 'intro':
                log_callback(f"Skipping block with title '{original_caption}' (considered intro).")
                block['eid'] = ''
                block['ssid'] = ''
                continue
            
            # Smart extraction of title line for matching
            title_line_for_matching = _extract_title_for_matching(processed_text, original_caption)

            blocks_total += 1

            # Match artist and song
            eid, ssid, artist_match_type, song_match_type = matcher.match_artist_and_song(
                title_line_for_matching, processed_text, video_id
            )

            # Add EID and SSID to block
            block['eid'] = eid if eid else ''
            block['ssid'] = ssid if ssid else ''

            if eid or ssid:
                blocks_matched += 1

            # Log detailed results
            artist_log = "Not found"
            if eid:
                if artist_match_type == 'exact':
                    artist_log = f"Found exact EID: {eid}"
                    stats['artist']['exact'] += 1
                elif artist_match_type == 'ai':
                    artist_log = f"Found EID via AI: {eid}"
                    stats['artist']['ai'] += 1
                elif artist_match_type == 'created':
                    artist_log = f"CREATED new EID: {eid}"
                    stats['artist']['created'] += 1

            song_log = "Not found"
            if ssid:
                if song_match_type == 'exact':
                    song_log = f"Found exact SSID: {ssid}"
                    stats['song']['exact'] += 1
                elif song_match_type == 'ai':
                    song_log = f"Found SSID via AI: {ssid}"
                    stats['song']['ai'] += 1
                elif song_match_type == 'created':
                    song_log = f"CREATED new SSID: {ssid}"
                    stats['song']['created'] += 1

            log_callback(f"Block '{original_caption[:50]}...':\n  - Artist: {artist_log}\n  - Song: {song_log}")

        # Create Clean Caption for all blocks with identified EID/SSID
        log_callback("Creating Clean Caption based on database data...")
        log_callback(f"DEBUG: Total blocks to process: {len(final_blocks)}")
        for block in final_blocks:
            clean_caption = create_clean_caption(
                block.get('caption', ''),
                block.get('eid', ''),
                block.get('ssid', ''),
                matcher
            )
            block['clean_caption'] = clean_caption

        log_callback(f"Search completed: found matches for {blocks_matched} out of {blocks_total} blocks.")

        return {
            'status': 'success',
            'final_blocks': final_blocks,
            'blocks_matched': blocks_matched,
            'blocks_total': blocks_total,
            'stats': stats
        }

    except Exception as e:
        error_msg = f"Error during database search: {str(e)}\n{traceback.format_exc()}"
        log_callback(error_msg, is_error=True)
        return {
            'status': 'error',
            'message': error_msg,
            'final_blocks': final_blocks,
            'blocks_matched': 0,
            'blocks_total': 0
        }


